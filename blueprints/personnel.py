#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
人员管理模块
负责员工信息管理、导入导出等功能
"""
import json
import os
import pymysql
from collections import Counter
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session, current_app
from openpyxl import Workbook, load_workbook

from config.settings import APP_TITLE, EXPORT_DIR
from models.database import get_db, get_year_month_concat
from .decorators import login_required, manager_required
from .helpers import (
    current_user_id, require_user_id, get_accessible_department_ids,
    get_accessible_departments, calculate_years_from_date, get_user_department,
    validate_employee_access, log_import_operation
)
from blueprints.safety import extract_score_from_assessment

# 创建 Blueprint
personnel_bp = Blueprint('personnel', __name__, url_prefix='/personnel')


# ==================== 常量定义 ====================

PERSONNEL_FIELD_SCHEME = [
    {"name": "emp_no", "label": "工号", "input_type": "text", "required": True},
    {"name": "name", "label": "姓名", "input_type": "text", "required": True},
    {"name": "department_id", "label": "所属部门", "input_type": "department_select", "required": True},
    {"name": "class_name", "label": "班级", "input_type": "text"},
    {"name": "position", "label": "岗位", "input_type": "text"},
    {"name": "birth_date", "label": "出生年月", "input_type": "date"},
    {"name": "certification_date", "label": "取证时间", "input_type": "date"},
    {"name": "solo_driving_date", "label": "单独驾驶时间", "input_type": "date"},
    {"name": "marital_status", "label": "婚姻状况", "input_type": "select"},
    {"name": "hometown", "label": "籍贯", "input_type": "text"},
    {"name": "political_status", "label": "政治面貌", "input_type": "select"},
    {"name": "education", "label": "学历", "input_type": "select"},
    {"name": "graduation_school", "label": "毕业院校", "input_type": "text"},
    {"name": "work_start_date", "label": "参加工作时间", "input_type": "date"},
    {"name": "entry_date", "label": "入司时间", "input_type": "date"},
    {"name": "specialty", "label": "特长及兴趣爱好", "input_type": "textarea"},
]

PERSONNEL_DB_COLUMNS = [
    field["name"] for field in PERSONNEL_FIELD_SCHEME if field["name"] not in {"emp_no", "name"}
]

PERSONNEL_DATE_FIELDS = {"birth_date", "work_start_date", "entry_date", "certification_date", "solo_driving_date"}

PERSONNEL_SELECT_OPTIONS = {
    "marital_status": ["未婚", "已婚", "离异", "其它"],
    "political_status": ["中共党员", "中共预备党员", "共青团员", "群众", "其它"],
    "education": ["博士研究生", "硕士研究生", "本科", "大专", "中专", "高中", "其它"],
}

PERSONNEL_IMPORT_HEADER_MAP = {
    "工号": "emp_no",
    "姓名": "name",
    "所属部门": "department_id",
    "部门": "department_id",
    "班级": "class_name",
    "岗位": "position",
    "出生年月": "birth_date",
    "取证时间": "certification_date",
    "取证日期": "certification_date",
    "单独驾驶时间": "solo_driving_date",
    "单独驾驶日期": "solo_driving_date",
    "婚否": "marital_status",
    "婚姻状况": "marital_status",
    "籍贯": "hometown",
    "政治面貌": "political_status",
    "特长及兴趣爱好": "specialty",
    "特长": "specialty",
    "学历": "education",
    "毕业院校": "graduation_school",
    "参加工作时间": "work_start_date",
    "入司时间": "entry_date",
}


# ==================== 辅助函数 ====================

def calculate_performance_score_monthly(grade: str, raw_score: float, config: dict = None) -> Dict:
    """
    绩效月度快照算法（参数化版本）

    Args:
        grade: 绩效等级 (A, B+, B, C, D)
        raw_score: 原始计算分 (100 + 加分 - 扣分)
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'radar_value': 雷达图显示值,
            'display_label': 显示标签,
            'status_color': 状态颜色 (RED/ORANGE/GREEN),
            'alert_tag': 警示标签,
            'grade': 等级
        }
    """
    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    grade_coefficients = config['performance']['grade_coefficients']
    grade_ranges = config['performance']['grade_ranges']

    grade = grade.upper() if grade else 'B+'

    # 等级锁定规则（使用配置参数）
    if grade == 'D':
        radar_value = grade_ranges['D']['radar_override']  # 从配置读取
        status_color = 'RED'
        alert_tag = '⛔ 绩效不合格'
        display_label = f'D级 (系数{grade_coefficients["D"]})'
    elif grade == 'C':
        radar_value = min(max(raw_score, grade_ranges['C']['min']), grade_ranges['C']['max'])
        status_color = 'ORANGE'
        alert_tag = '⚠️ 绩效预警'
        display_label = f'C级 (系数{grade_coefficients["C"]})'
    elif grade == 'B':
        radar_value = min(max(raw_score, grade_ranges['B']['min']), grade_ranges['B']['max'])
        status_color = 'ORANGE'
        alert_tag = '⚠️ 未达基准'
        display_label = f'B级 (系数{grade_coefficients["B"]})'
    elif grade == 'B+':
        radar_value = min(max(raw_score, grade_ranges['B+']['min']), grade_ranges['B+']['max'])
        status_color = 'GREEN'
        alert_tag = '✅ 达标'
        display_label = f'B+级 (系数{grade_coefficients["B+"]})'
    elif grade == 'A':
        radar_value = min(max(raw_score, grade_ranges['A']['min']), grade_ranges['A']['max'])
        status_color = 'GREEN'
        alert_tag = '✅ 优秀'
        display_label = f'A级 (系数{grade_coefficients["A"]})'
    else:  # 默认B+
        radar_value = min(max(raw_score, grade_ranges['B+']['min']), grade_ranges['B+']['max'])
        status_color = 'GREEN'
        alert_tag = '✅ 达标'
        display_label = f'B+级 (系数{grade_coefficients["B+"]})'

    return {
        'radar_value': round(radar_value, 1),
        'display_label': display_label,
        'status_color': status_color,
        'alert_tag': alert_tag,
        'grade': grade,
        'mode': 'MONTHLY'
    }


def calculate_performance_score_period(grade_list: List[str], grade_dates: Optional[List[str]] = None, config: dict = None) -> Dict:
    """
    绩效周期加权算法（跨月、季度、年度）（参数化版本）

    新增时间衰减机制：D级和C级的影响会随时间推移而减弱

    Args:
        grade_list: 周期内所有月份的等级列表，如 ['A', 'B+', 'B', 'C']
        grade_dates: 每个等级对应的日期列表（可选），如 ['2024-01', '2024-02', ...]
                     如果提供，将启用时间衰减机制
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'radar_value': 雷达图显示值,
            'display_label': 显示标签,
            'status_color': 状态颜色,
            'alert_tag': 警示标签
        }
    """
    if not grade_list:
        return {
            'radar_value': 95.0,
            'display_label': '暂无数据',
            'status_color': 'GREEN',
            'alert_tag': '✅ 暂无数据',
            'mode': 'PERIOD'
        }

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    grade_coefficients = config['performance']['grade_coefficients']
    contamination_rules = config['performance']['contamination_rules']
    time_decay = config['performance'].get('time_decay', {
        'enabled': True,
        'decay_months': 6,
        'decay_rate': 0.9
    })

    # Step 1: 系数映射（使用配置）
    coeff_map = grade_coefficients

    coeffs = []
    d_count = 0
    c_count = 0
    d_count_effective = 0.0  # 带时间衰减的有效D级计数
    c_count_effective = 0.0  # 带时间衰减的有效C级计数

    # 如果启用时间衰减且提供了日期信息
    use_time_decay = time_decay.get('enabled', True) and grade_dates and len(grade_dates) == len(grade_list)

    if use_time_decay:
        from datetime import datetime

        now = datetime.now()
        decay_months_threshold = time_decay.get('decay_months', 6)
        decay_rate_per_month = time_decay.get('decay_rate', 0.9)

        for i, (grade, date_str) in enumerate(zip(grade_list, grade_dates)):
            grade = grade.upper() if grade else 'B+'
            coeffs.append(coeff_map.get(grade, 1.0))

            try:
                # 解析日期（支持 YYYY-MM 或 YYYY-MM-DD 格式）
                if len(date_str) == 7:  # YYYY-MM
                    grade_date = datetime.strptime(date_str, '%Y-%m')
                else:  # YYYY-MM-DD
                    grade_date = datetime.strptime(date_str[:7], '%Y-%m')

                # 计算距今月数
                months_ago = (now.year - grade_date.year) * 12 + (now.month - grade_date.month)

                if grade == 'D':
                    d_count += 1
                    # 时间衰减逻辑：
                    # 1. 只计入最近 decay_months_threshold 个月内的D级
                    # 2. 每个月衰减 (1 - decay_rate_per_month)
                    if months_ago <= decay_months_threshold:
                        # 计算衰减权重：decay_rate^months_ago
                        decay_weight = (decay_rate_per_month ** months_ago)
                        d_count_effective += decay_weight
                elif grade == 'C':
                    c_count += 1
                    # C级同样应用时间衰减（但阈值和惩罚可能不同）
                    if months_ago <= decay_months_threshold:
                        decay_weight = (decay_rate_per_month ** months_ago)
                        c_count_effective += decay_weight

            except Exception:
                # 日期解析失败，按原逻辑计数
                if grade == 'D':
                    d_count += 1
                    d_count_effective += 1
                elif grade == 'C':
                    c_count += 1
                    c_count_effective += 1
    else:
        # 不使用时间衰减，按原逻辑
        for grade in grade_list:
            grade = grade.upper() if grade else 'B+'
            coeffs.append(coeff_map.get(grade, 1.0))
            if grade == 'D':
                d_count += 1
                d_count_effective = d_count
            elif grade == 'C':
                c_count += 1
                c_count_effective = c_count

    # Step 2: 计算平均系数
    avg_coeff = sum(coeffs) / len(coeffs) if coeffs else 1.0

    # Step 3: 还原基础分 (系数1.0对应95分)
    base_score = avg_coeff * 95

    # Step 4: 执行"污点熔断"规则（使用时间衰减后的计数）
    d_threshold = contamination_rules['d_count_threshold']
    c_threshold = contamination_rules['c_count_threshold']
    d_cap = contamination_rules['d_cap_score']
    c_cap = contamination_rules['c_cap_score']

    if d_count_effective >= d_threshold:
        # D级熔断规则（使用衰减后的计数）
        final_score = min(base_score, d_cap)
        status_color = 'RED'
        if use_time_decay and d_count_effective < d_count:
            alert_tag = f'⛔ 存在D级考核 (有效{d_count_effective:.1f}次)'
        else:
            alert_tag = '⛔ 存在D级考核'
    elif c_count_effective >= c_threshold:
        # C级熔断规则（使用衰减后的计数）
        final_score = min(base_score, c_cap)
        status_color = 'ORANGE'
        if use_time_decay and c_count_effective < c_count:
            alert_tag = f'⚠️ 多次C级预警 (有效{c_count_effective:.1f}次)'
        else:
            alert_tag = '⚠️ 多次C级预警'
    else:
        # 正常输出
        final_score = min(base_score, 110)
        if final_score >= 95:
            status_color = 'GREEN'
            alert_tag = '✅ 综合达标'
        elif final_score >= 80:
            status_color = 'ORANGE'
            alert_tag = '⚠️ 未达基准'
        else:
            status_color = 'RED'
            alert_tag = '⛔ 综合不合格'

    # 生成显示标签
    display_label = f'平均系数{avg_coeff:.2f}'

    return {
        'radar_value': round(final_score, 1),
        'display_label': display_label,
        'status_color': status_color,
        'alert_tag': alert_tag,
        'mode': 'PERIOD',
        'd_count_raw': d_count,  # 原始D级次数
        'd_count_effective': round(d_count_effective, 2),  # 时间衰减后有效次数
        'time_decay_applied': use_time_decay
    }


def calculate_safety_score_dual_track(violations_list: List[float], months_active: int = 1, config: dict = None) -> Dict:
    """
    安全意识双轨评分模型（参数化版本）

    Args:
        violations_list: 违规扣分值列表，例如 [1, 3, 6]
        months_active: 统计周期包含的月份数（月度传1，年度传12或实际在职月数）
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'score_a': 行为分（习惯维度）,
            'score_b': 严重性分（后果维度）,
            'final_score': 最终分数（取两者最低）,
            'status_color': 状态颜色（RED/ORANGE/GREEN）,
            'alert_tag': 警示标签
        }
    """
    import math

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    behavior_track = config['safety']['behavior_track']
    severity_track = config['safety']['severity_track']
    thresholds = config['safety']['thresholds']

    # 维度A：行为习惯（捉拿惯犯）
    violation_count = len(violations_list)
    avg_freq = math.ceil(violation_count / months_active) if months_active > 0 else 0

    # 根据月均频次扣分（使用配置参数）
    freq_thresholds = behavior_track['freq_thresholds']  # [2, 5, 6]
    freq_multipliers = behavior_track['freq_multipliers']  # [2, 5, 10]

    if avg_freq <= freq_thresholds[0]:
        score_a_deduction = avg_freq * freq_multipliers[0]
    elif freq_thresholds[0] < avg_freq <= freq_thresholds[1]:
        score_a_deduction = avg_freq * freq_multipliers[1]
    else:  # avg_freq >= freq_thresholds[2]
        score_a_deduction = avg_freq * freq_multipliers[2]

    score_a = max(0, 100 - score_a_deduction)

    # 维度B：后果严重性（精准打击）（使用配置参数）
    score_b_deduction = 0
    critical_threshold = severity_track['critical_threshold']
    has_critical_violation = False

    for score_value in violations_list:
        # 根据配置的score_ranges确定系数
        multiplier = 1.0
        for range_rule in severity_track['score_ranges']:
            if 'max' in range_rule and 'min' not in range_rule:
                # 只有max，表示 < max
                if score_value < range_rule['max']:
                    multiplier = range_rule['multiplier']
                    break
            elif 'min' in range_rule and 'max' in range_rule:
                # 有min和max，表示范围
                if range_rule['min'] <= score_value < range_rule['max']:
                    multiplier = range_rule['multiplier']
                    break
            elif 'min' in range_rule and 'max' not in range_rule:
                # 只有min，表示 >= min
                if score_value >= range_rule['min']:
                    multiplier = range_rule['multiplier']
                    break

        score_b_deduction += score_value * multiplier

        if score_value >= critical_threshold:
            has_critical_violation = True

    score_b = max(0, 100 - score_b_deduction)

    # 最终分数：取两者最低
    final_score = min(score_a, score_b)

    # 警示逻辑（使用配置阈值）
    fail_score = thresholds['fail_score']
    warning_score = thresholds['warning_score']

    if final_score < fail_score or has_critical_violation:
        # 红线熔断
        status_color = "RED"
        alert_tag = "⛔ 重大红线（存在高扣分）" if has_critical_violation else "⛔ 安全不合格"
    elif fail_score <= final_score < warning_score:
        # 黄色预警
        status_color = "ORANGE"
        if score_a < score_b:
            alert_tag = "⚠️ 高频违规风险"
        else:
            alert_tag = "⚠️ 扣分过多风险"
    else:  # final_score >= warning_score
        # 绿色安全
        status_color = "GREEN"
        alert_tag = "✅ 安全"

    return {
        'score_a': round(score_a, 1),
        'score_b': round(score_b, 1),
        'final_score': round(final_score, 1),
        'status_color': status_color,
        'alert_tag': alert_tag,
        'violation_count': violation_count,
        'avg_freq': avg_freq
    }


def calculate_training_score_with_penalty(
    training_records: List[tuple],
    duration_days: int = 30,
    cert_years: Optional[float] = None,
    config: dict = None
) -> Dict:
    """
    培训/实操能力高级评分算法 - 包含毒性惩罚和动态年化（参数化版本）

    新增动态AFR阈值：根据取证年限区分新老员工，使用不同的评判标准

    Args:
        training_records: 培训记录列表，每条记录为 (score, is_qualified, is_disqualified, training_date)
        duration_days: 统计周期天数（用于年化计算）
        cert_years: 取证年限（可选），用于判断新老员工。
                    None 或 <1年 为新员工，>=1年为老员工
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        dict: {
            'radar_score': 最终雷达图分数（已惩罚）,
            'original_score': 原始基础分,
            'penalty_coefficient': 惩罚系数,
            'stats': {'total_ops', 'fail_count', 'duration_days'},
            'risk_alert': {'show', 'level', 'text', 'description'},
            'status_color': 状态颜色（用于前端显示）
        }
    """
    import math

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    penalty_rules = config['training']['penalty_rules']
    duration_thresholds = config['training']['duration_thresholds']

    # Step 0: 数据准备
    total_ops = len(training_records)

    # 如果没有记录，根据统计周期判断严重程度（使用配置参数）
    if total_ops == 0:
        short_term_days = duration_thresholds['short_term_days']
        mid_term_days = duration_thresholds['mid_term_days']
        default_scores = duration_thresholds['default_scores']

        # 短期未培训：正常情况，给基础分
        if duration_days <= short_term_days:
            return {
                'radar_score': default_scores['short'],
                'original_score': default_scores['short'],
                'penalty_coefficient': 1.0,
                'stats': {
                    'total_ops': 0,
                    'fail_count': 0,
                    'duration_days': duration_days
                },
                'risk_alert': {
                    'show': True,
                    'level': 'NORMAL',
                    'text': '未开展培训',
                    'description': f'统计周期{duration_days}天内未开展培训，属于正常情况。'
                },
                'status_color': 'GREEN',
                'alert_tag': '未开展培训'
            }
        # 中期缺训：需要关注
        elif duration_days <= mid_term_days:
            return {
                'radar_score': default_scores['mid'],
                'original_score': default_scores['mid'],
                'penalty_coefficient': 1.0,
                'stats': {
                    'total_ops': 0,
                    'fail_count': 0,
                    'duration_days': duration_days
                },
                'risk_alert': {
                    'show': True,
                    'level': 'NOTICE',
                    'text': '⚠️ 长期未培训',
                    'description': f'统计周期{duration_days}天内未开展培训，建议安排培训。'
                },
                'status_color': 'YELLOW',
                'alert_tag': '⚠️ 长期未培训'
            }
        # 长期严重缺训：严重问题
        else:
            return {
                'radar_score': default_scores['long'],
                'original_score': default_scores['long'],
                'penalty_coefficient': 1.0,
                'stats': {
                    'total_ops': 0,
                    'fail_count': 0,
                    'duration_days': duration_days
                },
                'risk_alert': {
                    'show': True,
                    'level': 'CRITICAL',
                    'text': '❌ 严重缺训',
                    'description': f'统计周期{duration_days}天（超过半年）内未开展任何培训，严重影响业务能力。'
                },
                'status_color': 'RED',
                'alert_tag': '❌ 严重缺训'
            }

    # Step 1: 判定失格次数
    fail_count = 0
    total_score = 0

    for record in training_records:
        # 从字典中提取字段（MySQL DictCursor 返回字典）
        score = record['score']
        is_qualified = record['is_qualified']
        is_disqualified = record['is_disqualified']
        training_date = record['training_date']

        # 转换 score 为数值类型（MySQL 可能返回字符串）
        try:
            score_value = int(score) if score else 0
        except (ValueError, TypeError):
            score_value = 0

        # 失格判定：is_disqualified=1 OR score=0 OR is_qualified=0
        if is_disqualified == 1 or score_value == 0 or is_qualified == 0:
            fail_count += 1

        total_score += score_value

    # Step 2: 计算基础分（简单平均）
    avg_score = total_score / total_ops if total_ops > 0 else 0
    base_score = avg_score  # 可以根据需要调整权重，这里简化为平均分

    # Step 3: 确定惩罚系数（核心风控逻辑）
    coeff = 1.0
    tag_level = 'NORMAL'
    alert_msg = '✅ 能力达标'
    description = ''

    # Priority A: 绝对熔断红线（使用配置参数）
    absolute_threshold = penalty_rules['absolute_threshold']
    small_sample = penalty_rules['small_sample']

    if fail_count >= absolute_threshold['fail_count']:
        coeff = absolute_threshold['coefficient']
        tag_level = 'CRITICAL'
        alert_msg = '❌ 业务能力差 (高频失格)'
        description = f'检测到绝对失格次数 ≥ {absolute_threshold["fail_count"]}次（实际{fail_count}次），系统判定为不合格。'

    # Priority B: 小样本保护 & 高危标记（使用配置参数）
    elif total_ops < small_sample['sample_size'] and fail_count > 0:
        coeff = small_sample['coefficient']
        tag_level = 'HIGH_RISK'
        alert_msg = '⚠️ 观察期失格 (高风险-需带教)'
        description = f'样本量不足（仅{total_ops}次操作），但已出现{fail_count}次失格。建议加强带教。'

    # Priority C: 大样本年化推演（使用动态AFR阈值）
    elif total_ops >= small_sample['sample_size']:
        # 计算年化失格频率 (AFR - Annualized Failure Rate)
        duration_days = max(1, duration_days)  # 防止除零
        AFR = (fail_count / duration_days) * 365

        # 根据取证年限选择合适的AFR阈值（新增动态阈值逻辑）
        is_new_employee = cert_years is None or cert_years < 1.0

        if is_new_employee:
            # 新员工（取证1年内）：使用更宽松的阈值
            afr_thresholds = penalty_rules.get('afr_thresholds_new_employee', penalty_rules.get('afr_thresholds', []))
            employee_type = "新员工"
        else:
            # 老员工（取证1年以上）：使用标准阈值
            afr_thresholds = penalty_rules.get('afr_thresholds_experienced', penalty_rules.get('afr_thresholds', []))
            employee_type = "老员工"

        # 从高到低检查AFR阈值（支持新版可配置阈值）
        # 优先使用配置中的 threshold 键，如果不存在则回退到硬编码逻辑
        matched = False
        
        # 尝试按照 threshold 降序排序（如果有）
        sorted_rules = []
        try:
             # 过滤出有效的规则并排序
             valid_rules = [r for r in afr_thresholds if 'threshold' in r or 'min' in r]
             # 统一获取阈值用于排序
             def get_thresh(r):
                 return float(r.get('threshold', r.get('min', 0)))
             sorted_rules = sorted(valid_rules, key=get_thresh, reverse=True)
        except:
             sorted_rules = afr_thresholds

        for rule in sorted_rules:
            # 获取规则阈值
            limit = float(rule.get('threshold', rule.get('min', 0)))
            
            if AFR >= limit:
                coeff = rule['coefficient']
                
                # 确定警示级别
                if coeff <= 0.5:
                    tag_level = 'CRITICAL'
                    label = rule.get('label', '高频失格')
                elif coeff <= 0.8:
                    tag_level = 'WARNING'
                    label = rule.get('label', '频率偏高')
                else:
                    tag_level = 'NOTICE'
                    label = rule.get('label', '偶发失格')
                    
                alert_msg = f'⚠️ {label} (年化 {AFR:.1f} 次)'
                description = f'当前周期{duration_days}天内失格{fail_count}次，年化等效{AFR:.1f}次/年，触发{label}阈值({limit})。'
                matched = True
                break

        if not matched:
            # AFR < 最低阈值
            coeff = 1.0
            tag_level = 'NORMAL'
            alert_msg = '✅ 能力达标'
            description = ''

    # 如果没有失格记录，保持正常
    elif fail_count == 0:
        coeff = 1.0
        tag_level = 'NORMAL'
        alert_msg = '✅ 能力达标'
        description = ''

    # Step 4: 计算最终分数
    final_score = base_score * coeff

    # 映射到前端颜色
    if tag_level == 'CRITICAL':
        status_color = 'RED'
    elif tag_level == 'HIGH_RISK':
        status_color = 'PURPLE'
    elif tag_level == 'WARNING':
        status_color = 'ORANGE'
    elif tag_level == 'NOTICE':
        status_color = 'YELLOW'
    else:
        status_color = 'GREEN'

    return {
        'radar_score': round(final_score, 1),
        'original_score': round(base_score, 1),
        'penalty_coefficient': coeff,
        'stats': {
            'total_ops': total_ops,
            'fail_count': fail_count,
            'duration_days': duration_days
        },
        'risk_alert': {
            'show': fail_count > 0,
            'level': tag_level,
            'text': alert_msg,
            'description': description
        },
        'status_color': status_color,
        'alert_tag': alert_msg
    }


def calculate_learning_ability_monthly(score_curr: float, score_prev: float) -> Dict:
    """
    学习能力评分 - 月度模式 (Algorithm A: Short-Term Sensitivity)

    核心设计：学习能力值 = 现状锚点分 (Position) + 趋势动能分 (Momentum)

    Args:
        score_curr: 本月综合三维得分 (0-100)
        score_prev: 上月综合三维得分 (0-100)，新员工传入 score_curr

    Returns:
        {
            'learning_score': 学习能力分数 (0-100+, 可能超过100),
            'delta': 月度变化量,
            'status_color': 状态颜色 (RED/ORANGE/YELLOW/GREEN/GOLD),
            'alert_tag': 警示标签,
            'tier': 评级 (潜力股/稳健型/懈怠型/高位企稳/低位躺平)
        }
    """
    # Step 1: 计算增量
    delta = score_curr - score_prev

    # Step 2: 计算基础成长分
    # 公式：以本月得分为基准，叠加变化的 1.5 倍权重
    learning_score = score_curr + (delta * 1.5)

    # Step 3: 应用修正逻辑
    tier = '稳健型'
    status_color = 'GREEN'
    alert_tag = '✅ 状态正常'

    # 情形 1：高位企稳 (大师红利)
    if score_curr >= 95 and delta >= -2:
        learning_score = max(100, learning_score)
        tier = '高位企稳'
        status_color = 'GOLD'
        alert_tag = '🏆 顶尖水平 (大师红利)'

    # 情形 2：低位躺平 (差生陷阱)
    elif score_curr < 70 and delta <= 0:
        learning_score = learning_score * 0.8
        tier = '低位躺平'
        status_color = 'RED'
        alert_tag = '❌ 差且无进步 (学习态度有问题)'

    # 情形 3：显著进步
    elif delta > 10:
        tier = '潜力股'
        status_color = 'GOLD'
        alert_tag = f'⭐ 进步神速 (+{delta:.1f}分)'

    # 情形 4：显著退步
    elif delta < -10:
        tier = '懈怠型'
        status_color = 'RED'
        alert_tag = f'⚠️ 严重退步 ({delta:.1f}分)'

    # 情形 5：小幅进步
    elif delta > 0:
        tier = '稳健型'
        status_color = 'GREEN'
        alert_tag = f'✅ 稳中有进 (+{delta:.1f}分)'

    # 情形 6：小幅退步
    elif delta < 0:
        tier = '需关注'
        status_color = 'YELLOW'
        alert_tag = f'⚡ 略有下滑 ({delta:.1f}分)'

    # 限制分数范围（但允许超过100）
    learning_score = max(0, learning_score)

    return {
        'learning_score': round(learning_score, 1),
        'delta': round(delta, 1),
        'slope': 0,  # 月度模式无斜率概念，设为0
        'status_color': status_color,
        'alert_tag': alert_tag,
        'tier': tier,
        'months': 1  # 月度模式统计1个月
    }


def calculate_learning_ability_longterm(
    score_list: List[float],
    config: dict = None,
    current_three_dim_score: float = None,
    group_avg: float = 1.0,
    initial_prev_viol: Optional[int] = None
) -> Dict:
    """
    [V5.0 核心算法二] 长周期·风险惯性聚合 (L_period)

    这是本模型的灵魂。按以下步骤实现：

    步骤 1：基础加权 (Base Score)
    - 对周期内单月得分进行时间加权平均，得到 base_score
    - 公式：base_score = Σ(score[i] × (1.0 + i × time_decay)) / Σweights

    步骤 2：计算"风险惯性" (Risk Inertia)
    - 扫描周期内的 zone 状态序列，寻找 "连续处于 DANGER/CRITICAL 的最大月数" (K_max)
    - 若 K_max < inertia_start_months: 惯性为 0
    - 若 K_max >= inertia_start_months:
        惯性惩罚 = min((K_max - Start + 1) × Step, max_penalty)

    步骤 3：最终计算
    - final_score = base_score × (1.0 - inertia_penalty_rate)
    - 若曾触发熔断(CRITICAL)，分数上限压制到40分

    业务含义：
    一个连续 4 个月处于危险边缘的"老油条"，即使每个月得分有 60 分（及格），
    经过惯性惩罚（-45%）后，最终得分只有 33 分（高危）。
    这精准识别了"事故前兆群体"。

    风险概率映射 (Dashboard Mapping)：
    - [事故前兆] PRE_ACCIDENT: 惯性惩罚 > 40% 或 曾触发熔断
    - [高危] HIGH_RISK: 分数 < 60
    - [重点关注] WATCH_LIST: 处于危险区 但 惯性低
    - [安全] SAFE: 其他

    Args:
        score_list: 周期内的违规数量列表（按时间顺序）
        config: 算法配置
        current_three_dim_score: 当前三维综合分（可选）
        group_avg: 班组平均违规数
        initial_prev_viol: 周期前一个月的违规数（用于计算第一个月的趋势）

    Returns:
        {
            'learning_score': float,          # 最终评分 (0-100)
            'risk_level': str,                # 风险等级: SAFE|WATCH_LIST|HIGH_RISK|PRE_ACCIDENT
            'inertia_penalty_rate': float,    # 惯性扣减率 (0.0 ~ 0.6)
            'max_consecutive_danger': int,    # 最大连续危险月数
            'base_score': float,              # 基础加权分（惯性前）
            'has_meltdown': bool,             # 是否曾触发熔断
            'zone_sequence': list,            # 各月风险区域序列
            'monthly_scores': list,           # 各月得分序列
            'slope': float,                   # 线性趋势斜率
            'average_score': float,           # 简单平均分
            'status_color': str,              # UI颜色
            'alert_tag': str,                 # 中文警示标签
            'tier': str,                      # 分层标签
            'months': int                     # 统计月数
        }
    """
    import numpy as np

    # =====================================================
    # 1. 初始化
    # =====================================================
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    algo_new = config.get('learning_new', {})
    time_decay = algo_new.get('time_decay_rate', 0.2)
    history_list = score_list

    if not history_list:
        return {
            'learning_score': 0,
            'risk_level': 'UNKNOWN',
            'inertia_penalty_rate': 0,
            'max_consecutive_danger': 0,
            'base_score': 0,
            'has_meltdown': False,
            'zone_sequence': [],
            'monthly_scores': [],
            'slope': 0,
            'average_score': 0,
            'status_color': 'GRAY',
            'alert_tag': '无数据',
            'tier': '无数据',
            'months': 0
        }

    # =====================================================
    # 2. 逐月计算 & 构建状态序列
    # =====================================================
    monthly_scores = []
    zone_sequence = []
    prev_viol = initial_prev_viol
    has_meltdown = False  # 记录是否有过熔断（用于一票否决）

    for i, curr_viol in enumerate(history_list):
        res = calculate_learning_ability_new(curr_viol, prev_viol, group_avg, config)
        monthly_scores.append(res['learning_score'])
        zone_sequence.append(res['zone'])

        if res.get('trend_type') == 'meltdown' or res['zone'] == 'CRITICAL':
            has_meltdown = True

        prev_viol = curr_viol

    # =====================================================
    # 步骤 1：计算基础加权分 (Base Score)
    # 公式：base_score = Σ(score[i] × weight[i]) / Σweight[i]
    # 权重：weight[i] = 1.0 + (i × time_decay)，近期月份权重更高
    # =====================================================
    total_w = 0
    w_sum = 0
    for i, score in enumerate(monthly_scores):
        w = 1.0 + (i * time_decay)
        w_sum += score * w
        total_w += w

    base_score = w_sum / total_w if total_w > 0 else 0

    # =====================================================
    # 步骤 2：计算风险惯性 (Risk Inertia)
    # =====================================================
    inertia_res = calculate_inertia_penalty(zone_sequence, config)
    penalty_rate = inertia_res['penalty_rate']
    max_consecutive = inertia_res['max_consecutive']

    # =====================================================
    # 步骤 3：最终计算
    # final_score = base_score × (1.0 - penalty_rate)
    # =====================================================
    final_score = base_score * (1.0 - penalty_rate)

    # 特殊处理：如果有熔断记录，分数上限强行压制到40分
    if has_meltdown:
        final_score = min(final_score, 40)

    final_score = round(max(0, final_score), 1)

    # =====================================================
    # 风险概率映射 (Dashboard Mapping)
    # =====================================================
    risk_level = 'SAFE'
    status_color = 'GREEN'
    alert_tag = '✅ 状态良好'
    tier_display = '安全'

    # 规则1: [事故前兆] 惯性惩罚 > 40% 或 曾触发熔断
    if penalty_rate >= 0.4 or has_meltdown:
        risk_level = 'PRE_ACCIDENT'
        status_color = 'RED'
        tier_display = '⛔ 事故前兆'
        if has_meltdown:
            alert_tag = f'⛔ 极高危 (曾触发熔断)'
        else:
            alert_tag = f'⛔ 极高危 (惯性扣减{penalty_rate*100:.0f}%)'

    # 规则2: [高危] 分数 < 60
    elif final_score < 60:
        risk_level = 'HIGH_RISK'
        status_color = 'ORANGE'
        if penalty_rate > 0:
            status_color = 'RED'
        tier_display = '高危群体'
        if penalty_rate > 0:
            alert_tag = f'🔴 高风险 (惯性扣减{penalty_rate*100:.0f}%)'
        else:
            alert_tag = f'🔴 高风险 (得分{final_score})'

    # 规则3: [重点关注] 处于危险区 但 惯性低
    elif len(zone_sequence) >= 2 and 'DANGER' in zone_sequence[-2:]:
        risk_level = 'WATCH_LIST'
        status_color = 'YELLOW'
        tier_display = '重点关注'
        alert_tag = '⚠️ 重点关注'

    # 规则4: [安全] 其他
    # 已设置默认值

    # =====================================================
    # 计算斜率（仅供展示）
    # =====================================================
    slope = 0.0
    if len(history_list) >= 2:
        try:
            x = np.arange(len(history_list))
            y = np.array(history_list)
            res_poly = np.polyfit(x, y, 1)
            slope = float(res_poly[0])
        except:
            pass

    return {
        'learning_score': final_score,
        'risk_level': risk_level,
        'inertia_penalty_rate': round(penalty_rate, 3),
        'max_consecutive_danger': max_consecutive,
        'base_score': round(base_score, 1),
        'has_meltdown': has_meltdown,
        'zone_sequence': zone_sequence,
        'monthly_scores': monthly_scores,

        # 兼容旧字段
        'slope': round(slope, 2),
        'average_score': round(float(np.mean(history_list)), 1),
        'status_color': status_color,
        'alert_tag': alert_tag,
        'tier': tier_display,
        'months': len(history_list)
    }


def calculate_stability_score(
    birth_date: Optional[str],
    work_start_date: Optional[str],
    entry_date: Optional[str],
    certification_date: Optional[str],
    solo_driving_date: Optional[str],
    historical_scores: Optional[Dict[str, List[float]]] = None,
    config: dict = None
) -> Dict:
    """
    职业稳定性综合评分算法（新版）

    评分维度：
    1. 资历维度（60%）：基于年龄、工龄、司龄、取证年限、单独驾驶年限
    2. 表现稳定性维度（40%）：基于过去一年绩效、安全、培训分值的波动度

    Args:
        birth_date: 出生日期 (YYYY-MM-DD)
        work_start_date: 参加工作时间 (YYYY-MM-DD)
        entry_date: 入司时间 (YYYY-MM-DD)
        certification_date: 取证时间 (YYYY-MM-DD)
        solo_driving_date: 单独驾驶时间 (YYYY-MM-DD)
        historical_scores: 过去一年的分数历史，格式：
            {
                'performance': [95.0, 96.0, ...],  # 最多12个月
                'safety': [92.0, 94.0, ...],
                'training': [88.0, 90.0, ...]
            }
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'stability_score': 最终稳定性分数 (0-100),
            'seniority_score': 资历维度分数 (0-100),
            'volatility_score': 稳定性维度分数 (0-100),
            'metrics': {
                'age_years': 年龄,
                'working_years': 工龄,
                'company_years': 司龄,
                'cert_years': 取证年限,
                'solo_years': 单独驾驶年限,
                'volatility': 综合波动系数
            },
            'status_color': 状态颜色 (RED/ORANGE/GREEN),
            'alert_tag': 警示标签,
            'tier': 评级 (资深稳定/经验丰富/新手期/高波动风险)
        }
    """
    from datetime import datetime
    import numpy as np

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    stability_config = config.get('stability', {
        'seniority_weights': {
            'age': 0.15,
            'working_years': 0.20,
            'company_years': 0.25,
            'cert_years': 0.20,
            'solo_years': 0.20
        },
        'seniority_thresholds': {
            'age_cap': 30,  # 年龄满30年算满分
            'working_cap': 20,  # 工龄满20年算满分
            'company_cap': 10,  # 司龄满10年算满分
            'cert_cap': 10,  # 取证满10年算满分
            'solo_cap': 10  # 单独驾驶满10年算满分
        },
        'dimension_weights': {
            'seniority': 0.60,  # 资历维度权重
            'volatility': 0.40   # 稳定性维度权重
        },
        'volatility_penalty': {
            'low_threshold': 5.0,     # 低波动阈值（标准差）
            'high_threshold': 15.0,   # 高波动阈值（标准差）
            'max_penalty': 0.5        # 最大惩罚系数
        }
    })

    now = datetime.now()

    # 辅助函数：解析日期（支持字符串和date对象）
    def parse_date(date_val):
        if not date_val:
            return None
        if isinstance(date_val, datetime):
            return date_val
        if hasattr(date_val, 'year'):  # date对象
            return datetime(date_val.year, date_val.month, date_val.day)
        if isinstance(date_val, str):
            return datetime.strptime(date_val, '%Y-%m-%d')
        return None

    # ==================== 维度1：资历评分（60%） ====================
    seniority_weights = stability_config['seniority_weights']
    seniority_thresholds = stability_config['seniority_thresholds']

    # 1.1 年龄计算
    age_years = 0
    if birth_date:
        try:
            birth = parse_date(birth_date)
            if birth:
                age_years = (now - birth).days / 365.25
        except:
            pass
    age_score = min(100, (age_years / seniority_thresholds['age_cap']) * 100)

    # 1.2 工龄计算
    working_years = 0
    if work_start_date:
        try:
            work_start = parse_date(work_start_date)
            if work_start:
                working_years = (now - work_start).days / 365.25
        except:
            pass
    working_score = min(100, (working_years / seniority_thresholds['working_cap']) * 100)

    # 1.3 司龄计算
    company_years = 0
    if entry_date:
        try:
            entry = parse_date(entry_date)
            if entry:
                company_years = (now - entry).days / 365.25
        except:
            pass
    company_score = min(100, (company_years / seniority_thresholds['company_cap']) * 100)

    # 1.4 取证年限计算
    cert_years = 0
    if certification_date:
        try:
            cert = parse_date(certification_date)
            if cert:
                cert_years = (now - cert).days / 365.25
        except:
            pass
    cert_score = min(100, (cert_years / seniority_thresholds['cert_cap']) * 100)

    # 1.5 单独驾驶年限计算
    solo_years = 0
    if solo_driving_date:
        try:
            solo = parse_date(solo_driving_date)
            if solo:
                solo_years = (now - solo).days / 365.25
        except:
            pass
    solo_score = min(100, (solo_years / seniority_thresholds['solo_cap']) * 100)

    # 计算资历加权分数
    seniority_score = (
        age_score * seniority_weights['age'] +
        working_score * seniority_weights['working_years'] +
        company_score * seniority_weights['company_years'] +
        cert_score * seniority_weights['cert_years'] +
        solo_score * seniority_weights['solo_years']
    )

    # ==================== 维度2：表现稳定性评分（40%） ====================
    volatility_score = 100  # 默认满分（无波动数据时）
    volatility_coefficient = 0

    if historical_scores and any(historical_scores.values()):
        # 计算每个维度的标准差
        std_devs = []

        for dimension in ['performance', 'safety', 'training']:
            scores = historical_scores.get(dimension, [])
            if scores and len(scores) >= 2:
                std_dev = float(np.std(scores))
                std_devs.append(std_dev)

        if std_devs:
            # 综合波动系数：使用平均标准差
            volatility_coefficient = float(np.mean(std_devs))

            # 根据波动系数计算分数
            low_threshold = stability_config['volatility_penalty']['low_threshold']
            high_threshold = stability_config['volatility_penalty']['high_threshold']
            max_penalty = stability_config['volatility_penalty']['max_penalty']

            if volatility_coefficient <= low_threshold:
                # 低波动：满分
                volatility_score = 100
            elif volatility_coefficient >= high_threshold:
                # 高波动：应用最大惩罚
                volatility_score = 100 * (1 - max_penalty)
            else:
                # 中等波动：线性惩罚
                penalty_ratio = (volatility_coefficient - low_threshold) / (high_threshold - low_threshold)
                penalty = max_penalty * penalty_ratio
                volatility_score = 100 * (1 - penalty)

    # ==================== 综合评分 ====================
    dimension_weights = stability_config['dimension_weights']
    final_score = (
        seniority_score * dimension_weights['seniority'] +
        volatility_score * dimension_weights['volatility']
    )

    # ==================== 分级和状态判定 ====================
    # 判定资历等级
    if company_years >= 5 and cert_years >= 5:
        seniority_tier = "资深员工"
    elif company_years >= 2 and cert_years >= 2:
        seniority_tier = "经验员工"
    elif cert_years >= 1:
        seniority_tier = "新手期"
    else:
        seniority_tier = "新员工"

    # 判定稳定性等级
    if volatility_coefficient == 0:
        volatility_tier = "无历史数据"
    elif volatility_coefficient <= low_threshold:
        volatility_tier = "表现稳定"
    elif volatility_coefficient <= high_threshold:
        volatility_tier = "波动适中"
    else:
        volatility_tier = "高波动风险"

    # 综合评级
    if final_score >= 85:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'GREEN'
        alert_tag = '✅ 稳定可靠'
    elif final_score >= 70:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'GREEN'
        alert_tag = '✅ 基本稳定'
    elif final_score >= 50:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'ORANGE'
        alert_tag = '⚠️ 稳定性一般'
    else:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'RED'
        alert_tag = '⛔ 不稳定'

    return {
        'stability_score': round(final_score, 1),
        'seniority_score': round(seniority_score, 1),
        'volatility_score': round(volatility_score, 1),
        'metrics': {
            'age_years': round(age_years, 1),
            'working_years': round(working_years, 1),
            'company_years': round(company_years, 1),
            'cert_years': round(cert_years, 1),
            'solo_years': round(solo_years, 1),
            'volatility': round(volatility_coefficient, 2)
        },
        'status_color': status_color,
        'alert_tag': alert_tag,
        'tier': tier
    }


def _month_index(month_str: str) -> int:
    year, month = map(int, month_str.split('-'))
    return year * 12 + (month - 1)


def _month_shift(month_str: str, delta: int) -> str:
    idx = _month_index(month_str) + delta
    year = idx // 12
    month = idx % 12 + 1
    return f"{year:04d}-{month:02d}"


def _month_range(start_month: str, end_month: str) -> List[str]:
    if _month_index(start_month) > _month_index(end_month):
        return []
    months = []
    curr = start_month
    while _month_index(curr) <= _month_index(end_month):
        months.append(curr)
        curr = _month_shift(curr, 1)
    return months


def _resolve_stability_window(start_date: Optional[str], end_date: Optional[str], config: dict) -> tuple:
    from datetime import datetime

    stability_config = config.get('stability_new', {})
    window_months = stability_config.get('window_months', 12)
    min_effective_months = stability_config.get('min_effective_months', 6)

    end_month = end_date or datetime.now().strftime('%Y-%m')
    window_start = _month_shift(end_month, -(window_months - 1))

    if start_date and _month_index(start_date) > _month_index(window_start):
        window_start = start_date

    span = _month_index(end_month) - _month_index(window_start) + 1
    if span < min_effective_months:
        window_start = _month_shift(end_month, -(min_effective_months - 1))

    return window_start, end_month


def _load_monthly_safety_violations(cur, emp_name: str, start_month: str, end_month: str) -> Dict[str, List[float]]:
    from blueprints.safety import extract_score_from_assessment

    start_date = f"{start_month}-01"
    end_next = f"{_month_shift(end_month, 1)}-01"

    cur.execute("""
        SELECT assessment, inspection_date
        FROM safety_inspection_records
        WHERE inspected_person = %s
        AND inspection_date >= %s
        AND inspection_date < %s
    """, [emp_name, start_date, end_next])
    rows = cur.fetchall()

    violations_by_month = {}
    for row in rows:
        insp_date = row['inspection_date']
        month_str = insp_date.strftime('%Y-%m') if hasattr(insp_date, 'strftime') else str(insp_date)[:7]
        score_val = extract_score_from_assessment(row['assessment'])
        if score_val > 0:
            violations_by_month.setdefault(month_str, []).append(float(score_val))

    return violations_by_month


def _build_monthly_safety_scores(violations_by_month: Dict[str, List[float]], months: List[str], config: dict) -> tuple:
    monthly_scores = {}
    monthly_issue_counts = {}

    for m in months:
        violations = violations_by_month.get(m, [])
        monthly_issue_counts[m] = len(violations)
        monthly_scores[m] = calculate_safety_score_dual_track(violations, 1, config)['final_score']

    return monthly_scores, monthly_issue_counts


def calculate_stability_for_employee(
    emp_name: str,
    start_date: Optional[str],
    end_date: Optional[str],
    config: dict,
    cur,
    safety_score_for_tip: Optional[float] = None,
    monthly_comprehensive_scores: Optional[Dict[str, float]] = None
) -> Dict:
    window_start, window_end = _resolve_stability_window(start_date, end_date, config)
    window_months = _month_range(window_start, window_end)

    last_12_start = _month_shift(window_end, -11)
    query_start = window_start if _month_index(window_start) <= _month_index(last_12_start) else last_12_start

    violations_by_month = _load_monthly_safety_violations(cur, emp_name, query_start, window_end)
    monthly_safety_scores, monthly_issue_counts = _build_monthly_safety_scores(violations_by_month, window_months, config)

    last_12_months = _month_range(last_12_start, window_end)
    issue_counts_last_12 = [len(violations_by_month.get(m, [])) for m in last_12_months]

    if not window_months:
        return {
            'stability_score': 50.0,
            'stability_label': '暂无数据',
            'status_color': 'GRAY',
            'alert_tag': '暂无数据'
        }

    return calculate_stability_score_new(
        window_months=window_months,
        monthly_safety_scores=monthly_safety_scores,
        monthly_issue_counts=monthly_issue_counts,
        issue_counts_last_12=issue_counts_last_12,
        monthly_comprehensive_scores=monthly_comprehensive_scores,
        safety_score_for_tip=safety_score_for_tip,
        config=config
    )


def calculate_stability_score_new(
    window_months: List[str],
    monthly_safety_scores: Dict[str, float],
    monthly_issue_counts: Dict[str, int],
    issue_counts_last_12: List[int],
    config: dict = None,
    monthly_comprehensive_scores: Optional[Dict[str, float]] = None,
    safety_score_for_tip: Optional[float] = None
) -> Dict:
    """
    稳定度评分算法（波动型）- 仅衡量安全表现波动

    设计原则：
    1. 使用近 window_months 个月安全分序列（不足 6 个月则标记低置信度）
    2. 0 记录按“无问题/未覆盖”规则区分
    3. 使用波动指标映射为稳定度分数（0-100）

    Args:
        window_months: 月份序列（YYYY-MM）
        monthly_safety_scores: 月度安全分（按月）
        monthly_issue_counts: 月度问题数（按月）
        issue_counts_last_12: 近 12 个月月度问题数列表（用于 0 记录判断）
        config: 算法配置
        monthly_comprehensive_scores: 月度综合分（用于 CV 对比提示，可选）
        safety_score_for_tip: 安全维度雷达分（用于低水平提示，可选）

    Returns:
        {
            'stability_score': 稳定度分数,
            'stability_label': 标签,
            'status_color': 状态颜色,
            'alert_tag': 标签文案,
            'volatility_metric': 波动指标,
            'volatility_value': 波动值,
            'coverage': 覆盖率 (有效月/窗口月),
            'confidence': 置信度,
            'volatility_tip': 波动异常提示,
            'low_level_tip': 低水平提示,
            'sample_tip': 样本不足提示,
            'safety_cv': 安全CV,
            'comprehensive_cv': 综合CV,
            'mean_safety': 安全均值
        }
    """
    import statistics

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    stability_config = config.get('stability_new', {})
    metric = stability_config.get('volatility_metric', 'mean_abs_delta')
    min_effective_months = stability_config.get('min_effective_months', 6)
    high_vol_threshold = stability_config.get('high_vol_threshold', 0.0667)
    k_multiplier = stability_config.get('k_multiplier', 1.2)
    score_floor = stability_config.get('score_floor', 40.0)
    score_ceiling = stability_config.get('score_ceiling', 100.0)

    # 0 记录判定规则
    avg_issues_12 = sum(issue_counts_last_12) / len(issue_counts_last_12) if issue_counts_last_12 else 0.0
    zero_streak_months = set()
    current_streak = []
    for m in window_months:
        if monthly_issue_counts.get(m, 0) == 0:
            current_streak.append(m)
        else:
            if len(current_streak) >= 3:
                zero_streak_months.update(current_streak)
            current_streak = []
    if len(current_streak) >= 3:
        zero_streak_months.update(current_streak)

    effective_months = []
    effective_scores = []
    for m in window_months:
        issue_count = monthly_issue_counts.get(m, 0)
        is_zero = issue_count == 0
        is_effective = (not is_zero) or (avg_issues_12 < 1) or (m in zero_streak_months)
        if is_effective:
            effective_months.append(m)
            effective_scores.append(float(monthly_safety_scores.get(m, 100.0)))

    window_count = len(window_months)
    effective_count = len(effective_scores)
    coverage = f"{effective_count}/{window_count}" if window_count else "0/0"

    # 波动指标计算
    metric_value = 0.0
    if metric == 'mean_abs_delta':
        if effective_count >= 2:
            diffs = [abs(effective_scores[i] - effective_scores[i - 1]) for i in range(1, effective_count)]
            metric_value = sum(diffs) / len(diffs) if diffs else 0.0
    elif metric == 'mad':
        if effective_count >= 1:
            median_val = statistics.median(effective_scores)
            deviations = [abs(s - median_val) for s in effective_scores]
            metric_value = statistics.median(deviations) if deviations else 0.0
    elif metric == 'cv':
        if effective_count >= 2:
            mean_val = statistics.mean(effective_scores)
            if mean_val > 0:
                metric_value = statistics.pstdev(effective_scores) / mean_val
    else:
        if effective_count >= 2:
            diffs = [abs(effective_scores[i] - effective_scores[i - 1]) for i in range(1, effective_count)]
            metric_value = sum(diffs) / len(diffs) if diffs else 0.0

    # 分位映射（线性插值）
    map_low_value = stability_config.get('score_map_low', 1.09)
    map_high_value = stability_config.get('score_map_high', 6.00)
    map_low_score = stability_config.get('score_map_low_score', 90.0)
    map_high_score = stability_config.get('score_map_high_score', 60.0)

    if map_high_value == map_low_value:
        stability_score = (map_low_score + map_high_score) / 2
    else:
        slope = (map_high_score - map_low_score) / (map_high_value - map_low_value)
        stability_score = map_low_score + slope * (metric_value - map_low_value)

    stability_score = max(score_floor, min(score_ceiling, stability_score))

    # 标签判定
    label_cutoffs = stability_config.get('label_cutoffs') or {}
    stable_cut = label_cutoffs.get('stable')
    medium_cut = label_cutoffs.get('medium')

    if not isinstance(stable_cut, (int, float)):
        stable_cut = 75
    if not isinstance(medium_cut, (int, float)):
        medium_cut = 60

    if stability_score >= stable_cut:
        stability_label = '稳定'
        status_color = 'GREEN'
        alert_tag = '✅ 稳定'
    elif stability_score >= medium_cut:
        stability_label = '波动偏大'
        status_color = 'ORANGE'
        alert_tag = '⚠️ 波动偏大'
    else:
        stability_label = '波动较大'
        status_color = 'RED'
        alert_tag = '⛔ 波动较大'

    # CV 计算（用于提示）
    safety_cv = 0.0
    if effective_count >= 2:
        mean_safety = statistics.mean(effective_scores)
        if mean_safety > 0:
            safety_cv = statistics.pstdev(effective_scores) / mean_safety
    else:
        mean_safety = statistics.mean(effective_scores) if effective_scores else 0.0

    comprehensive_cv = None
    comp_scores = []
    if monthly_comprehensive_scores:
        for m in effective_months:
            score_val = monthly_comprehensive_scores.get(m)
            if isinstance(score_val, (int, float)):
                comp_scores.append(float(score_val))
    if len(comp_scores) >= 2:
        mean_comp = statistics.mean(comp_scores)
        if mean_comp > 0:
            comprehensive_cv = statistics.pstdev(comp_scores) / mean_comp

    volatility_tip = None
    if comprehensive_cv is not None:
        if safety_cv >= high_vol_threshold and safety_cv > comprehensive_cv * k_multiplier:
            volatility_tip = "安全表现波动明显，波动高于整体表现"

    low_level_tip = None
    low_level_threshold = stability_config.get('low_level_threshold')
    if low_level_threshold is None:
        low_level_threshold = config.get('safety', {}).get('thresholds', {}).get('fail_score', 60)
    tip_basis = None
    if isinstance(safety_score_for_tip, (int, float)):
        tip_basis = float(safety_score_for_tip)
    elif effective_scores:
        tip_basis = mean_safety

    if tip_basis is not None and tip_basis <= low_level_threshold:
        low_level_tip = "整体安全水平偏低（即使稳定，仍需关注）"

    sample_tip = None
    if effective_count < min_effective_months:
        sample_tip = "样本不足，稳定度参考价值有限"

    metric_labels = {
        'mean_abs_delta': 'Mean |Δ|',
        'mad': 'MAD',
        'cv': 'CV'
    }

    return {
        'stability_score': round(float(stability_score), 1),
        'stability_label': stability_label,
        'status_color': status_color,
        'alert_tag': alert_tag,
        'volatility_metric': metric,
        'volatility_metric_label': metric_labels.get(metric, metric),
        'volatility_value': round(float(metric_value), 3),
        'coverage': coverage,
        'confidence': 'LOW' if effective_count < min_effective_months else 'OK',
        'volatility_tip': volatility_tip,
        'low_level_tip': low_level_tip,
        'sample_tip': sample_tip,
        'safety_cv': round(float(safety_cv), 3) if effective_scores else None,
        'comprehensive_cv': round(float(comprehensive_cv), 3) if comprehensive_cv is not None else None,
        'mean_safety': round(float(mean_safety), 2) if effective_scores else None
    }



def calculate_inertia_penalty(zone_sequence: List[str], config: dict) -> Dict:
    """
    [V5.0 核心] 计算风险惯性惩罚 (Risk Inertia) - 识别长尾高风险群体

    核心理念：防止"短期洗白"。扫描周期内的 zone 状态序列，
    寻找"连续处于 DANGER/CRITICAL 的最大月数" (K_max)。

    判定逻辑：
    - 若 K_max < inertia_start_months: 惯性为 0，不触发惩罚
    - 若 K_max >= inertia_start_months:
        惯性惩罚 = min((K_max - Start + 1) × Step, max_penalty)

    示例（标准档 Start=2, Step=0.15, Max=0.6）：
    - 连续2个月危险 → (2-2+1)×0.15 = 15% 惩罚
    - 连续3个月危险 → (3-2+1)×0.15 = 30% 惩罚
    - 连续4个月危险 → (4-2+1)×0.15 = 45% 惩罚
    - 连续5个月危险 → min(60%, 60%) = 60% 惩罚（封顶）

    Args:
        zone_sequence: 状态序列，例如 ['SAFE', 'DANGER', 'DANGER', 'SAFE']
        config: 算法配置

    Returns:
        {
            'penalty_rate': float,        # 惯性扣减率 (0.0 ~ max_penalty)
            'max_consecutive': int,       # 最大连续危险月数 (K_max)
            'is_triggered': bool,         # 是否触发惯性惩罚
            'start_threshold': int,       # 启动阈值
            'step': float,                # 步长
            'max_penalty': float          # 最大惩罚
        }
    """
    cfg = config.get('learning_new', {})

    # =====================================================
    # C. 风险惯性配置 (The Risk Inertia)
    # =====================================================
    inertia_start = cfg.get('inertia_start_months', 2)   # 惯性启动阈值
    inertia_step = cfg.get('inertia_step', 0.15)         # 惯性累积步长
    inertia_max = cfg.get('inertia_max_penalty', 0.6)    # 最大惯性惩罚

    if not zone_sequence:
        return {
            'penalty_rate': 0.0,
            'max_consecutive': 0,
            'is_triggered': False,
            'start_threshold': inertia_start,
            'step': inertia_step,
            'max_penalty': inertia_max
        }

    # =====================================================
    # 扫描连续危险月数
    # =====================================================
    max_conse = 0       # 最大连续危险月数
    current_conse = 0   # 当前连续计数

    for zone in zone_sequence:
        if zone in ['DANGER', 'CRITICAL']:
            current_conse += 1
        else:
            max_conse = max(max_conse, current_conse)
            current_conse = 0

    # 处理序列末尾的连续危险
    max_conse = max(max_conse, current_conse)

    # =====================================================
    # 计算惯性惩罚
    # =====================================================
    penalty_rate = 0.0
    is_triggered = False

    if max_conse >= inertia_start:
        is_triggered = True
        # 公式: (K_max - Start + 1) × Step
        raw_penalty = (max_conse - inertia_start + 1) * inertia_step
        penalty_rate = min(raw_penalty, inertia_max)

    return {
        'penalty_rate': round(penalty_rate, 3),
        'max_consecutive': max_conse,
        'is_triggered': is_triggered,
        'start_threshold': inertia_start,
        'step': inertia_step,
        'max_penalty': inertia_max
    }


def calculate_learning_ability_new(
    current_violations: int,
    previous_violations: Optional[int],
    group_avg_violations: float,
    config: dict = None
) -> Dict:
    """
    [V5.0 核心算法一] 单月风险状态判定 (L_month)

    升级说明：
    - 不再单纯看趋势，而是根据动态水位判定当月处于哪个"风险区域"
    - 必须返回当月的"区域状态 (Zone Status)"，供长周期算法计算惯性

    水位线计算：
    - warning_line = max(group_avg × ratio, warning_floor, historical_baseline)
    - warning_line = min(warning_line, ceiling_floor)  # 绝对天花板限制
    - critical_line = max(group_avg × critical_ratio, critical_floor)

    区域判定：
    - CRITICAL: N >= critical_line → 分数 0.0, 一票否决
    - DANGER:   N >= warning_line  → 分数按危险区系数计算
    - SAFE:     N < warning_line   → 分数按安全区系数计算

    Args:
        current_violations: 本月违规数
        previous_violations: 上月违规数 (用于辅助判定改善/恶化)
        group_avg_violations: 班组均值
        config: 算法配置

    Returns:
        {
            'score': float,           # 单月得分 (0-100)
            'learning_score': float,  # 兼容旧字段名
            'zone': str,              # 'SAFE' | 'DANGER' | 'CRITICAL'
            'count': int,             # 当月违规数
            'trend_type': str,        # 细分类型
            'status_color': str,      # UI颜色
            'alert_tag': str,         # 警示标签
            'warning_line': float,    # 关注线
            'critical_line': float    # 熔断线
        }
    """
    # 1. 提取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    cfg = config.get('learning_new', {})

    # =====================================================
    # A. 动态水位配置 (The Filter) - 防止群体漂移的核心红线
    # =====================================================
    ceiling_floor = cfg.get('trend_ceiling_floor', 5)         # 绝对天花板
    warning_ratio = cfg.get('trend_warning_ratio', 1.5)       # 关注线倍率
    warning_floor = cfg.get('trend_warning_floor', 2)         # 关注线保底
    critical_ratio = cfg.get('trend_critical_ratio', 3.0)     # 熔断线倍率
    critical_floor = cfg.get('trend_critical_floor', 5)       # 熔断线保底
    historical_baseline = cfg.get('historical_baseline', 3)   # 历史基准

    # =====================================================
    # B. 阶梯趋势系数 (The Matrix)
    # =====================================================
    # 安全区系数
    factor_reward = cfg.get('factor_reward', cfg.get('factor_improvement', 1.2))
    factor_stable = cfg.get('factor_stable', 1.0)
    factor_safe_fluctuation = cfg.get('factor_safe_fluctuation', 0.9)
    # 危险区系数
    factor_mitigation = cfg.get('factor_mitigation', cfg.get('factor_high_improvement', 0.8))
    factor_warning = cfg.get('factor_warning', 0.6)
    factor_solidification = cfg.get('factor_solidification', 0.4)
    factor_deterioration = cfg.get('factor_deterioration', 0.3)

    # =====================================================
    # 计算水位线
    # =====================================================
    # 关注线 = max(group_avg × ratio, warning_floor, historical_baseline)
    warning_line_dynamic = max(
        group_avg_violations * warning_ratio,
        warning_floor,
        historical_baseline
    )
    # 应用绝对天花板：水位线不能超过 ceiling_floor
    warning_line = warning_line_dynamic
    if ceiling_floor > 0:
        warning_line = min(warning_line, ceiling_floor)

    # 熔断线
    critical_line = max(group_avg_violations * critical_ratio, critical_floor)
    # 保证 critical >= warning + 1 (至少差1)
    critical_line = max(critical_line, warning_line + 1)

    # =====================================================
    # 区域判定 (Zone Detection)
    # =====================================================
    zone = 'SAFE'
    score_base = 90
    coeff = 1.0
    trend_type = 'stable'
    status_color = 'GREEN'
    alert_tag = ''

    # --- (1) CRITICAL ZONE: 触达熔断线 → 一票否决 ---
    if current_violations >= critical_line:
        return {
            'score': 0.0,
            'learning_score': 0.0,
            'zone': 'CRITICAL',
            'count': current_violations,
            'trend_type': 'meltdown',
            'status_color': 'RED',
            'alert_tag': f'⛔ 触达熔断线 ({current_violations}≥{critical_line:.0f})',
            'warning_line': round(warning_line, 1),
            'critical_line': round(critical_line, 1)
        }

    # --- (2) DANGER ZONE: 高于关注线 ---
    elif current_violations >= warning_line:
        zone = 'DANGER'
        score_base = 60  # 危险区及格分起点

        # 细分趋势判定
        if previous_violations is not None:
            if current_violations < previous_violations:
                # 危险区改善 (Mitigation): 减轻惩罚但不奖励
                coeff = factor_mitigation  # 0.8
                trend_type = 'high_improvement'
                alert_tag = '⚠️ 高位改善 (未脱险)'
                status_color = 'YELLOW'
            elif current_violations == previous_violations:
                # 危险区固化 (Solidification): 严厉惩罚
                coeff = factor_solidification  # 0.4
                trend_type = 'solidification'
                alert_tag = '⛔ 风险固化'
                status_color = 'ORANGE'
            else:
                # 危险区恶化 (Warning): 更严厉
                coeff = factor_deterioration  # 0.3
                trend_type = 'deterioration'
                alert_tag = '🔴 高位恶化'
                status_color = 'RED'
        else:
            # 无历史数据 (冷启动高位)
            coeff = factor_warning  # 0.6
            trend_type = 'cold_start_warning'
            alert_tag = '⚠️ 起步高危'
            status_color = 'YELLOW'

    # --- (3) SAFE ZONE: 低于关注线 ---
    else:
        zone = 'SAFE'
        score_base = 95

        # 奖励机制
        if previous_violations is not None:
            if current_violations < previous_violations:
                # 安全区改善: 奖励
                coeff = factor_reward  # 1.2
                trend_type = 'improvement'
                alert_tag = '📈 持续改善'
                status_color = 'GREEN'
            elif current_violations == previous_violations:
                # 安全区稳定: 保持
                coeff = factor_stable  # 1.0
                trend_type = 'safe_stable'
                alert_tag = '✅ 保持平稳'
                status_color = 'GREEN'
            else:
                # 安全区波动: 轻微惩罚
                coeff = factor_safe_fluctuation  # 0.9
                trend_type = 'safe_fluctuation'
                alert_tag = '📉 安全波动'
                status_color = 'BLUE'
        else:
            # 冷启动良好
            coeff = factor_stable  # 1.0
            trend_type = 'cold_start_good'
            alert_tag = '✅ 表现良好'
            status_color = 'GREEN'

    # =====================================================
    # 计算最终得分
    # =====================================================
    final_score = score_base * coeff

    # 群体校准补偿：优于班组平均 +10%
    if current_violations < group_avg_violations:
        final_score *= 1.1

    final_score = min(100, max(0, final_score))

    return {
        'score': round(final_score, 1),
        'learning_score': round(final_score, 1),  # 兼容旧字段名
        'zone': zone,
        'count': current_violations,
        'trend_type': trend_type,
        'status_color': status_color,
        'alert_tag': alert_tag,
        'warning_line': round(warning_line, 1),
        'critical_line': round(critical_line, 1)
    }




def calculate_stability_period_aggregated(monthly_data, config):
    """
    长周期稳定度聚合逻辑 (V4.0)
    Args:
        monthly_data: list of dicts, item: {'score': float, 'has_redline': bool}
        config: EvaluationConfig dict
    Returns:
        dict: {
            'final_score': float,
            'is_veto': bool,
            'avg_score': float,
            'cv_discount': float,
            'cv': float,
            'alert_tag': str
        }
    """
    import statistics
    
    # 获取配置参数
    stability_config = config.get('stability_new', {})
    period_cv_sensitivity = stability_config.get('period_cv_sensitivity', 1.2)
    time_decay = config.get('learning_new', {}).get('time_decay_rate', 0.2)
    
    # 1. 短板熔断检查 (Veto Check)
    scores = []
    veto_triggered = False
    
    for month in monthly_data:
        # 条件：红线违规 或 得分为0
        # 注意：浮点数比较用 epsilon
        if month.get('has_redline', False) or month.get('score', 0) <= 0.001:
            veto_triggered = True
            return {
                'final_score': 0.0,
                'is_veto': True,
                'avg_score': 0.0,
                'cv_discount': 0.0,
                'cv': 0.0,
                'alert_tag': '❌ 熔断 (红线/零分)'
            }
        scores.append(month['score'])
        
    if not scores: # 无数据
        return {
            'final_score': 100.0,
            'is_veto': False,
            'avg_score': 100.0,
            'cv_discount': 1.0,
            'cv': 0.0,
            'alert_tag': '✅ 稳定'
        }

    # 2. 加权平均 (Weighted Average)
    weighted_sum = 0
    total_w = 0
    # 假设 scores 顺序为 [最早月 ... 最近月]
    for i, s in enumerate(scores):
        w = 1.0 + (i * time_decay)
        weighted_sum += s * w
        total_w += w
    
    avg_score = weighted_sum / total_w if total_w > 0 else 0
    
    # 3. 波动惩罚 (CV Discount)
    if len(scores) < 2:
        cv = 0.0
        discount = 1.0
    else:
        mean_val = statistics.mean(scores)
        if mean_val > 0.001:
            stdev_val = statistics.pstdev(scores) 
            cv = stdev_val / mean_val
        else:
            cv = 0.0
            
        discount = 1.0 - (cv * period_cv_sensitivity)
        discount = max(0.0, min(1.0, discount))
        
    final_score = avg_score * discount
    
    # 确定标签
    if final_score >= 80:
        tag = '✅ 稳定'
    elif final_score >= 60:
        tag = '⚠️ 波动'
    else:
        tag = '❌ 不稳定'
        
    return {
        'final_score': final_score,
        'is_veto': False,
        'avg_score': avg_score,
        'cv_discount': discount,
        'cv': cv,
        'alert_tag': tag
    }


def _parse_date_string(value: Optional[str]) -> Optional[date]:
    """解析日期字符串为date对象"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    fmts = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y%m%d",
        "%Y-%m",
        "%Y/%m",
        "%Y.%m",
        "%Y%m",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(raw, fmt)
            if fmt in {"%Y-%m", "%Y/%m", "%Y.%m", "%Y%m"}:
                dt = dt.replace(day=1)
            return dt.date()
        except ValueError:
            continue
    return None


def _normalize_date_to_str(value: Optional[str]) -> Optional[str]:
    """标准化日期为字符串"""
    parsed = _parse_date_string(value)
    return parsed.strftime("%Y-%m-%d") if parsed else None


def _calculate_age(birth_date: Optional[str]) -> Optional[int]:
    """计算年龄"""
    parsed = _parse_date_string(birth_date)
    if not parsed:
        return None
    today = date.today()
    years = today.year - parsed.year
    if (today.month, today.day) < (parsed.month, parsed.day):
        years -= 1
    return max(years, 0)


def _calculate_years_since(date_str: Optional[str]) -> Optional[float]:
    """计算从指定日期到今天的年数"""
    parsed = _parse_date_string(date_str)
    if not parsed:
        return None
    today = date.today()
    if parsed > today:
        return 0.0
    years = (today - parsed).days / 365.25
    return round(years, 1)


def _serialize_person(row: Dict) -> Dict:
    """序列化人员数据，添加计算字段"""
    data = dict(row)
    data["age"] = _calculate_age(data.get("birth_date"))
    data["working_years"] = _calculate_years_since(data.get("work_start_date"))
    data["tenure_years"] = _calculate_years_since(data.get("entry_date"))
    return data


def _build_personnel_charts(rows: List[Dict]) -> Dict:
    """构建人员统计图表数据"""
    # 年龄分布
    age_labels = ["25岁及以下", "26-35岁", "36-45岁", "46岁及以上"]
    age_counts = [0, 0, 0, 0]
    for row in rows:
        age = row.get("age")
        if age is None:
            continue
        if age <= 25:
            age_counts[0] += 1
        elif 26 <= age <= 35:
            age_counts[1] += 1
        elif 36 <= age <= 45:
            age_counts[2] += 1
        else:
            age_counts[3] += 1

    # 学历分布
    education_counter = Counter(
        row.get("education") or "未填写" for row in rows
    )
    education_labels = list(education_counter.keys())
    education_counts = [education_counter[label] for label in education_labels]

    # 工龄分布
    tenure_labels = ["1年以下", "1-3年", "3-5年", "5-10年", "10年以上"]
    tenure_counts = [0, 0, 0, 0, 0]
    for row in rows:
        tenure = row.get("tenure_years")
        if tenure is None:
            continue
        if tenure < 1:
            tenure_counts[0] += 1
        elif 1 <= tenure < 3:
            tenure_counts[1] += 1
        elif 3 <= tenure < 5:
            tenure_counts[2] += 1
        elif 5 <= tenure < 10:
            tenure_counts[3] += 1
        else:
            tenure_counts[4] += 1

    return {
        "age": {"labels": age_labels, "values": age_counts},
        "education": {"labels": education_labels, "values": education_counts},
        "tenure": {"labels": tenure_labels, "values": tenure_counts},
    }


# ==================== 数据库访问函数 ====================

def list_personnel():
    """列出所有可访问的人员"""
    from flask import session
    user_role = session.get('role', 'user')

    conn = get_db()
    cur = conn.cursor()

    # 管理员可以看到所有员工，其他角色只能看到可访问部门的员工
    if user_role == 'admin':
        query = """
            SELECT e.emp_no, e.name, e.department_id, d.name as department_name,
                   e.class_name, e.position, e.birth_date, e.certification_date,
                   e.solo_driving_date, e.marital_status, e.hometown,
                   e.political_status, e.education, e.graduation_school,
                   e.work_start_date, e.entry_date, e.specialty
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            ORDER BY CAST(e.emp_no AS SIGNED)
        """
        try:
            cur.execute(query)
        except Exception:
            cur.execute(query.replace("CAST(e.emp_no AS SIGNED)", "e.emp_no"))
    else:
        accessible_dept_ids = get_accessible_department_ids()
        if not accessible_dept_ids:
            return []

        placeholders = ','.join(['%s'] * len(accessible_dept_ids))
        query = f"""
            SELECT e.emp_no, e.name, e.department_id, d.name as department_name,
                   e.class_name, e.position, e.birth_date, e.certification_date,
                   e.solo_driving_date, e.marital_status, e.hometown,
                   e.political_status, e.education, e.graduation_school,
                   e.work_start_date, e.entry_date, e.specialty
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            WHERE e.department_id IN ({placeholders})
            ORDER BY CAST(e.emp_no AS SIGNED)
        """
        try:
            cur.execute(query, accessible_dept_ids)
        except Exception:
            cur.execute(
                query.replace("CAST(e.emp_no AS SIGNED)", "e.emp_no"),
                accessible_dept_ids,
            )

    rows = cur.fetchall()
    result = []
    for row in rows:
        person_dict = _serialize_person(row)
        # 添加计算字段
        if person_dict.get('certification_date'):
            person_dict['certification_years'] = calculate_years_from_date(person_dict['certification_date'])
        else:
            person_dict['certification_years'] = None

        if person_dict.get('solo_driving_date'):
            person_dict['solo_driving_years'] = calculate_years_from_date(person_dict['solo_driving_date'])
        else:
            person_dict['solo_driving_years'] = None

        result.append(person_dict)

    return result


def get_personnel(emp_no: str) -> Optional[Dict]:
    """获取指定工号的人员信息"""
    uid = require_user_id()

    # 🔒 权限检查: 非管理员需要验证是否有权访问该员工
    from flask import session
    user_role = session.get('role', 'user')
    if user_role != 'admin':
        if not validate_employee_access(emp_no):
            return None

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT e.emp_no, e.name, e.department_id, d.name as department_name,
               e.class_name, e.position, e.birth_date, e.certification_date,
               e.solo_driving_date, e.marital_status, e.hometown,
               e.political_status, e.education, e.graduation_school,
               e.work_start_date, e.entry_date, e.specialty, e.created_at
        FROM employees e
        LEFT JOIN departments d ON e.department_id = d.id
        WHERE e.emp_no=%s
        """,
        (emp_no,),
    )
    row = cur.fetchone()
    if not row:
        return None

    person_dict = _serialize_person(row)
    # 添加计算字段
    if person_dict.get('certification_date'):
        person_dict['certification_years'] = calculate_years_from_date(person_dict['certification_date'])
    if person_dict.get('solo_driving_date'):
        person_dict['solo_driving_years'] = calculate_years_from_date(person_dict['solo_driving_date'])

    return person_dict


def _sanitize_person_payload(data: Dict[str, Optional[str]]) -> Dict[str, Optional[str]]:
    """清理和标准化人员数据"""
    sanitized: Dict[str, Optional[str]] = {}
    for field in PERSONNEL_DB_COLUMNS + ["emp_no", "name"]:
        if field == "emp_no":
            value = str(data.get(field) or "").strip()
            sanitized[field] = value or None
            continue
        raw_val = data.get(field)
        if raw_val is None:
            sanitized[field] = None
            continue
        if field in PERSONNEL_DATE_FIELDS:
            sanitized[field] = _normalize_date_to_str(raw_val)
        else:
            sanitized[field] = str(raw_val).strip() or None
    return sanitized


def upsert_personnel(data: Dict[str, Optional[str]]) -> bool:
    """插入或更新人员信息"""
    payload = _sanitize_person_payload(data)
    emp_no = payload.get("emp_no")
    name = payload.get("name")
    department_id = payload.get("department_id")

    if not emp_no or not name:
        return False

    # department_id是必填项，如果没有提供则返回False
    if department_id is None or department_id == "":
        return False

    # 转换department_id为整数
    try:
        department_id = int(department_id)
    except (ValueError, TypeError):
        return False

    uid = require_user_id()
    conn = get_db()
    cur = conn.cursor()

    # 注意: UNIQUE约束是emp_no（全局唯一），数据以department_id为基准隔离
    columns = ["emp_no", "name", "created_by", "department_id"] + [col for col in PERSONNEL_DB_COLUMNS if col != "department_id"]
    values = [emp_no, name, uid, department_id] + [payload.get(col) for col in PERSONNEL_DB_COLUMNS if col != "department_id"]
    update_clause = ", ".join(
        f"{col}=VALUES({col})" for col in ["name", "department_id"] + [col for col in PERSONNEL_DB_COLUMNS if col != "department_id"]
    )
    cur.execute(
        f"""
        INSERT INTO employees ({", ".join(columns)})
        VALUES ({", ".join(["%s"] * len(columns))})
        ON DUPLICATE KEY UPDATE {update_clause}
        """,
        values,
    )
    conn.commit()
    return True


def bulk_import_personnel(records: List[Dict[str, Optional[str]]]) -> int:
    """批量导入人员信息"""
    imported = 0
    for record in records:
        if upsert_personnel(record):
            imported += 1
    return imported


def update_personnel_field(emp_no: str, field: str, value: Optional[str]) -> bool:
    """更新人员的单个字段"""
    if field not in {"name", *PERSONNEL_DB_COLUMNS}:
        return False

    # 🔒 权限检查: 非管理员需要验证是否有权修改该员工
    from flask import session
    user_role = session.get('role', 'user')
    if user_role != 'admin':
        if not validate_employee_access(emp_no):
            return False

    payload = _sanitize_person_payload({field: value})
    uid = require_user_id()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        f"""
        UPDATE employees
        SET {field} = %s
        WHERE emp_no=%s
        """,
        (payload.get(field), emp_no),
    )
    conn.commit()
    affected = cur.rowcount > 0
    return affected


def delete_employee(emp_no):
    """删除员工"""
    uid = require_user_id()

    # 🔒 权限检查: 非管理员需要验证是否有权删除该员工
    from flask import session
    user_role = session.get('role', 'user')
    if user_role != 'admin':
        if not validate_employee_access(emp_no):
            return False

    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM employees WHERE emp_no=%s", (emp_no,))
    conn.commit()
    return True


# ==================== 路由处理 ====================

@personnel_bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    """人员管理首页"""
    if request.method == 'POST':
        # 🔒 权限检查: 创建/更新员工需要管理员权限
        from flask import session
        user_role = session.get('role', 'user')
        if user_role not in ['admin', 'manager']:
            flash("您没有权限执行此操作，需要部门管理员或系统管理员权限", "danger")
            return redirect(url_for("personnel.index"))

        form_payload = {}
        for field in PERSONNEL_FIELD_SCHEME:
            key = field["name"]
            if field["input_type"] == "textarea":
                form_payload[key] = request.form.get(key, "")
            else:
                form_payload[key] = request.form.get(key, "").strip()
        saved = upsert_personnel(form_payload)
        if saved:
            flash("人员信息已保存。", "success")
        else:
            flash("请填写有效的工号和姓名。", "warning")
        return redirect(url_for("personnel.index"))

    rows = list_personnel()
    accessible_departments = get_accessible_departments()

    return render_template(
        "personnel.html",
        title=f"人员管理 | {APP_TITLE}",
        rows=rows,
        field_scheme=PERSONNEL_FIELD_SCHEME,
        select_options=PERSONNEL_SELECT_OPTIONS,
        accessible_departments=accessible_departments,
    )


@personnel_bp.route('/template')
@login_required
def template():
    """下载人员导入模板"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人员导入模板"

    headers = [field["label"] for field in PERSONNEL_FIELD_SCHEME]
    sheet.append(headers)

    examples = {
        "emp_no": "1001",
        "name": "张三",
        "class_name": "一班",
        "position": "班长",
        "birth_date": "1990-01-01",
        "marital_status": "已婚",
        "hometown": "江苏南京",
        "political_status": "群众",
        "education": "本科",
        "graduation_school": "某某大学",
        "work_start_date": "2012-07-01",
        "entry_date": "2018-03-15",
        "specialty": "摄影、篮球",
    }
    sheet.append([examples.get(field["name"], "") for field in PERSONNEL_FIELD_SCHEME])

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"personnel_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@personnel_bp.route('/import', methods=['POST'])
@manager_required
def import_data():
    """批量导入人员数据"""
    file_obj = request.files.get("file")
    if not file_obj or file_obj.filename == "":
        flash("请选择包含花名册数据的 Excel 文件。", "warning")
        return redirect(url_for("personnel.index"))
    ext = file_obj.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx"}:
        flash("目前仅支持上传 .xlsx 文件。", "warning")
        return redirect(url_for("personnel.index"))
    try:
        workbook = load_workbook(file_obj, data_only=True)
        sheet = workbook.active
    except Exception as exc:  # noqa: BLE001
        flash(f"无法读取 Excel 文件：{exc}", "danger")
        return redirect(url_for("personnel.index"))

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        flash("Excel 文件为空。", "warning")
        return redirect(url_for("personnel.index"))

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    field_map = [PERSONNEL_IMPORT_HEADER_MAP.get(header) for header in headers]

    if "emp_no" not in field_map or "name" not in field_map:
        flash('Excel 首行必须包含"工号"与"姓名"列。', "warning")
        return redirect(url_for("personnel.index"))

    # 获取部门映射，用于处理Excel中的部门信息
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM departments")
    dept_name_map = {row['name']: row['id'] for row in cur.fetchall()}

    # 获取当前用户可访问的部门ID列表（用于权限验证）
    accessible_dept_ids = get_accessible_department_ids()

    records: List[Dict[str, Optional[str]]] = []
    skipped_no_dept = 0
    skipped_no_permission = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: Dict[str, Optional[str]] = {}
        for idx, cell in enumerate(row):
            field = field_map[idx] if idx < len(field_map) else None
            if not field:
                continue
            record[field] = cell
            
        # 处理部门ID：支持名称匹配
        raw_dept = record.get('department_id')
        final_dept_id = None

        if raw_dept:
            raw_dept_str = str(raw_dept).strip()
            if raw_dept_str.isdigit():
                final_dept_id = int(raw_dept_str)
            elif raw_dept_str in dept_name_map:
                final_dept_id = dept_name_map[raw_dept_str]

        if not final_dept_id:
            # 未填写部门或部门无效
            skipped_no_dept += 1
        elif final_dept_id not in accessible_dept_ids:
            # 部门存在但无权限导入到该部门
            skipped_no_permission += 1
        else:
            # 部门有效且有权限
            record['department_id'] = str(final_dept_id)
            records.append(record)

    if not records:
        msg_parts = ["未导入任何数据。"]
        if skipped_no_dept > 0:
            msg_parts.append(f"{skipped_no_dept} 条记录因未填写部门或部门无效被跳过。")
        if skipped_no_permission > 0:
            msg_parts.append(f"{skipped_no_permission} 条记录因无权限导入到该部门被跳过。")
        if not skipped_no_dept and not skipped_no_permission:
            msg_parts.append("未识别到任何有效行。")
        flash(" ".join(msg_parts), "warning")

        # 记录失败的导入操作
        log_import_operation(
            module='personnel',
            operation='import',
            file_name=file_obj.filename,
            total_rows=skipped_no_dept + skipped_no_permission,
            success_rows=0,
            failed_rows=0,
            skipped_rows=skipped_no_dept + skipped_no_permission,
            error_message=" ".join(msg_parts),
            import_details={
                'skipped_no_dept': skipped_no_dept,
                'skipped_no_permission': skipped_no_permission
            }
        )
        return redirect(url_for("personnel.index"))

    imported = bulk_import_personnel(records)

    # 计算总行数
    total_rows = len(records) + skipped_no_dept + skipped_no_permission

    # 构建提示消息
    msg = f"已导入/更新 {imported} 名员工信息。"
    msg_parts = []
    if skipped_no_dept > 0:
        msg_parts.append(f"{skipped_no_dept} 条记录因未填写部门或部门无效被跳过")
    if skipped_no_permission > 0:
        msg_parts.append(f"{skipped_no_permission} 条记录因无权限导入到该部门被跳过")

    if msg_parts:
        msg += " 另有 " + "、".join(msg_parts) + "。"
        flash_type = "warning"
    else:
        flash_type = "success"

    flash(msg, flash_type)

    # 记录导入操作日志
    log_import_operation(
        module='personnel',
        operation='import',
        file_name=file_obj.filename,
        total_rows=total_rows,
        success_rows=imported,
        failed_rows=0,
        skipped_rows=skipped_no_dept + skipped_no_permission,
        import_details={
            'imported': imported,
            'skipped_no_dept': skipped_no_dept,
            'skipped_no_permission': skipped_no_permission,
            'accessible_departments': len(accessible_dept_ids)
        }
    )

    return redirect(url_for("personnel.index"))


@personnel_bp.route('/<emp_no>')
@login_required
def preview(emp_no):
    """查看人员详情"""
    person = get_personnel(emp_no)
    if not person:
        flash("未找到该员工。", "warning")
        return redirect(url_for("personnel.index"))
    return render_template(
        "personnel_preview.html",
        title=f"{person.get('name', '')} | 人员档案 · {APP_TITLE}",
        person=person,
        field_scheme=PERSONNEL_FIELD_SCHEME,
        select_options=PERSONNEL_SELECT_OPTIONS,
    )


@personnel_bp.route('/<emp_no>/update', methods=['POST'])
@login_required
@manager_required
def update(emp_no):
    """更新人员信息字段（仅限部门管理员及以上权限）"""
    payload = request.get_json(silent=True) or request.form
    field = (payload.get("field") or "").strip()
    value = payload.get("value")
    if field in PERSONNEL_DATE_FIELDS and isinstance(value, str):
        value = value.strip()
    if not field:
        return jsonify({"ok": False, "message": "未指定字段"}), 400
    updated = update_personnel_field(emp_no, field, value)
    if not updated:
        return jsonify({"ok": False, "message": "更新失败或字段不受支持"}), 400
    person = get_personnel(emp_no)
    return jsonify({"ok": True, "person": person})


@personnel_bp.route('/batch-delete', methods=['POST'])
@login_required
@manager_required
def batch_delete():
    """批量删除员工（仅限部门管理员及以上权限）"""
    emp_nos = request.form.getlist('emp_nos')

    if not emp_nos:
        flash("未选择要删除的员工", "warning")
        return redirect(url_for("personnel.index"))

    uid = require_user_id()
    from flask import session
    user_role = session.get('role', 'user')

    conn = get_db()
    cur = conn.cursor()

    deleted_count = 0
    skipped_count = 0
    for emp_no in emp_nos:
        emp_no = emp_no.strip()
        if emp_no:
            # 🔒 权限检查: 非管理员需要验证是否有权删除每个员工
            if user_role != 'admin':
                if not validate_employee_access(emp_no):
                    skipped_count += 1
                    continue

            cur.execute("DELETE FROM employees WHERE emp_no=%s", (emp_no,))
            if cur.rowcount > 0:
                deleted_count += 1

    conn.commit()

    if deleted_count > 0:
        message = f"成功删除 {deleted_count} 名员工"
        if skipped_count > 0:
            message += f"，跳过 {skipped_count} 名无权删除的员工"
        flash(message, "success")
    elif skipped_count > 0:
        flash(f"跳过 {skipped_count} 名无权删除的员工", "warning")
    else:
        flash("未删除任何员工", "info")

    return redirect(url_for("personnel.index"))


@personnel_bp.route('/employees')
@login_required
def employees_legacy_redirect():
    """旧版employees路由重定向"""
    flash("花名册入口已升级为人员管理，请使用新页面。", "info")
    return redirect(url_for("personnel.index"))


@personnel_bp.route('/dashboard')
@login_required
def dashboard():
    """人员工作台首页"""
    # 先获取部门权限（内部会使用自己的数据库连接）
    accessible_dept_ids = get_accessible_department_ids()

    # 再获取新的数据库连接用于统计查询
    conn = get_db()
    cur = conn.cursor()

    # ===== 统计数据查询 =====
    # 1. 员工总数
    if accessible_dept_ids:
        placeholders = ','.join(['%s'] * len(accessible_dept_ids))
        cur.execute(f"SELECT COUNT(*) AS cnt FROM employees WHERE department_id IN ({placeholders})", accessible_dept_ids)
    else:
        cur.execute("SELECT COUNT(*) AS cnt FROM employees")
    employee_count = cur.fetchone()['cnt']

    # 2. 部门/班组数量
    if accessible_dept_ids:
        placeholders = ','.join(['%s'] * len(accessible_dept_ids))
        cur.execute(f"SELECT COUNT(*) AS cnt FROM departments WHERE id IN ({placeholders})", accessible_dept_ids)
    else:
        cur.execute("SELECT COUNT(*) AS cnt FROM departments")
    dept_count = cur.fetchone()['cnt']

    # 3. 培训覆盖率（有培训记录的员工 / 总员工数）
    try:
        if accessible_dept_ids:
            placeholders = ','.join(['%s'] * len(accessible_dept_ids))
            cur.execute(f"""
                SELECT COUNT(DISTINCT tr.emp_no) AS trained
                FROM training_records tr
                JOIN employees e ON tr.emp_no = e.emp_no
                WHERE e.department_id IN ({placeholders})
            """, accessible_dept_ids)
        else:
            cur.execute("SELECT COUNT(DISTINCT emp_no) AS trained FROM training_records")
        trained_count = cur.fetchone()['trained']
        training_coverage = round(trained_count / max(employee_count, 1) * 100, 1)
    except Exception:
        training_coverage = 0

    # 4. 风险预警数量（安全检查中未整改的记录）
    try:
        cur.execute("""
            SELECT COUNT(*) AS cnt FROM safety_inspection_records
            WHERE rectification_status IS NULL
               OR rectification_status = ''
               OR rectification_status = '未整改'
        """)
        risk_count = cur.fetchone()['cnt']
    except Exception:
        risk_count = 0

    dashboard_stats = {
        'employee_count': f"{employee_count:,}",
        'dept_count': dept_count,
        'training_coverage': f"{training_coverage}%",
        'risk_count': risk_count,
    }

    feature_cards = [
        {
            "title": "人员管理",
            "description": "管理员工基础档案信息，支持增删改查。",
            "endpoint": "personnel.index",
            "icon": "fas fa-users",
        },
        {
            "title": "数据分析",
            "description": "多维度的员工数据交叉分析。",
            "endpoint": "personnel.analytics",
            "icon": "fas fa-chart-pie",
        },
        {
            "title": "能力画像",
            "description": "查看员工个人综合能力雷达图。",
            "endpoint": "personnel.capability_profile",
            "icon": "fas fa-user-circle",
        },
        {
            "title": "人才九宫格",
            "description": "基于绩效和潜力的九宫格人才分布。",
            "endpoint": "personnel.page_nine_grid",
            "icon": "fas fa-th",
        },
        {
            "title": "风险挖掘",
            "description": "挖掘潜在的人员风险因素。",
            "endpoint": "personnel.risk_mining_page",
            "icon": "fas fa-search",
        },
    ]
    return render_template(
        "personnel_dashboard.html",
        title=f"人员工作台 | {APP_TITLE}",
        feature_cards=feature_cards,
        stats=dashboard_stats,
    )


@personnel_bp.route('/analytics')
@login_required
def analytics():
    """人员数据分析页面"""
    return render_template(
        "personnel_analytics.html",
        title=f"人员数据分析 | {APP_TITLE}"
    )


@personnel_bp.route('/api/analytics-data')
@login_required
def api_analytics_data():
    """获取人员分析数据API"""
    rows = list_personnel()

    # 岗位筛选：只统计电客车司机，排除副队长和队长
    def is_driver(row):
        position = (row.get("position") or "").strip()
        # 排除副队长和队长
        if "队长" in position:
            return False
        # 只要包含"司机"就算
        return "司机" in position

    # 除了政治面貌统计，其他都只统计司机
    driver_rows = [row for row in rows if is_driver(row)]

    # 1. 安全风险等级分布 - 按入司后单独驾驶年限分级
    risk_levels = {"新手(<1年)": 0, "成长(1-3年)": 0, "熟练(3-5年)": 0, "资深(≥5年)": 0, "未知": 0}
    for row in driver_rows:
        solo_years = row.get("solo_driving_years")
        if solo_years is None:
            risk_levels["未知"] += 1
        elif solo_years < 1:
            risk_levels["新手(<1年)"] += 1
        elif 1 <= solo_years < 3:
            risk_levels["成长(1-3年)"] += 1
        elif 3 <= solo_years < 5:
            risk_levels["熟练(3-5年)"] += 1
        else:
            risk_levels["资深(≥5年)"] += 1

    # 2. 部门战力雷达图 - 各部门的平均司龄、驾龄、取证年限（只统计司机）
    # 获取当前用户可访问的部门列表
    accessible_depts = get_accessible_departments()

    # 获取所有底层部门（没有子部门的部门）
    conn = get_db()
    cur = conn.cursor()

    # 找出所有可访问部门中的底层部门
    accessible_dept_ids = [dept['id'] for dept in accessible_depts]
    if not accessible_dept_ids:
        team_power = []
    else:
        # 查询每个部门是否有子部门
        placeholders = ','.join(['%s'] * len(accessible_dept_ids))
        cur.execute(f"""
            SELECT d.id, d.name, d.level,
                   CASE
                       WHEN EXISTS(SELECT 1 FROM departments child WHERE child.parent_id = d.id)
                       THEN 1 ELSE 0
                   END as has_children
            FROM departments d
            WHERE d.id IN ({placeholders})
            ORDER BY d.level, d.name
        """, accessible_dept_ids)

        dept_info = {row['id']: dict(row) for row in cur.fetchall()}

        # 对于最底层用户，只显示自己部门；对于上级用户，显示所有下级底层部门
        user_dept_info = get_user_department()
        if user_dept_info and user_dept_info['department_id']:
            user_dept_id = user_dept_info['department_id']
            # 检查用户部门是否是底层部门
            if user_dept_id in dept_info and dept_info[user_dept_id]['has_children'] == 0:
                # 用户是底层部门，只显示自己部门
                display_dept_ids = [user_dept_id]
            else:
                # 用户是上级部门，显示所有可访问的底层部门
                display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]
        else:
            # 管理员或无部门用户，显示所有底层部门
            display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]

        # 按部门统计司机数据
        dept_stats = {}
        for row in driver_rows:
            dept_id = row.get("department_id")
            if dept_id not in display_dept_ids:
                continue

            if dept_id not in dept_stats:
                dept_stats[dept_id] = {
                    "name": dept_info.get(dept_id, {}).get('name', '未知部门'),
                    "tenure_years": [],
                    "solo_driving_years": [],
                    "certification_years": []
                }

            if row.get("tenure_years") is not None:
                dept_stats[dept_id]["tenure_years"].append(row["tenure_years"])
            if row.get("solo_driving_years") is not None:
                dept_stats[dept_id]["solo_driving_years"].append(row["solo_driving_years"])
            if row.get("certification_years") is not None:
                dept_stats[dept_id]["certification_years"].append(row["certification_years"])

        team_power = []
        for dept_id, stats in dept_stats.items():
            avg_tenure = sum(stats["tenure_years"]) / len(stats["tenure_years"]) if stats["tenure_years"] else 0
            avg_solo = sum(stats["solo_driving_years"]) / len(stats["solo_driving_years"]) if stats["solo_driving_years"] else 0
            avg_cert = sum(stats["certification_years"]) / len(stats["certification_years"]) if stats["certification_years"] else 0

            team_power.append({
                "team": stats["name"],
                "avg_tenure": round(avg_tenure, 1),
                "avg_solo": round(avg_solo, 1),
                "avg_cert": round(avg_cert, 1),
                "member_count": len([r for r in driver_rows if r.get("department_id") == dept_id])
            })

    # 3. 经验溢出分析 - 散点图数据（只统计司机）
    experience_scatter = []
    for row in driver_rows:
        cert_years = row.get("certification_years")
        solo_years = row.get("solo_driving_years")
        if cert_years is not None and solo_years is not None:
            experience_scatter.append({
                "name": row.get("name"),
                "emp_no": row.get("emp_no"),
                "cert_years": round(cert_years, 1),
                "solo_years": round(solo_years, 1),
                # 分类：准师傅(取证久但单驾短)、资深师傅(两项都高)、新手
                "category": _categorize_experience(cert_years, solo_years)
            })

    # 4. 排班压力预警 - 籍贯分布（只统计司机）+ 政治面貌统计（统计所有人）
    hometown_stats = {}
    political_stats = {"中共党员": 0, "中共预备党员": 0, "共青团员": 0, "群众": 0, "其它": 0}

    # 籍贯统计只统计司机
    for row in driver_rows:
        hometown = row.get("hometown") or "未填写"
        # 河南省内细分到市/县，省外只显示省份
        location = _extract_location(hometown)
        hometown_stats[location] = hometown_stats.get(location, 0) + 1

    # 政治面貌统计所有人员
    for row in rows:
        political = row.get("political_status") or "未填写"
        if political in political_stats:
            political_stats[political] += 1
        else:
            political_stats["其它"] += 1

    # 5. 职业稳定性分析 - 司龄 vs 工龄散点图（只统计司机）
    stability_scatter = []
    for row in driver_rows:
        tenure = row.get("tenure_years")
        working = row.get("working_years")
        if tenure is not None and working is not None:
            stability_scatter.append({
                "name": row.get("name"),
                "emp_no": row.get("emp_no"),
                "tenure": round(tenure, 1),
                "working": round(working, 1),
                # 分类：应届入职、社招新员工、社招老员工
                "category": _categorize_stability(tenure, working)
            })

    return jsonify({
        "risk_distribution": risk_levels,
        "team_power": team_power,
        "experience_scatter": experience_scatter,
        "hometown_stats": hometown_stats,
        "political_stats": political_stats,
        "stability_scatter": stability_scatter,
        "total_count": len(rows),
        "driver_count": len(driver_rows)
    })


def _categorize_experience(cert_years: float, solo_years: float) -> str:
    """分类经验等级"""
    if cert_years >= 5 and solo_years < 3:
        return "准师傅"  # 取证很久但单驾时间较短
    elif cert_years >= 5 and solo_years >= 5:
        return "资深师傅"  # 两项指标都高
    elif cert_years < 2:
        return "新手"
    else:
        return "普通"


def _categorize_stability(tenure: float, working: float) -> str:
    """分类职业稳定性

    Args:
        tenure: 司龄（在本单位工作年限）
        working: 工龄（总工作年限）

    Returns:
        分类标签：应届入职、社招(新)、社招(老)
    """
    work_exp_diff = working - tenure  # 入职前的工作经验

    if work_exp_diff < 1:
        # 工龄和司龄相近，基本是应届生或毕业后很快入职
        return "应届入职"
    elif tenure < 3:
        # 有工作经验，但在本单位时间不长
        return "社招(新)"
    else:
        # 有工作经验，且在本单位时间较长
        return "社招(老)"


def _extract_location(hometown: str) -> str:
    """提取地域信息

    河南省内细分到市/县，其他省份只显示省外或省份名称

    Args:
        hometown: 籍贯字符串，如"河南郑州"、"河南省洛阳市"、"江苏南京"等

    Returns:
        地域标签：河南省内返回市/县名，省外返回省份名或"省外"
    """
    if not hometown or hometown == "未填写":
        return "未填写"

    hometown = hometown.strip()

    # 河南省内的地级市和县
    henan_cities = [
        "郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁",
        "新乡", "焦作", "濮阳", "许昌", "漯河", "三门峡",
        "南阳", "商丘", "信阳", "周口", "驻马店", "济源"
    ]

    # 常见县级市/县（可根据实际情况扩展）
    henan_counties = [
        "巩义", "荥阳", "新密", "新郑", "登封", "中牟",
        "兰考", "杞县", "通许", "尉氏", "偃师", "孟津",
        "新安", "栾川", "嵩县", "汝阳", "宜阳", "洛宁",
        "伊川", "汝州", "舞钢", "林州", "卫辉", "辉县",
        "沁阳", "孟州", "禹州", "长葛", "义马", "灵宝",
        "永城", "项城", "邓州", "固始", "鹿邑", "新蔡"
    ]

    # 检查是否为河南省内
    is_henan = False
    if "河南" in hometown:
        is_henan = True
    else:
        # 如果没有明确写"河南"，但包含河南的市/县名，也认为是河南
        for city in henan_cities + henan_counties:
            if city in hometown:
                is_henan = True
                break

    if is_henan:
        # 河南省内，提取市/县名
        # 优先匹配县级市/县（更具体）
        for county in henan_counties:
            if county in hometown:
                return f"河南·{county}"

        # 再匹配地级市
        for city in henan_cities:
            if city in hometown:
                return f"河南·{city}"

        # 如果只写了"河南"，返回"河南·未详"
        return "河南·未详"

    else:
        # 非河南省，提取省份
        provinces = [
            "北京", "天津", "上海", "重庆",
            "河北", "山西", "辽宁", "吉林", "黑龙江",
            "江苏", "浙江", "安徽", "福建", "江西", "山东",
            "湖北", "湖南", "广东", "海南",
            "四川", "贵州", "云南", "陕西", "甘肃",
            "青海", "台湾", "内蒙古", "广西", "西藏",
            "宁夏", "新疆", "香港", "澳门"
        ]

        for province in provinces:
            if hometown.startswith(province) or province in hometown:
                return f"省外·{province}"

        # 如果无法识别，返回"省外·其他"
        return "省外·其他"


# ==================== 个人综合能力画像 API ====================

@personnel_bp.route('/capability-profile')
@login_required
def capability_profile():
    """个人综合能力画像页面"""
    return render_template('personnel_capability_profile.html', title='个人综合能力画像')


@personnel_bp.route('/api/key-personnel-config')
@login_required
def api_key_personnel_config():
    """API: 获取关键人员配置参数（供前端动态显示使用）"""
    from services.algorithm_config_service import AlgorithmConfigService

    try:
        algo_config = AlgorithmConfigService.get_active_config()
        key_personnel_config = algo_config.get('key_personnel', {})

        return jsonify({
            'success': True,
            'config': {
                'comprehensive_threshold': key_personnel_config.get('comprehensive_threshold', 75),
                'monthly_violation_threshold': key_personnel_config.get('monthly_violation_threshold', 3)
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'config': {
                'comprehensive_threshold': 75,
                'monthly_violation_threshold': 3
            }
        })


@personnel_bp.route('/nine-grid')
@login_required
def page_nine_grid():
    """人才九宫格页面"""
    return render_template('personnel_nine_grid.html')


@personnel_bp.route('/api/departments')
@login_required
def api_departments():
    """API: 获取可访问部门列表"""
    return jsonify(get_accessible_departments())


@personnel_bp.route('/api/nine-grid-data')
@login_required
def api_nine_grid_data():
    """API: 获取九宫格数据"""
    from datetime import datetime
    
    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department_filter = request.args.get('department_id')

    # 默认当月
    if not start_date and not end_date:
        current_month = datetime.now().strftime('%Y-%m')
        start_date = current_month
        end_date = current_month

    # 读取配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    
    rows = list_personnel()
    
    # 筛选
    if department_filter:
        try:
            dept_id_filter = int(department_filter)
            rows = [r for r in rows if r.get('department_id') == dept_id_filter]
        except ValueError:
            pass

    data = []
    
    # 获取权重配置
    score_weights = algo_config['comprehensive']['score_weights']
    nine_grid_weights = algo_config.get('nine_grid', {}).get('y_axis_weights', {
        'stability': 0.4,
        'learning': 0.6,
    })
    
    # 三维分权重归一化（去除稳定性和学习能力后的相对权重）
    w_perf = score_weights.get('performance', 35)
    w_safe = score_weights.get('safety', 30)
    w_train = score_weights.get('training', 20)
    w_x_total = w_perf + w_safe + w_train
    if w_x_total <= 0: w_x_total = 1

    for row in rows:
        try:
            scores = _calculate_single_employee_score(row, start_date, end_date, algo_config, cur)
            
            # 计算X轴（三维综合分）
            x_raw = (scores['performance'] * w_perf + 
                     scores['safety'] * w_safe + 
                     scores['training'] * w_train)
            x_score = round(x_raw / w_x_total * 100, 1) if w_x_total > 1 else round(x_raw, 1) # 假设配置是百分比整数(35)或小数(0.35)
            # 如果配置是小数(0.35)，w_x_total=0.85。 x_raw = P*0.35 ... -> x_raw / 0.85 * 100? No.
            # 如果配置是0.35，x_raw 是加权后的分。
            # 如果 P=100, x_raw = 35. 35/0.85 approx 41. ???
            # 综合分通常是加权和。
            # x_score 应该是满分100。
            # 如 P=100, S=100, T=100. x_raw = 35+30+20 = 85.
            # 那么 x_score = 85 / 0.85 = 100. Correct.
            if w_x_total < 5: # 检测是否为小数配置 (e.g. 0.35)
                # 系数本来就是小数，不需要 * 100 ?
                # 0.35+0.30+0.20 = 0.85
                # raw = 100*0.35 + ... = 85
                # 85 / 0.85 = 100.
                x_score = round(x_raw / w_x_total, 1)
            else:
                # 配置是整数 (35, 30, 20) -> Sum 85
                # raw = 100*35 + ... = 8500
                # 8500 / 85 = 100 ?
                # 通常 weighted average = sum(val * weight) / sum(weights)
                x_score = round(x_raw / w_x_total, 1)

            
            # 计算Y轴（稳定 + 学习）
            y_w_stab = nine_grid_weights.get('stability', 0.4) 
            y_w_learn = nine_grid_weights.get('learning', 0.6)
            y_total = y_w_stab + y_w_learn
            if y_total <= 0: y_total = 1
            
            y_raw = (scores['stability'] * y_w_stab + scores['learning'] * y_w_learn)
            y_score = round(y_raw / y_total, 1)

            # 判定九宫格位置 (3x3)
            # 简单的三分法：<75(Low), 75-90(Mid), >=90(High)
            # 可以根据实际需求调整阈值
            
            x_level = 1
            if x_score >= 90: x_level = 3
            elif x_score >= 75: x_level = 2
            
            y_level = 1
            if y_score >= 90: y_level = 3
            elif y_score >= 75: y_level = 2
            
            # 映射到行和列
            # 界面布局：
            # Row 1 (High Y): Cell 1-1, 1-2, 1-3
            # Row 2 (Mid Y)
            # Row 3 (Low Y)
            # Col 1 (Low X), Col 2 (Mid X), Col 3 (High X)
            
            grid_row = 4 - y_level # 3->1, 2->2, 1->3
            grid_col = x_level     # 1->1, 2->2, 3->3
            
            data.append({
                'emp_no': row.get('emp_no'),
                'name': row.get('name'),
                'department_name': row.get('department_name'),
                'x_score': x_score,
                'y_score': y_score,
                'grid_row': grid_row,
                'grid_col': grid_col,
                'details': scores
            })
        except Exception as e:
            current_app.logger.error(f"Error calculating score for {row.get('name')}: {e}")
            continue

    conn.close()
    return jsonify(data)


@personnel_bp.route('/nine-grid/export')
@login_required
def export_nine_grid():
    """导出人才九宫格数据"""
    from datetime import datetime
    
    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department_filter = request.args.get('department_id')

    # 默认当月
    if not start_date and not end_date:
        current_month = datetime.now().strftime('%Y-%m')
        start_date = current_month
        end_date = current_month

    # 读取配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    
    rows = list_personnel()
    
    # 筛选
    if department_filter:
        try:
            dept_id_filter = int(department_filter)
            rows = [r for r in rows if r.get('department_id') == dept_id_filter]
        except ValueError:
            pass

    # 获取权重配置
    score_weights = algo_config['comprehensive']['score_weights']
    nine_grid_weights = algo_config.get('nine_grid', {}).get('y_axis_weights', {
        'stability': 0.4,
        'learning': 0.6,
    })
    
    # 三维分权重归一化
    w_perf = score_weights.get('performance', 35)
    w_safe = score_weights.get('safety', 30)
    w_train = score_weights.get('training', 20)
    w_x_total = w_perf + w_safe + w_train
    if w_x_total <= 0: w_x_total = 1

    # 九宫格标签映射
    grid_labels = {
        (1, 1): '培养对象',
        (1, 2): '潜力新星',
        (1, 3): '明星员工',
        (2, 1): '改善对象',
        (2, 2): '中坚力量',
        (2, 3): '骨干员工',
        (3, 1): '问题员工',
        (3, 2): '需关注',
        (3, 3): '待观察稳定'
    }

    # 汇总数据：按九宫格位置统计
    summary_data = {}
    detail_data = []
    
    for row in rows:
        try:
            scores = _calculate_single_employee_score(row, start_date, end_date, algo_config, cur)
            
            # 计算X轴（三维综合分）
            x_raw = (scores['performance'] * w_perf + 
                     scores['safety'] * w_safe + 
                     scores['training'] * w_train)
            if w_x_total < 5:
                x_score = round(x_raw / w_x_total, 1)
            else:
                x_score = round(x_raw / w_x_total, 1)
            
            # 计算Y轴（稳定 + 学习）
            y_w_stab = nine_grid_weights.get('stability', 0.4) 
            y_w_learn = nine_grid_weights.get('learning', 0.6)
            y_total = y_w_stab + y_w_learn
            if y_total <= 0: y_total = 1
            
            y_raw = (scores['stability'] * y_w_stab + scores['learning'] * y_w_learn)
            y_score = round(y_raw / y_total, 1)

            # 判定九宫格位置
            x_level = 1
            if x_score >= 90: x_level = 3
            elif x_score >= 75: x_level = 2
            
            y_level = 1
            if y_score >= 90: y_level = 3
            elif y_score >= 75: y_level = 2
            
            grid_row = 4 - y_level
            grid_col = x_level
            
            # 汇总统计
            grid_key = (grid_row, grid_col)
            if grid_key not in summary_data:
                summary_data[grid_key] = []
            summary_data[grid_key].append(row.get('name'))
            
            # 明细数据
            detail_data.append({
                'emp_no': row.get('emp_no'),
                'name': row.get('name'),
                'department_name': row.get('department_name'),
                'x_score': x_score,
                'y_score': y_score,
                'grid_label': grid_labels.get(grid_key, ''),
                'performance': scores['performance'],
                'safety': scores['safety'],
                'training': scores['training'],
                'stability': scores['stability'],
                'learning': scores['learning']
            })
        except Exception as e:
            current_app.logger.error(f"Error calculating score for {row.get('name')}: {e}")
            continue

    conn.close()

    # 创建Excel工作簿
    wb = Workbook()
    
    # 第一个工作表：汇总（9个单元格）
    ws_summary = wb.active
    ws_summary.title = '汇总'
    
    # 设置汇总表的标题和数据
    ws_summary['A1'] = '九宫格汇总'
    ws_summary['A1'].font = ws_summary['A1'].font.copy()
    
    # 按照九宫格布局填充汇总数据
    # 行标题（Y轴）
    ws_summary['A2'] = '高'
    ws_summary['A3'] = '中'
    ws_summary['A4'] = '低'
    
    # 列标题（X轴）
    ws_summary['B1'] = '低'
    ws_summary['C1'] = '中'
    ws_summary['D1'] = '高'
    
    # 填充9个单元格
    for row_idx in range(1, 4):
        for col_idx in range(1, 4):
            grid_key = (row_idx, col_idx)
            cell_row = row_idx + 1
            cell_col = col_idx + 1
            cell = ws_summary.cell(row=cell_row, column=cell_col)
            
            label = grid_labels.get(grid_key, '')
            count = len(summary_data.get(grid_key, []))
            names = ', '.join(summary_data.get(grid_key, []))
            
            cell.value = f"{label}\n({count}人)\n{names}"
            cell.alignment = cell.alignment.copy()
    
    # 第二个工作表：明细
    ws_detail = wb.create_sheet('明细')
    ws_detail.append(['工号', '姓名', '部门', 'X分数', 'Y分数', '九宫格位置', '绩效', '安全', '培训', '稳定性', '学习能力'])
    
    for item in detail_data:
        ws_detail.append([
            item['emp_no'],
            item['name'],
            item['department_name'],
            item['x_score'],
            item['y_score'],
            item['grid_label'],
            item['performance'],
            item['safety'],
            item['training'],
            item['stability'],
            item['learning']
        ])
    
    # 保存文件
    export_filename = f"人才九宫格_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    xlsx_path = os.path.join(EXPORT_DIR, export_filename)
    wb.save(xlsx_path)
    
    return send_file(xlsx_path, as_attachment=True, download_name=export_filename)


def _calculate_single_employee_score(row, start_date, end_date, algo_config, cur):
    """
    辅助函数：计算单个员工的各项评分
    """
    from blueprints.safety import extract_score_from_assessment
    from datetime import datetime, timedelta
    import calendar

    emp_no = row.get('emp_no')
    emp_name = row.get('name')
    dept_id = row.get('department_id')
    entry_date = row.get('entry_date')

    # 计算取证年限
    cert_date = row.get('certification_date')
    cert_years = calculate_years_from_date(cert_date) if cert_date else None

    # 1. 培训能力
    training_query = "SELECT score, is_qualified, is_disqualified, training_date FROM training_records WHERE emp_no = %s"
    training_params = [emp_no]

    if start_date:
        training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') >= %s"
        training_params.append(start_date)
    if end_date:
        training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') <= %s"
        training_params.append(end_date)

    training_query += " ORDER BY training_date ASC"
    cur.execute(training_query, training_params)
    training_records_list = cur.fetchall()

    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            end_year, end_month = int(end_date.split('-')[0]), int(end_date.split('-')[1])
            last_day = calendar.monthrange(end_year, end_month)[1]
            end_dt = end_dt.replace(day=last_day)
            duration_days = max(1, (end_dt - start_dt).days + 1)
        except:
            duration_days = 30
    else:
        duration_days = 30

    training_result = calculate_training_score_with_penalty(training_records_list, duration_days, cert_years, algo_config)
    training_score = training_result['radar_score']

    # 2. 安全意识
    safety_query = "SELECT assessment, inspection_date FROM safety_inspection_records WHERE inspected_person = %s"
    safety_params = [emp_name]

    if start_date:
        safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') >= %s"
        safety_params.append(start_date)
    if end_date:
        safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') <= %s"
        safety_params.append(end_date)

    safety_query += " ORDER BY inspection_date ASC"
    cur.execute(safety_query, safety_params)
    safety_rows = cur.fetchall()

    violations_list = []
    for s_row in safety_rows:
        score = extract_score_from_assessment(s_row['assessment'])
        if score > 0:
            violations_list.append(float(score))

    months_active = 1
    if start_date:
        try:
            start = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end = datetime.strptime(end_date + '-01', '%Y-%m-%d') if end_date else datetime.now()
            if not end_date:
                # 只指定开始，到今天
                months_active = max(1, int((end - start).days / 30) + 1)
            else:
                months_active = max(1, int((end - start).days / 30) + 1)
        except:
            months_active = 1
    elif entry_date:
        try:
             # entry_date可能是date对象或str
            entry = entry_date if isinstance(entry_date, datetime) or hasattr(entry_date, 'year') else datetime.strptime(str(entry_date), '%Y-%m-%d')
            months_active = max(1, int((datetime.now() - entry).days / 30))
        except:
            months_active = 1

    safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
    safety_score = safety_result['final_score']

    # 3. 工作绩效
    is_monthly = (start_date == end_date) if start_date and end_date else True
    perf_query = "SELECT score, grade, year, month FROM performance_records WHERE emp_no = %s"
    perf_params = [emp_no]

    if start_date:
        perf_query += " AND CAST(CONCAT(year, '-', LPAD(month, 2, '0')) AS CHAR) >= %s"
        perf_params.append(start_date)
    if end_date:
        perf_query += " AND CAST(CONCAT(year, '-', LPAD(month, 2, '0')) AS CHAR) <= %s"
        perf_params.append(end_date)

    perf_query += " ORDER BY year, month"
    cur.execute(perf_query, perf_params)
    perf_rows = cur.fetchall()

    if perf_rows:
        if is_monthly and len(perf_rows) == 1:
            perf_row = perf_rows[0]
            raw_score = float(perf_row['score']) if perf_row['score'] else 95
            grade = perf_row['grade'] if perf_row['grade'] else 'B+'
            perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
            performance_score = perf_result['radar_value']
        else:
            grade_list = [row['grade'] if row['grade'] else 'B+' for row in perf_rows]
            grade_dates = [f"{row['year']}-{row['month']:02d}" for row in perf_rows]
            perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
            performance_score = perf_result['radar_value']
    else:
        performance_score = 0

    # 4. 学习能力
    is_long_term = False
    if start_date and end_date and start_date != end_date:
        is_long_term = True
    
    learning_score = 0
    calculated_learning = False
    
    if is_long_term:
        try:
            monthly_counts = {}
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            
            # 预查上月数据
            pre_period_month = (start_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
            pre_period_count = 0
            try:
                cur.execute("SELECT assessment FROM safety_inspection_records WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s", [emp_name, pre_period_month])
                pre_rows = cur.fetchall()
                pre_period_count = sum(1 for r in pre_rows if extract_score_from_assessment(r['assessment']) > 0)
            except:
                pre_period_count = None # None means no record/new employee

            curr = start_dt
            months_seq = []
            while curr <= end_dt:
                m_str = curr.strftime('%Y-%m')
                monthly_counts[m_str] = 0
                months_seq.append(m_str)
                curr = (curr.replace(day=1) + timedelta(days=32)).replace(day=1)
            
            # 填充违规计数
            for row in safety_rows:
                if extract_score_from_assessment(row['assessment']) > 0:
                    insp_date = row['inspection_date']
                    m_str = insp_date.strftime('%Y-%m') if hasattr(insp_date, 'strftime') else str(insp_date)[:7]
                    
                    if m_str in monthly_counts:
                        monthly_counts[m_str] += 1
            
            # 班组平均
            period_group_avg = 1.0 
            if dept_id:
                try:
                    cur.execute("SELECT COUNT(*) / COUNT(DISTINCT e.name) / %s as avg_viol FROM safety_inspection_records s JOIN employees e ON s.inspected_person = e.name WHERE e.department_id = %s AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') >= %s AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') <= %s", [max(1, len(months_seq)), dept_id, start_date, end_date])
                    avg_res = cur.fetchone()
                    if avg_res and avg_res['avg_viol']:
                         period_group_avg = float(avg_res['avg_viol'])
                except:
                    pass

            monthly_scores = []
            last_violations = pre_period_count
            
            for m_str in months_seq:
                curr_viol = monthly_counts[m_str]
                res = calculate_learning_ability_new(curr_viol, last_violations, period_group_avg, algo_config)
                monthly_scores.append(res['learning_score'])
                last_violations = curr_viol
            
            # 加权平均
            total_weight = 0
            weighted_sum = 0
            time_decay_rate = algo_config.get('learning_new', {}).get('time_decay_rate', 0.2)
            for i, score in enumerate(monthly_scores):
                weight = 1.0 + (i * time_decay_rate) 
                weighted_sum += score * weight
                total_weight += weight
            
            learning_score = weighted_sum / total_weight if total_weight > 0 else 0
            calculated_learning = True
            
        except Exception as e:
            # Fallback
            pass
            
    if not calculated_learning:
        # 单月/Short term logic
        current_violations = 0
        if end_date:
            target_month = end_date
        else:
            target_month = datetime.now().strftime('%Y-%m')
            
        current_violations = sum(1 for r in list(filter(lambda x: x['inspection_date'].strftime('%Y-%m') == target_month, safety_rows)) if extract_score_from_assessment(x['assessment']) > 0)

        previous_violations = None
        if start_date:
            try:
                curr_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                prev_month = (curr_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
                cur.execute("SELECT assessment FROM safety_inspection_records WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s", [emp_name, prev_month])
                prev_rows = cur.fetchall()
                previous_violations = sum(1 for r in prev_rows if extract_score_from_assessment(r['assessment']) > 0)
            except:
                pass
        
        group_avg = 1.0
        if dept_id and start_date:
            try:
                cur.execute("SELECT COUNT(*) / COUNT(DISTINCT e.name) as avg_viol FROM safety_inspection_records s JOIN employees e ON s.inspected_person = e.name WHERE e.department_id = %s AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') = %s", [dept_id, start_date])
                res = cur.fetchone()
                if res and res['avg_viol']: group_avg = float(res['avg_viol'])
            except:
                pass
                
        l_res = calculate_learning_ability_new(current_violations, previous_violations, group_avg, algo_config)
        learning_score = l_res['learning_score']

    # 5. 稳定性（波动型）
    stability_result = calculate_stability_for_employee(
        emp_name,
        start_date,
        end_date,
        algo_config,
        cur,
        safety_score_for_tip=safety_score
    )
    stability_score = stability_result.get('stability_score', 50)
    if not isinstance(stability_score, (int, float)):
        stability_score = 50

    return {
        'performance': performance_score,
        'safety': safety_score,
        'training': training_score,
        'learning': round(float(learning_score), 1),
        'stability': purity_score(stability_score) # Just ensure float/int
    }

def purity_score(s):
    if isinstance(s, (int, float)): return round(float(s), 1)
    return 50.0



@personnel_bp.route('/api/students-list')
@login_required
def api_students_list():
    """API: 获取人员列表及综合评分（带权限过滤和关键人员标记）"""
    from datetime import datetime
    from blueprints.safety import extract_score_from_assessment

    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    start_date = request.args.get('start_date')  # 格式：YYYY-MM
    end_date = request.args.get('end_date')      # 格式：YYYY-MM
    department_filter = request.args.get('department')
    position_filter = request.args.get('position')

    # 如果没有指定日期筛选，默认使用当月（1号到今天）
    if not start_date and not end_date:
        current_month = datetime.now().strftime('%Y-%m')
        start_date = current_month
        end_date = current_month

    # 获取当前月份（用于关键人员标记）
    current_month = datetime.now().strftime('%Y-%m')

    # 读取算法配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    score_weights = algo_config['comprehensive']['score_weights']
    key_personnel_config = algo_config['key_personnel']

    # 安全获取字典值的辅助函数
    def safe_get(obj, key, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        else:
            try:
                return obj[key] if obj[key] is not None else default
            except (KeyError, IndexError):
                return default

    # 使用现有的 list_personnel() 函数获取权限过滤后的人员列表
    rows = list_personnel()

    # 应用部门和岗位筛选
    if department_filter:
        rows = [r for r in rows if safe_get(r, 'department_name') == department_filter]

    if position_filter:
        position_filter_lower = position_filter.lower()
        rows = [r for r in rows if position_filter_lower in (safe_get(r, 'position') or '').lower()]

    students = []
    for row in rows:
        emp_no = safe_get(row, 'emp_no')
        emp_name = safe_get(row, 'name')
        dept_id = safe_get(row, 'department_id')

        # 计算取证年限（用于培训和稳定性算法）
        cert_date = safe_get(row, 'certification_date')
        cert_years = calculate_years_from_date(cert_date) if cert_date else None

        # 获取部门名称
        if dept_id:
            cur.execute("SELECT name FROM departments WHERE id = %s", (dept_id,))
            dept_row = cur.fetchone()
            dept_name = dept_row['name'] if dept_row else None
        else:
            dept_name = None

        # 1. 培训能力（使用高级评分算法，应用日期筛选）
        training_query = """
            SELECT score, is_qualified, is_disqualified, training_date
            FROM training_records
            WHERE emp_no = %s
        """
        training_params = [emp_no]

        if start_date:
            training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') >= %s"
            training_params.append(start_date)

        if end_date:
            training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') <= %s"
            training_params.append(end_date)

        training_query += " ORDER BY training_date ASC"
        cur.execute(training_query, training_params)
        training_records_list = cur.fetchall()

        # 计算统计周期天数
        if start_date and end_date and start_date == end_date:
            duration_days = 30
        elif start_date and end_date:
            try:
                start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
                import calendar
                end_year, end_month = int(end_date.split('-')[0]), int(end_date.split('-')[1])
                last_day = calendar.monthrange(end_year, end_month)[1]
                end_dt = end_dt.replace(day=last_day)
                duration_days = max(1, (end_dt - start_dt).days + 1)
            except:
                duration_days = 30
        else:
            duration_days = 30

        # 使用新的评分算法
        training_result = calculate_training_score_with_penalty(training_records_list, duration_days, cert_years, algo_config)
        training_score = training_result['radar_score']

        # 2. 安全意识（使用双轨评分模型）
        # 构建日期筛选条件
        safety_query = """
            SELECT assessment, inspection_date
            FROM safety_inspection_records
            WHERE inspected_person = %s
        """
        safety_params = [emp_name]

        if start_date:
            safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') >= %s"
            safety_params.append(start_date)

        if end_date:
            safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') <= %s"
            safety_params.append(end_date)

        safety_query += " ORDER BY inspection_date ASC"
        cur.execute(safety_query, safety_params)
        safety_rows = cur.fetchall()

        # 收集所有违规扣分
        violations_list = []
        for s_row in safety_rows:
            assessment = s_row['assessment']
            score = extract_score_from_assessment(assessment)
            if score > 0:
                violations_list.append(float(score))

        # 计算统计周期月数
        months_active = 1
        if start_date and end_date:
            # 如果指定了日期范围，计算该范围的月数
            try:
                start = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end = datetime.strptime(end_date + '-01', '%Y-%m-%d')
                months_active = max(1, int((end - start).days / 30) + 1)
            except:
                months_active = 1
        elif start_date:
            # 只指定了开始日期，从开始日期到现在
            try:
                start = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                months_active = max(1, int((datetime.now() - start).days / 30) + 1)
            except:
                months_active = 1
        elif entry_date:
            # 没有日期筛选，使用入职以来的月数
            try:
                entry = datetime.strptime(entry_date, '%Y-%m-%d')
                months_active = max(1, int((datetime.now() - entry).days / 30))
            except:
                months_active = 1

        # 使用双轨评分模型
        safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
        safety_score = safety_result['final_score']
        safety_status_color = safety_result['status_color']
        safety_alert_tag = safety_result['alert_tag']

        # 3. 工作绩效（使用双算法系统，应用日期筛选）
        is_monthly = (start_date == end_date) if start_date and end_date else True

        perf_query = """
            SELECT score, grade, year, month
            FROM performance_records
            WHERE emp_no = %s
        """
        perf_params = [emp_no]

        if start_date:
            perf_query += f" AND ({get_year_month_concat()}) >= %s"
            perf_params.append(start_date)

        if end_date:
            perf_query += f" AND ({get_year_month_concat()}) <= %s"
            perf_params.append(end_date)

        perf_query += " ORDER BY year, month"
        cur.execute(perf_query, perf_params)
        perf_rows = cur.fetchall()

        if perf_rows:
            if is_monthly and len(perf_rows) == 1:
                # 月度快照算法
                perf_row = perf_rows[0]
                raw_score = float(perf_row['score']) if perf_row['score'] else 95
                grade = perf_row['grade'] if perf_row['grade'] else 'B+'
                perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
                performance_score = perf_result['radar_value']
            else:
                # 周期加权算法（带时间衰减）
                grade_list = [row['grade'] if row['grade'] else 'B+' for row in perf_rows]
                grade_dates = [f"{row['year']}-{row['month']:02d}" for row in perf_rows]  # 构建日期列表
                perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
                performance_score = perf_result['radar_value']
        else:
            # 没有绩效数据
            performance_score = 0

        # 4. 学习能力评估（新版：基于违规数量变化趋势）
        current_violations = 0
        previous_violations = None
        learning_result = None  # 初始化，避免UnboundLocalError
        
        # 判断是否为长周期（跨月）
        is_long_term = False
        if start_date and end_date and start_date != end_date:
            is_long_term = True
            
        # 初始化班组平均违规数（默认值）
        group_avg_violations = 1.0
            
        if is_long_term:
            # 长周期模式：使用 calculate_learning_ability_longterm 保持与详情页一致
            try:
                # 1. 初始化每月计数
                monthly_counts = {}
                start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')

                # 预先查询周期前一个月的数据（作为第一个月的previous基础）
                pre_period_month = (start_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
                pre_period_count = 0
                try:
                    cur.execute("""
                        SELECT assessment FROM safety_inspection_records
                        WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s
                    """, [emp_name, pre_period_month])
                    pre_rows = cur.fetchall()
                    pre_period_count = sum(1 for r in pre_rows if extract_score_from_assessment(r['assessment']) > 0)
                except:
                    pre_period_count = None # 无数据

                # 构建月份序列
                curr = start_dt
                months_seq = []
                while curr <= end_dt:
                    m_str = curr.strftime('%Y-%m')
                    monthly_counts[m_str] = 0
                    months_seq.append(m_str)
                    curr = (curr.replace(day=1) + timedelta(days=32)).replace(day=1)

                # 2. 填充数据
                for row in safety_rows:
                    if extract_score_from_assessment(row['assessment']) > 0:
                        insp_date = row['inspection_date']
                        m_str = insp_date.strftime('%Y-%m') if hasattr(insp_date, 'strftime') else str(insp_date)[:7]
                        if m_str in monthly_counts:
                            monthly_counts[m_str] += 1

                # 获取班组平均作为参考（使用周期内的整体平均）
                period_group_avg = 1.0
                if dept_id:
                     try:
                        cur.execute("""
                            SELECT COUNT(*) / COUNT(DISTINCT e.name) / %s as avg_viol
                            FROM safety_inspection_records s
                            JOIN employees e ON s.inspected_person = e.name
                            WHERE e.department_id = %s
                            AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') >= %s
                            AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') <= %s
                        """, [max(1, len(months_seq)), dept_id, start_date, end_date])
                        avg_res = cur.fetchone()
                        if avg_res and avg_res['avg_viol']:
                            period_group_avg = float(avg_res['avg_viol'])
                     except:
                        pass

                # 3. 构建违规数量列表（按月份顺序）
                score_list = [monthly_counts[m] for m in months_seq]

                # 4. 调用 calculate_learning_ability_longterm（与详情页一致，包含风险惯性惩罚）
                learning_result = calculate_learning_ability_longterm(
                    score_list=score_list,
                    config=algo_config,
                    group_avg=period_group_avg,
                    initial_prev_viol=pre_period_count
                )

                # 补全部分前端需要的字段
                if learning_result['risk_level'] == 'SAFE':
                    learning_result['trend_type'] = 'safe'
                elif learning_result['risk_level'] in ['HIGH_RISK', 'PRE_ACCIDENT']:
                    learning_result['trend_type'] = 'deterioration'
                elif learning_result['risk_level'] == 'WATCH_LIST':
                    learning_result['trend_type'] = 'yellow_alert'
                else:
                    learning_result['trend_type'] = 'fluctuation'

                # 设置current/previous用于后续兼容（使用最后两个月数据）
                current_violations = monthly_counts[months_seq[-1]]
                if len(months_seq) >= 2:
                    previous_violations = monthly_counts[months_seq[-2]]
                else:
                    previous_violations = pre_period_count

            except Exception as e:
                current_app.logger.error(f": 长周期学习能力计算异常: {e}")
                is_long_term = False # 回退到单月模式


        if not is_long_term:
            # 单月模式：本月 vs 上月
            if end_date:
                # 统计end_date当月的违规数（安全处理 datetime 和 str 类型）
                end_target = end_date
                current_violations = sum(1 for r in safety_rows 
                                       if (r['inspection_date'].strftime('%Y-%m') if hasattr(r['inspection_date'], 'strftime') 
                                           else str(r['inspection_date'])[:7]) == end_target 
                                       and extract_score_from_assessment(r['assessment']) > 0)
            else:
                current_violations = len(violations_list)

            # 获取上月违规数
            if start_date:
                try:
                    current_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                    prev_dt = current_dt.replace(day=1) - timedelta(days=1)
                    prev_month = prev_dt.strftime('%Y-%m')

                    cur.execute("""
                        SELECT assessment FROM safety_inspection_records
                        WHERE inspected_person = %s
                        AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s
                    """, [emp_name, prev_month])
                    prev_rows = cur.fetchall()
                    previous_violations = sum(1 for r in prev_rows if extract_score_from_assessment(r['assessment']) > 0)
                except:
                    previous_violations = None
        
        if not learning_result:
            # 获取班组平均违规数
            group_avg_violations = 1.0
            if dept_id and start_date:
                try:
                    cur.execute("""
                        SELECT COUNT(*) / COUNT(DISTINCT e.name) as avg_viol
                        FROM safety_inspection_records s
                        JOIN employees e ON s.inspected_person = e.name
                        WHERE e.department_id = %s
                        AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') = %s
                    """, [dept_id, start_date])
                    avg_result = cur.fetchone()
                    if avg_result and avg_result['avg_viol']:
                        group_avg_violations = float(avg_result['avg_viol'])
                except:
                    pass
            
            # 调用新算法
            learning_result = calculate_learning_ability_new(
                current_violations=current_violations,
                previous_violations=previous_violations,
                group_avg_violations=group_avg_violations,
                config=algo_config
            )
        learning_score = learning_result['learning_score']

        # 5. 稳定性（波动型）
        stability_result = calculate_stability_for_employee(
            emp_name,
            start_date,
            end_date,
            algo_config,
            cur,
            safety_score_for_tip=safety_score
        )
        stability_score = stability_result.get('stability_score', 50)

        # 异常情况：使用简单计算作为降级方案
        # Note: The new stability algorithm is designed to be robust.
        # If `calculate_stability_score_new` itself fails, it should handle its own defaults.
        # The previous fallback logic for `entry_date` is now less relevant
        # as the new algorithm doesn't primarily rely on `entry_date` for its core calculation.
        # However, if `stability_score` is still not set or is invalid, a final fallback can be applied.
        if stability_score is None:
            current_app.logger.debug(f" [api_students_list-员工{emp_no}]: 稳定性算法返回None，使用默认值50")
            stability_score = 50
        elif not isinstance(stability_score, (int, float)):
            current_app.logger.debug(f" [api_students_list-员工{emp_no}]: 稳定性算法返回非数值，使用默认值50")
            stability_score = 50

        # 综合评分（加权平均 - 使用配置权重）
        comprehensive_score = round(
            performance_score * score_weights['performance'] +
            safety_score * score_weights['safety'] +
            training_score * score_weights['training'] +
            stability_score * score_weights['stability'] +
            learning_score * score_weights['learning'],
            1
        )

        # 判断是否为关键人员（基于筛选日期范围）（使用配置阈值）
        # 复用已计算的违规数据和月数，避免重复查询
        import math
        violation_count = len(violations_list)
        avg_freq = math.ceil(violation_count / months_active) if months_active > 0 else 0

        is_key_personnel = (comprehensive_score < key_personnel_config['comprehensive_threshold']) or (avg_freq >= key_personnel_config['monthly_violation_threshold'])

        students.append({
            'emp_no': emp_no,
            'name': emp_name,
            'department_name': dept_name,
            'position': safe_get(row, 'position'),
            'comprehensive_score': comprehensive_score,
            'is_key_personnel': bool(is_key_personnel),  # 显式转换为JSON兼容的布尔值
            'safety_status_color': safety_status_color,
            'safety_alert_tag': safety_alert_tag
        })

    # 按综合分升序排序
    students.sort(key=lambda x: x['comprehensive_score'])

    return jsonify(students)


@personnel_bp.route('/api/comprehensive-profile/<emp_no>')
@login_required
def api_comprehensive_profile(emp_no):
    """API: 获取个人综合能力画像（人员+培训+安全+绩效）"""
    from datetime import datetime, timedelta
    from blueprints.safety import extract_score_from_assessment

    # 读取算法配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    score_weights = algo_config['comprehensive']['score_weights']

    conn = get_db()
    cur = conn.cursor()

    # 1. 获取员工基本信息
    cur.execute("""
        SELECT
            name, department_id, position, education, entry_date,
            birth_date, work_start_date, certification_date, solo_driving_date
        FROM employees
        WHERE emp_no = %s
    """, (emp_no,))
    employee = cur.fetchone()

    if not employee:
        return jsonify({'error': '员工不存在'}), 404

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # DictCursor返回字典，使用字典访问方式
    emp_name = employee['name']
    dept_id = employee['department_id']
    position = employee['position']
    education = employee['education']
    entry_date = employee['entry_date']
    birth_date = employee['birth_date']
    work_start_date = employee['work_start_date']
    cert_date = employee['certification_date']
    solo_date = employee['solo_driving_date']

    # 计算各项年限
    working_years = calculate_years_from_date(work_start_date) if work_start_date else None
    tenure_years = calculate_years_from_date(entry_date) if entry_date else None
    cert_years = calculate_years_from_date(cert_date) if cert_date else None
    solo_years = calculate_years_from_date(solo_date) if solo_date else None

    # 获取日期筛选参数（如果有）
    start_date = request.args.get('start_date')  # 格式：YYYY-MM
    end_date = request.args.get('end_date')      # 格式：YYYY-MM

    # DEBUG: 打印接收到的日期参数
    current_app.logger.debug(f" [comprehensive-profile]: 原始参数 - start_date='{start_date}', end_date='{end_date}'")
    current_app.logger.debug(f" [comprehensive-profile]: 参数类型 - start_date type={type(start_date)}, end_date type={type(end_date)}")
    current_app.logger.debug(f" [comprehensive-profile]: 参数布尔值 - bool(start_date)={bool(start_date)}, bool(end_date)={bool(end_date)}")

    # 如果没有指定日期，默认使用当月
    if not start_date and not end_date:
        current_month = datetime.now().strftime('%Y-%m')
        start_date = current_month
        end_date = current_month
        current_app.logger.debug(f" [comprehensive-profile]: 无日期参数，使用默认当月: {current_month}")

    # 2. 培训能力分析（使用高级评分算法 - 包含毒性惩罚和动态年化）
    training_query = """
        SELECT
            score,
            is_qualified,
            is_disqualified,
            training_date
        FROM training_records
        WHERE emp_no = %s
    """
    training_params = [emp_no]

    if start_date:
        training_query += " AND training_date >= %s"
        training_params.append(start_date + '-01')

    if end_date:
        # 计算下个月1号作为上限（开区间 <）
        try:
            curr = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            # 简单处理：加上32天然后设为1号
            next_month = (curr + timedelta(days=32)).replace(day=1)
            training_query += " AND training_date < %s"
            training_params.append(next_month.strftime('%Y-%m-%d'))
        except:
            # 回退方案
            training_query += " AND training_date <= %s"
            training_params.append(end_date + '-31')

    training_query += " ORDER BY training_date ASC"
    cur.execute(training_query, training_params)
    training_records = cur.fetchall()

    # 计算统计周期天数
    if start_date and end_date and start_date == end_date:
        # 单月统计，按30天计算
        duration_days = 30
    elif start_date and end_date:
        # 多月统计，计算实际天数
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            # 计算到月末
            import calendar
            end_year, end_month = int(end_date.split('-')[0]), int(end_date.split('-')[1])
            last_day = calendar.monthrange(end_year, end_month)[1]
            end_dt = end_dt.replace(day=last_day)
            duration_days = max(1, (end_dt - start_dt).days + 1)
        except:
            duration_days = 30
    else:
        # 默认按30天计算
        duration_days = 30

    # 使用新的评分算法
    training_result = calculate_training_score_with_penalty(training_records, duration_days, cert_years, algo_config)
    training_score = training_result['radar_score']
    training_status_color = training_result['status_color']
    training_alert_tag = training_result['alert_tag']
    training_original_score = training_result['original_score']
    training_penalty_coeff = training_result['penalty_coefficient']
    total_training_count = training_result['stats']['total_ops']
    training_fail_count = training_result['stats']['fail_count']

    # 3. 安全能力分析（使用双轨评分模型，应用日期筛选）
    safety_query = """
        SELECT
            inspection_date,
            assessment,
            inspected_person,
            rectifier
        FROM safety_inspection_records
        WHERE (inspected_person = %s OR rectifier = %s)
    """
    safety_params = [emp_name, emp_name]

    if start_date:
        safety_query += " AND inspection_date >= %s"
        safety_params.append(start_date + '-01')

    if end_date:
        try:
            curr = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            next_month = (curr + timedelta(days=32)).replace(day=1)
            safety_query += " AND inspection_date < %s"
            safety_params.append(next_month.strftime('%Y-%m-%d'))
        except:
            safety_query += " AND inspection_date <= %s"
            safety_params.append(end_date + '-31')

    safety_query += " ORDER BY inspection_date ASC"
    cur.execute(safety_query, safety_params)
    
    safety_rows = cur.fetchall()

    violations_list = []
    safety_as_inspector = 0
    safety_as_rectifier = 0

    for row in safety_rows:
        assessment = row['assessment']
        inspected = row['inspected_person']
        rectifier = row['rectifier']
        score = extract_score_from_assessment(assessment)

        if inspected == emp_name and score > 0:
            violations_list.append(float(score))

        if inspected == emp_name:
            safety_as_inspector += 1
        if rectifier == emp_name:
            safety_as_rectifier += 1

    # 计算统计周期月数（使用筛选日期范围的月数）
    months_active = 1
    if start_date and end_date:
        # 如果指定了日期范围，计算该范围的月数
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            months_active = max(1, int((end_dt - start_dt).days / 30) + 1)
        except:
            months_active = 1
    elif start_date:
        # 只指定了开始日期，从开始日期到现在
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            months_active = max(1, int((datetime.now() - start_dt).days / 30) + 1)
        except:
            months_active = 1
    elif entry_date:
        # 没有日期筛选，使用入职以来的月数
        try:
            entry = datetime.strptime(entry_date, '%Y-%m-%d')
            months_active = max(1, int((datetime.now() - entry).days / 30))
        except:
            months_active = 1

    # 使用双轨评分模型
    safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
    safety_score = safety_result['final_score']
    safety_status_color = safety_result['status_color']
    safety_alert_tag = safety_result['alert_tag']
    safety_violations = len(violations_list)
    safety_total_score = sum(violations_list)

    # 4. 绩效能力分析（使用双算法系统）
    # 判断是月度还是周期（使用前面已经设置的 start_date 和 end_date）
    is_monthly = (start_date == end_date) if start_date and end_date else True
    current_app.logger.debug(f" [comprehensive-profile]: is_monthly={is_monthly}, start_date={start_date}, end_date={end_date}")

    # 构建绩效查询
    perf_query = """
        SELECT score, grade, year, month
        FROM performance_records
        WHERE emp_no = %s
    """
    perf_params = [emp_no]

    if start_date:
        s_year, s_month = map(int, start_date.split('-'))
        # 性能优化: 避免函数索引失效,使用元组比较 (year, month) >= (s_year, s_month)
        perf_query += " AND (year > %s OR (year = %s AND month >= %s))"
        perf_params.extend([s_year, s_year, s_month])

    if end_date:
        e_year, e_month = map(int, end_date.split('-'))
        perf_query += " AND (year < %s OR (year = %s AND month <= %s))"
        perf_params.extend([e_year, e_year, e_month])

    perf_query += " ORDER BY year, month"
    cur.execute(perf_query, perf_params)
    perf_rows = cur.fetchall()

    if perf_rows:
        if is_monthly and len(perf_rows) == 1:
            # 月度快照算法
            perf_row = perf_rows[0]
            raw_score = float(perf_row['score']) if perf_row['score'] else 95
            grade = perf_row['grade'] if perf_row['grade'] else 'B+'
            perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
            performance_score = perf_result['radar_value']
            performance_status_color = perf_result['status_color']
            performance_alert_tag = perf_result['alert_tag']
            performance_display_label = perf_result['display_label']
            performance_mode = 'MONTHLY'
        else:
            # 周期加权算法（带时间衰减）
            grade_list = [row['grade'] if row['grade'] else 'B+' for row in perf_rows]
            grade_dates = [f"{row['year']}-{row['month']:02d}" for row in perf_rows]  # 构建日期列表
            perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
            performance_score = perf_result['radar_value']
            performance_status_color = perf_result['status_color']
            performance_alert_tag = perf_result['alert_tag']
            performance_display_label = perf_result['display_label']
            performance_mode = 'PERIOD'
        performance_count = len(perf_rows)
    else:
        # 没有绩效数据
        performance_score = 0
        performance_count = 0
        performance_status_color = 'GREEN'
        performance_alert_tag = '暂无数据'
        performance_display_label = '暂无数据'
        performance_mode = 'MONTHLY'

    # 计算三维综合分（用于返回数据）
    current_comprehensive = (
        performance_score * score_weights.get('performance', 0.35) +
        safety_score * score_weights.get('safety', 0.30) +
        training_score * score_weights.get('training', 0.20)
    )
    
    # 5. 学习能力评估（新版：基于违规数量变化趋势）
    current_violations = 0
    previous_violations = None
    
    # 判断是否为长周期（跨月）
    is_long_term = False
    if start_date and end_date and start_date != end_date:
        is_long_term = True
        
    learning_result = None
    
    if is_long_term:
        # 长周期模式：V5.0 风险惯性模型
        try:
            # 1. 初始化每月计数
            monthly_counts = {}
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            
            # 预先查询周期前一个月的数据
            pre_period_month = (start_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
            pre_period_count = 0
            try:
                cur.execute("""
                    SELECT assessment FROM safety_inspection_records 
                    WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s
                """, [emp_name, pre_period_month])
                pre_rows = cur.fetchall()
                pre_period_count = sum(1 for r in pre_rows if extract_score_from_assessment(r['assessment']) > 0)
            except:
                pre_period_count = None

            curr = start_dt
            months_seq = []
            while curr <= end_dt:
                m_str = curr.strftime('%Y-%m')
                monthly_counts[m_str] = 0
                months_seq.append(m_str)
                curr = (curr.replace(day=1) + timedelta(days=32)).replace(day=1)
            
            # 2. 填充数据
            for row in safety_rows:
                if row['inspected_person'] == emp_name and extract_score_from_assessment(row['assessment']) > 0:
                    insp_date = row['inspection_date']
                    m_str = insp_date.strftime('%Y-%m') if hasattr(insp_date, 'strftime') else str(insp_date)[:7]
                    
                    if m_str in monthly_counts:
                        monthly_counts[m_str] += 1
                        
            # 3. 准备参数调用核心算法
            score_list = [monthly_counts[m] for m in months_seq]
            
            # 获取班组平均（周期整体）
            period_group_avg = 1.0 
            if dept_id:
                 try:
                    cur.execute("""
                        SELECT COUNT(*) / COUNT(DISTINCT e.name) / %s as avg_viol
                        FROM safety_inspection_records s
                        JOIN employees e ON s.inspected_person = e.name
                        WHERE e.department_id = %s
                        AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') >= %s
                        AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') <= %s
                    """, [max(1, len(months_seq)), dept_id, start_date, end_date])
                    avg_res = cur.fetchone()
                    if avg_res and avg_res['avg_viol']:
                        period_group_avg = float(avg_res['avg_viol'])
                        group_avg_violations = period_group_avg  # 更新外部变量
                 except:
                    pass

            # 4. 调用 calculate_learning_ability_longterm (V5.0 核心)
            # 注意：此函数已更新接受 initial_prev_viol
            learning_result = calculate_learning_ability_longterm(
                score_list=score_list,
                config=algo_config,
                group_avg=period_group_avg,
                initial_prev_viol=pre_period_count
            )
            
            # 补全部分前端需要的字段（如果 missed）
            # long-term函数返回了 risk_level, inertia_penalty_rate, max_consecutive_danger 等关键字段
            # 我们只需要补充 trend_type 供兼容旧代码逻辑判断
            if learning_result['risk_level'] == 'SAFE':
                learning_result['trend_type'] = 'safe'
            elif learning_result['risk_level'] in ['HIGH_RISK', 'PRE_ACCIDENT']:
                learning_result['trend_type'] = 'deterioration'
            elif learning_result['risk_level'] == 'WATCH_LIST':
                 learning_result['trend_type'] = 'yellow_alert'
            else:
                 learning_result['trend_type'] = 'fluctuation'
                 
            # 兼容设置
            current_violations = monthly_counts[months_seq[-1]]
            if len(months_seq) >= 2:
                previous_violations = monthly_counts[months_seq[-2]]
            else:
                previous_violations = pre_period_count
                
            monthly_scores = [0] * len(months_seq) # 标记为非空列表以触发 learning_months 计算

        except Exception as e:
            current_app.logger.error(f": 长周期学习能力计算异常: {e}")
            is_long_term = False

    if not learning_result:
        # 单月模式或降级
        # 单月模式：本月 vs 上月
        if end_date:
            try:
                # 统计end_date当月的违规数
                end_target = end_date
                # 注意：safety_rows中包含筛选范围内的数据，如果范围仅为单月，这里直接统计
                current_violations = sum(1 for r in safety_rows if r['inspected_person'] == emp_name and r['inspection_date'].strftime('%Y-%m') == end_target and extract_score_from_assessment(r['assessment']) > 0)
            except:
                current_violations = len(violations_list)
        else:
            current_violations = len(violations_list)

        # 获取上月违规数
        if start_date:
            try:
                current_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                prev_dt = current_dt.replace(day=1) - timedelta(days=1)
                prev_month = prev_dt.strftime('%Y-%m')
                
                cur.execute("""
                    SELECT assessment FROM safety_inspection_records
                    WHERE inspected_person = %s
                    AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s
                """, [emp_name, prev_month])
                prev_rows = cur.fetchall()
                previous_violations = sum(1 for r in prev_rows if extract_score_from_assessment(r['assessment']) > 0)
            except:
                previous_violations = None            

    
        # 获取班组平均违规数
        group_avg_violations = 1.0
        if dept_id and start_date:
            try:
                cur.execute("""
                    SELECT COUNT(*) / COUNT(DISTINCT e.name) as avg_viol
                    FROM safety_inspection_records s
                    JOIN employees e ON s.inspected_person = e.name
                    WHERE e.department_id = %s
                    AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') = %s
                """, [dept_id, start_date])
                avg_result = cur.fetchone()
                if avg_result and avg_result['avg_viol']:
                    group_avg_violations = float(avg_result['avg_viol'])
            except:
                pass
        
        # 调用新算法
        learning_result = calculate_learning_ability_new(
            current_violations=current_violations,
            previous_violations=previous_violations,
            group_avg_violations=group_avg_violations,
            config=algo_config
        )
    
    # 提取学习能力分值和详情
    learning_score = learning_result['learning_score']
    learning_status_color = learning_result['status_color']
    learning_alert_tag = learning_result['alert_tag']
    
    # [V5.0 补全] 确保 risk_level 等关键字段存在 (兼容单月/短周期模式)
    if 'risk_level' not in learning_result:
        zone = learning_result.get('zone', 'SAFE')
        if zone == 'CRITICAL':
             learning_result['risk_level'] = 'PRE_ACCIDENT'
        elif zone == 'DANGER':
             learning_result['risk_level'] = 'HIGH_RISK'
        elif zone == 'SAFE':
             learning_result['risk_level'] = 'SAFE'
        else:
             learning_result['risk_level'] = zone
             
    if 'inertia_penalty_rate' not in learning_result:
        learning_result['inertia_penalty_rate'] = 0.0
        
    if 'max_consecutive_danger' not in learning_result:
        learning_result['max_consecutive_danger'] = 0

    raw_trend_type = learning_result.get('trend_type', '未知')
    if raw_trend_type == 'high_improvement':
        learning_tier = 'improvement'
    elif raw_trend_type in ['deterioration_mild', 'meltdown']:
        learning_tier = 'deterioration'
    else:
        learning_tier = raw_trend_type
    
    # 获取新算法的详细指标
    learning_warning_line = learning_result.get('warning_line', 0)
    learning_critical_line = learning_result.get('critical_line', 0)
    
    # 尝试获取统计周期月数（如果是长周期模式）
    # 注意：在前面的代码中我们可能定义了 monthly_scores
    if 'monthly_scores' in locals() and monthly_scores:
        learning_months = len(monthly_scores)
    elif 'learning_result' in locals() and 'months' in learning_result:
        learning_months = learning_result['months']
    else:
        learning_months = 1


    # 6. 稳定性评估（波动型）
    stability_window_start, stability_window_end = _resolve_stability_window(start_date, end_date, algo_config)
    stability_window_months = _month_range(stability_window_start, stability_window_end)
    last_12_start = _month_shift(stability_window_end, -11)
    stability_query_start = (
        stability_window_start
        if _month_index(stability_window_start) <= _month_index(last_12_start)
        else last_12_start
    )

    violations_by_month = _load_monthly_safety_violations(cur, emp_name, stability_query_start, stability_window_end)
    monthly_safety_scores, monthly_issue_counts = _build_monthly_safety_scores(
        violations_by_month, stability_window_months, algo_config
    )
    issue_counts_last_12 = [
        len(violations_by_month.get(m, []))
        for m in _month_range(last_12_start, stability_window_end)
    ]

    # 计算月度综合分（用于CV对比提示）
    monthly_comprehensive_scores = {}
    if stability_window_months:
        perf_scores_by_month = {}
        train_scores_by_month = {}

        perf_start_year, perf_start_month = map(int, stability_window_months[0].split('-'))
        perf_end_year, perf_end_month = map(int, stability_window_months[-1].split('-'))

        cur.execute("""
            SELECT score, grade, year, month
            FROM performance_records
            WHERE emp_no = %s
            AND (year > %s OR (year = %s AND month >= %s))
            AND (year < %s OR (year = %s AND month <= %s))
        """, [emp_no, perf_start_year, perf_start_year, perf_start_month,
              perf_end_year, perf_end_year, perf_end_month])
        perf_rows_window = cur.fetchall()

        for row in perf_rows_window:
            m_str = f"{int(row['year']):04d}-{int(row['month']):02d}"
            raw_score = float(row['score']) if row['score'] else 95
            grade = row['grade'] if row['grade'] else 'B+'
            perf_score = calculate_performance_score_monthly(grade, raw_score, algo_config)['radar_value']
            perf_scores_by_month.setdefault(m_str, []).append(perf_score)

        for m_str, scores in perf_scores_by_month.items():
            if scores:
                perf_scores_by_month[m_str] = sum(scores) / len(scores)

        train_start_date = f"{stability_window_months[0]}-01"
        train_end_date = f"{_month_shift(stability_window_months[-1], 1)}-01"
        cur.execute("""
            SELECT score, is_qualified, is_disqualified, training_date
            FROM training_records
            WHERE emp_no = %s
            AND training_date >= %s
            AND training_date < %s
        """, [emp_no, train_start_date, train_end_date])
        train_rows_window = cur.fetchall()

        train_by_month = {}
        for row in train_rows_window:
            t_date = row['training_date']
            m_str = t_date.strftime('%Y-%m') if hasattr(t_date, 'strftime') else str(t_date)[:7]
            train_by_month.setdefault(m_str, []).append(row)

        for m_str, records in train_by_month.items():
            if records:
                train_score = calculate_training_score_with_penalty(records, 30, cert_years, algo_config)['radar_score']
                train_scores_by_month[m_str] = train_score

        for m_str in stability_window_months:
            components = {}
            if m_str in monthly_safety_scores:
                components['safety'] = monthly_safety_scores[m_str]
            if m_str in perf_scores_by_month:
                components['performance'] = perf_scores_by_month[m_str]
            if m_str in train_scores_by_month:
                components['training'] = train_scores_by_month[m_str]

            if components:
                weight_sum = 0.0
                weighted_sum = 0.0
                for key, value in components.items():
                    w = score_weights.get(key, 0)
                    weight_sum += w
                    weighted_sum += value * w
                if weight_sum > 0:
                    monthly_comprehensive_scores[m_str] = weighted_sum / weight_sum

    stability_result = calculate_stability_score_new(
        window_months=stability_window_months,
        monthly_safety_scores=monthly_safety_scores,
        monthly_issue_counts=monthly_issue_counts,
        issue_counts_last_12=issue_counts_last_12,
        monthly_comprehensive_scores=monthly_comprehensive_scores,
        safety_score_for_tip=safety_score,
        config=algo_config
    )
    stability_score = stability_result.get('stability_score', 50)
    if not isinstance(stability_score, (int, float)):
        stability_score = 50




    # 7. 计算综合能力分数（加权平均 - 使用配置权重）
    comprehensive_score = round(
        performance_score * score_weights['performance'] +
        safety_score * score_weights['safety'] +
        training_score * score_weights['training'] +
        stability_score * score_weights['stability'] +
        learning_score * score_weights['learning'],
        1
    )

    # 格式化日期为字符串（MySQL返回date对象，JSON无法直接序列化）
    def format_date(d):
        if d is None:
            return None
        if isinstance(d, str):
            return d
        if hasattr(d, 'strftime'):
            return d.strftime('%Y-%m-%d')
        return str(d)

    return jsonify({
        'employee': {
            'emp_no': emp_no,
            'name': emp_name,
            'position': position,
            'education': education,
            'entry_date': format_date(entry_date)
        },
        'scores': {
            'comprehensive': round(comprehensive_score, 1),
            'training': round(training_score, 1),
            'safety': round(safety_score, 1),
            'performance': round(performance_score, 1),
            'learning': round(learning_score, 1),
            'stability': round(stability_score, 1)
        },
        'personnel_details': {
            'working_years': round(working_years, 1) if working_years else None,
            'tenure_years': round(tenure_years, 1) if tenure_years else None,
            'certification_years': round(cert_years, 1) if cert_years else None,
            'solo_driving_years': round(solo_years, 1) if solo_years else None,
            'education': education
        },
        'safety_details': {
            'violations': safety_violations,
            'total_deduction': safety_total_score,
            'as_inspector': safety_as_inspector,
            'as_rectifier': safety_as_rectifier,
            'status_color': safety_status_color,
            'alert_tag': safety_alert_tag,
            'score_a': safety_result['score_a'],
            'score_b': safety_result['score_b'],
            'avg_freq': safety_result['avg_freq']
        },
        'statistics': {
            'total_trainings': total_training_count,
            'avg_training_score': training_score,
            'recent_trainings': len(training_records) if training_records else 0
        },
        'training_details': {
            'radar_score': training_score,
            'original_score': training_original_score,
            'penalty_coefficient': training_penalty_coeff,
            'status_color': training_status_color,
            'alert_tag': training_alert_tag,
            'total_ops': total_training_count,
            'fail_count': training_fail_count,
            'duration_days': duration_days
        },
        'performance_details': {
            'recent_avg': performance_score,
            'range': f'{"当月" if is_monthly else "统计周期"}',
            'count': performance_count,
            'status_color': performance_status_color,
            'alert_tag': performance_alert_tag,
            'display_label': performance_display_label,
            'mode': performance_mode
        },
        'learning_details': {
            'learning_score': round(learning_score, 1),
            'status_color': learning_status_color,
            'alert_tag': learning_alert_tag,
            'tier': learning_tier,
            'current_violations': current_violations,
            'previous_violations': previous_violations if previous_violations is not None else -1,  # -1表示无记录(冷启动)
            'group_avg': round(group_avg_violations, 1),
            'warning_line': learning_warning_line,
            'critical_line': learning_critical_line,
            'months': learning_months,
            # V5.0 新增字段
            'risk_level': learning_result.get('risk_level', 'UNKNOWN'),
            'inertia_penalty_rate': learning_result.get('inertia_penalty_rate', 0),
            'max_consecutive_danger': learning_result.get('max_consecutive_danger', 0),
            'base_score': learning_result.get('base_score', 0),  # 基础加权分（惯性扣减前）
            'has_meltdown': learning_result.get('has_meltdown', False),  # 是否曾触发熔断
            'zone': learning_result.get('zone', 'UNKNOWN'),  # 当前区域状态
            'slope': learning_result.get('slope', 0)
        },
        'stability_details': {
            'stability_score': round(stability_score, 1),
            'status_color': stability_result.get('status_color', 'GRAY'),
            'alert_tag': stability_result.get('alert_tag', '暂无数据'),
            'stability_label': stability_result.get('stability_label'),
            'volatility_metric': stability_result.get('volatility_metric'),
            'volatility_metric_label': stability_result.get('volatility_metric_label'),
            'volatility_value': stability_result.get('volatility_value'),
            'coverage': stability_result.get('coverage'),
            'confidence': stability_result.get('confidence'),
            'volatility_tip': stability_result.get('volatility_tip'),
            'low_level_tip': stability_result.get('low_level_tip'),
            'sample_tip': stability_result.get('sample_tip'),
            'safety_cv': stability_result.get('safety_cv'),
            'comprehensive_cv': stability_result.get('comprehensive_cv'),
            'mean_safety': stability_result.get('mean_safety')
        }
    })


@personnel_bp.route('/api/student-detail/<emp_no>')
@login_required
def api_student_detail(emp_no):
    """API: 获取学员详细数据（培训雷达图数据：按项目分类的平均分）"""
    conn = get_db()
    cur = conn.cursor()

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # 获取时间筛选参数
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    # 构建时间筛选条件
    time_filter = ""
    time_params = [emp_no]
    if year and month:
        time_filter = " AND YEAR(training_date) = %s AND MONTH(training_date) = %s"
        time_params.extend([str(year), str(month).zfill(2)])
    elif year:
        time_filter = " AND YEAR(training_date) = %s"
        time_params.append(str(year))

    # 查询该学员各项目分类的平均分
    query = f"""
        SELECT
            c.name as category_name,
            ROUND(AVG(tr.score), 1) as avg_score,
            COUNT(*) as count
        FROM training_records tr
        LEFT JOIN training_projects p ON tr.project_id = p.id
        LEFT JOIN training_project_categories c ON p.category_id = c.id
        WHERE tr.emp_no = %s AND c.name IS NOT NULL{time_filter}
        GROUP BY c.id, c.name
        ORDER BY c.display_order ASC
    """
    cur.execute(query, time_params)
    student_data = {}
    for row in cur.fetchall():
        student_data[row['category_name']] = {
            'avg_score': row['avg_score'],
            'count': row['count']
        }

    # 查询团队平均（基于权限过滤的可见员工）
    accessible_dept_ids = get_accessible_department_ids()
    if not accessible_dept_ids:
        return jsonify({
            'student_data': student_data,
            'team_data': {},
            'categories': sorted(list(student_data.keys()))
        })

    placeholders = ','.join(['%s'] * len(accessible_dept_ids))

    # 构建团队查询的时间筛选
    team_time_filter = ""
    team_time_params = accessible_dept_ids.copy()
    if year and month:
        team_time_filter = " AND YEAR(tr.training_date) = %s AND MONTH(tr.training_date) = %s"
        team_time_params.extend([str(year), str(month).zfill(2)])
    elif year:
        team_time_filter = " AND YEAR(tr.training_date) = %s"
        team_time_params.append(str(year))

    query = f"""
        SELECT
            c.name as category_name,
            ROUND(AVG(tr.score), 1) as avg_score
        FROM training_records tr
        LEFT JOIN employees e ON tr.emp_no = e.emp_no
        LEFT JOIN training_projects p ON tr.project_id = p.id
        LEFT JOIN training_project_categories c ON p.category_id = c.id
        WHERE (e.department_id IN ({placeholders}) OR e.emp_no IS NULL)
            AND c.name IS NOT NULL{team_time_filter}
        GROUP BY c.id, c.name
        ORDER BY c.display_order ASC
    """
    cur.execute(query, team_time_params)
    team_data = {}
    for row in cur.fetchall():
        team_data[row['category_name']] = row['avg_score']

    # 合并所有分类
    all_categories = set(student_data.keys()) | set(team_data.keys())

    return jsonify({
        'student_data': student_data,
        'team_data': team_data,
        'categories': sorted(list(all_categories))
    })


@personnel_bp.route('/api/student-growth/<emp_no>')
@login_required
def api_student_growth(emp_no):
    """API: 获取学员成长趋势数据（按时间的平均分变化）"""
    conn = get_db()
    cur = conn.cursor()

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # 查询该学员按月份的平均分趋势
    query = """
        SELECT
            DATE_FORMAT(training_date, '%%Y-%%m') as month,
            ROUND(AVG(score), 1) as avg_score,
            COUNT(*) as count
        FROM training_records
        WHERE emp_no = %s
        GROUP BY month
        ORDER BY month ASC
    """
    cur.execute(query, (emp_no,))

    growth_data = []
    for row in cur.fetchall():
        growth_data.append({
            'month': row['month'],
            'avg_score': row['avg_score'],
            'count': row['count']
        })

    return jsonify(growth_data)


# ============================================================
# Risk Mining API - 风险挖掘接口
# ============================================================

@personnel_bp.route('/api/risk-mining')
@login_required
def api_risk_mining():
    """
    API: 高风险人员挖掘分析

    Query Parameters:
        start_date: 开始日期 (YYYY-MM), 默认12个月前
        end_date: 结束日期 (YYYY-MM), 默认当前月
        enable_ai: 是否启用AI诊断 (true/false), 默认true

    Returns:
        {
            high_risk_list: [...],   # 按风险分排序的员工列表
            keyword_cloud: [...],     # 关键词词云数据
            survival_curve: [...],    # 生存曲线数据
            summary: {...}            # 统计摘要
        }
    """
    try:
        from services.risk_mining_service import RiskMiningService

        # Parse parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        enable_ai = request.args.get('enable_ai', 'true').lower() == 'true'

        # 使用标准权限控制函数进行安全过滤
        role = session.get('role')
        department_path = None

        if role != 'admin':
            # 非管理员用户需要检查部门权限
            accessible_dept_ids = get_accessible_department_ids()

            if not accessible_dept_ids:
                # 用户没有可访问的部门，返回空结果（安全保护）
                return jsonify({
                    'success': True,
                    'high_risk_list': [],
                    'keyword_cloud': [],
                    'survival_curve': [],
                    'summary': {
                        'total_employees': 0,
                        'high_risk_count': 0,
                        'anomaly_count': 0,
                        'analysis_period': f'{start_date} ~ {end_date}' if start_date and end_date else '最近12个月'
                    },
                    'message': '您未被分配到任何部门，无法查看风险数据'
                })

            # 获取用户部门路径用于过滤
            conn = get_db()
            cur = conn.cursor()
            user_id = session.get('user_id')
            cur.execute("""
                SELECT d.path FROM users u
                JOIN departments d ON u.department_id = d.id
                WHERE u.id = %s
            """, (user_id,))
            row = cur.fetchone()
            if row:
                department_path = row['path']
            else:
                # 用户有可访问部门但没有自己的部门路径（异常情况），返回空结果
                return jsonify({
                    'success': True,
                    'high_risk_list': [],
                    'keyword_cloud': [],
                    'survival_curve': [],
                    'summary': {
                        'total_employees': 0,
                        'high_risk_count': 0,
                        'anomaly_count': 0,
                        'analysis_period': f'{start_date} ~ {end_date}' if start_date and end_date else '最近12个月'
                    },
                    'message': '您的用户账户未关联部门，无法查看风险数据'
                })

        # Perform risk analysis
        result = RiskMiningService.analyze_all(
            start_date=start_date,
            end_date=end_date,
            department_path=department_path,
            enable_ai_diagnosis=enable_ai
        )

        return jsonify({
            'success': True,
            **result
        })

    except ImportError as e:
        return jsonify({
            'success': False,
            'error': f'缺少必要的依赖库: {str(e)}。请运行 pip install pandas scikit-learn jieba lifelines httpx'
        }), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'风险分析失败: {str(e)}'
        }), 500


@personnel_bp.route('/risk-mining')
@login_required
def risk_mining_page():
    """风险挖掘分析页面"""
    return render_template('personnel_risk_mining.html', title='高风险人员挖掘')


@personnel_bp.route('/api/ai-diagnosis', methods=['POST'])
@login_required
def api_ai_diagnosis():
    """
    API: 单个员工AI诊断（带缓存机制）

    Request Body (JSON):
        emp_no: 员工工号
        name: 员工姓名
        risk_score: 风险评分
        risk_data: 风险数据（基础统计信息，详细记录由后端重新获取）
        start_date: 开始日期 (YYYY-MM 格式，用于过滤详细记录)
        end_date: 结束日期 (YYYY-MM 格式，用于过滤详细记录)

    Returns:
        {
            success: true/false,
            diagnosis: {...},  # 诊断结果
            source: "cache" | "api",  # 数据来源（节省token的关键指标）
            tokens_used: 0  # 如果命中缓存则为0
        }
    """
    try:
        from services.ai_diagnosis_service import AIDiagnosisService
        from services.risk_mining_service import RiskMiningService

        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': '请求数据为空'
            }), 400

        emp_no = data.get('emp_no')
        name = data.get('name', '未知')
        risk_score = data.get('risk_score', 0)
        basic_risk_data = data.get('risk_data', {})
        # 新增：获取日期范围参数，用于过滤详细记录
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        if not emp_no:
            return jsonify({
                'success': False,
                'error': '缺少员工工号'
            }), 400

        # 检查AI是否已配置
        if not AIDiagnosisService.is_configured():
            return jsonify({
                'success': False,
                'error': 'AI未配置。请在系统设置中添加AI提供商配置。'
            }), 400

        # 关键：从数据库重新获取详细记录，确保与批量分析时的数据一致
        # 这样才能命中缓存！使用日期范围过滤，确保AI诊断的数据与用户筛选一致
        risk_data = {
            # 基础统计数据（来自前端）
            'performance_slope': basic_risk_data.get('performance_slope', 0),
            'performance_mean': basic_risk_data.get('performance_score', 0),
            'safety_count': basic_risk_data.get('safety_count', 0),
            'training_disqualified_count': basic_risk_data.get('training_disqualified_count', 0),
            'is_anomaly': basic_risk_data.get('is_anomaly', False),
            'anomaly_score': basic_risk_data.get('anomaly_score', 0),
            'risk_factors': basic_risk_data.get('risk_factors', []),
            # 详细记录（从数据库重新获取，使用日期范围过滤，与批量分析保持一致）
            'recent_violations': RiskMiningService._get_recent_violations(emp_no, 10, start_date, end_date),
            'severe_violations': RiskMiningService._get_severe_violations(emp_no, start_date, end_date),
            'failed_training': RiskMiningService._get_failed_training(emp_no, start_date, end_date)
        }

        # 调用AI诊断（内置缓存逻辑）
        result = AIDiagnosisService.diagnose_sync(
            emp_no=emp_no,
            name=name,
            risk_score=risk_score,
            risk_data=risk_data
        )

        if result.success:
            return jsonify({
                'success': True,
                'diagnosis': result.parsed_result or result.diagnosis,
                'raw_diagnosis': result.diagnosis,
                'source': result.source,  # "cache" 或 "api"
                'tokens_used': result.tokens_used or 0,
                'model_used': result.model_used,
                'provider_name': result.provider_name
            })
        else:
            return jsonify({
                'success': False,
                'error': result.error
            }), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'AI诊断失败: {str(e)}'
        }), 500
