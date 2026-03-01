#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""人员管理 - 数据分析路由（统计/九宫格/学员管理）"""
import os
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session, current_app

from config.settings import APP_TITLE, EXPORT_DIR
from models.database import get_db, close_db, get_year_month_concat
from ..decorators import login_required, manager_required
from ..helpers import (
    current_user_id, require_user_id, get_accessible_department_ids,
    get_accessible_departments, calculate_years_from_date, get_user_department,
    validate_employee_access, log_import_operation, parse_time_range,
    month_range_to_dates
)
from . import personnel_bp
from . import (
    list_personnel, get_personnel, _serialize_person,
    _parse_date_string, _normalize_date_to_str,
    _calculate_age, _calculate_years_since,
    calculate_learning_ability_monthly, calculate_learning_ability_longterm,
    calculate_stability_score, calculate_stability_for_employee,
    calculate_stability_score_new, calculate_stability_period_aggregated,
    calculate_inertia_penalty, calculate_learning_ability_new,
    _month_index, _month_shift, _month_range, _resolve_stability_window,
    _load_monthly_safety_violations, _build_monthly_safety_scores,
)
from services.domain.personnel_algo import (
    calculate_performance_score_monthly,
    calculate_performance_score_period,
    calculate_safety_score_dual_track,
    calculate_training_score_with_penalty,
)
from services.domain.safety_utils import extract_score_from_assessment
import json
import re
from collections import Counter
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO
from openpyxl import Workbook



@personnel_bp.route('/analytics')
@login_required
def analytics():
    """人员数据分析页面"""
    return render_template(
        "personnel_analytics.html",
        title=f"人员数据分析 | {APP_TITLE}"
    )


@personnel_bp.route('/api/analytics-data')
@login_required
def api_analytics_data():
    """获取人员分析数据API"""
    rows = list_personnel()

    # 岗位筛选：只统计电客车司机，排除副队长和队长
    def is_driver(row):
        position = (row.get("position") or "").strip()
        # 排除副队长和队长
        if "队长" in position:
            return False
        # 只要包含"司机"就算
        return "司机" in position

    # 除了政治面貌统计，其他都只统计司机
    driver_rows = [row for row in rows if is_driver(row)]

    # 1. 安全风险等级分布 - 按入司后单独驾驶年限分级
    risk_levels = {"新手(<1年)": 0, "成长(1-3年)": 0, "熟练(3-5年)": 0, "资深(≥5年)": 0, "未知": 0}
    for row in driver_rows:
        solo_years = row.get("solo_driving_years")
        if solo_years is None:
            risk_levels["未知"] += 1
        elif solo_years < 1:
            risk_levels["新手(<1年)"] += 1
        elif 1 <= solo_years < 3:
            risk_levels["成长(1-3年)"] += 1
        elif 3 <= solo_years < 5:
            risk_levels["熟练(3-5年)"] += 1
        else:
            risk_levels["资深(≥5年)"] += 1

    # 2. 部门战力雷达图 - 各部门的平均司龄、驾龄、取证年限（只统计司机）
    # 获取当前用户可访问的部门列表
    accessible_depts = get_accessible_departments()

    # 获取所有底层部门（没有子部门的部门）
    conn = get_db()
    cur = conn.cursor()

    # 找出所有可访问部门中的底层部门
    accessible_dept_ids = [dept['id'] for dept in accessible_depts]
    if not accessible_dept_ids:
        team_power = []
    else:
        # 查询每个部门是否有子部门
        placeholders = ','.join(['%s'] * len(accessible_dept_ids))
        cur.execute(f"""
            SELECT d.id, d.name, d.level,
                   CASE
                       WHEN EXISTS(SELECT 1 FROM departments child WHERE child.parent_id = d.id)
                       THEN 1 ELSE 0
                   END as has_children
            FROM departments d
            WHERE d.id IN ({placeholders})
            ORDER BY d.level, d.name
        """, accessible_dept_ids)

        dept_info = {row['id']: dict(row) for row in cur.fetchall()}

        # 管理员始终显示所有可访问的底层部门。
        # 非管理员：最底层用户只显示自己部门；上级用户显示所有下级底层部门。
        # P1 统一出口
        from services.access_control_service import AccessControlService
        user_dept_info = get_user_department()
        if AccessControlService.is_admin():
            display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]
        elif user_dept_info and user_dept_info['department_id']:
            user_dept_id = user_dept_info['department_id']
            # 检查用户部门是否是底层部门
            if user_dept_id in dept_info and dept_info[user_dept_id]['has_children'] == 0:
                # 用户是底层部门，只显示自己部门
                display_dept_ids = [user_dept_id]
            else:
                # 用户是上级部门，显示所有可访问的底层部门
                display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]
        else:
            # 管理员或无部门用户，显示所有底层部门
            display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]

        # 按部门统计司机数据
        dept_stats = {}
        for row in driver_rows:
            dept_id = row.get("department_id")
            if dept_id not in display_dept_ids:
                continue

            if dept_id not in dept_stats:
                dept_stats[dept_id] = {
                    "name": dept_info.get(dept_id, {}).get('name', '未知部门'),
                    "tenure_years": [],
                    "solo_driving_years": [],
                    "certification_years": []
                }

            if row.get("tenure_years") is not None:
                dept_stats[dept_id]["tenure_years"].append(row["tenure_years"])
            if row.get("solo_driving_years") is not None:
                dept_stats[dept_id]["solo_driving_years"].append(row["solo_driving_years"])
            if row.get("certification_years") is not None:
                dept_stats[dept_id]["certification_years"].append(row["certification_years"])

        team_power = []
        for dept_id, stats in dept_stats.items():
            avg_tenure = sum(stats["tenure_years"]) / len(stats["tenure_years"]) if stats["tenure_years"] else 0
            avg_solo = sum(stats["solo_driving_years"]) / len(stats["solo_driving_years"]) if stats["solo_driving_years"] else 0
            avg_cert = sum(stats["certification_years"]) / len(stats["certification_years"]) if stats["certification_years"] else 0

            team_power.append({
                "team": stats["name"],
                "avg_tenure": round(avg_tenure, 1),
                "avg_solo": round(avg_solo, 1),
                "avg_cert": round(avg_cert, 1),
                "member_count": len([r for r in driver_rows if r.get("department_id") == dept_id])
            })

    # 3. 经验溢出分析 - 散点图数据（只统计司机）
    experience_scatter = []
    for row in driver_rows:
        cert_years = row.get("certification_years")
        solo_years = row.get("solo_driving_years")
        if cert_years is not None and solo_years is not None:
            experience_scatter.append({
                "name": row.get("name"),
                "emp_no": row.get("emp_no"),
                "cert_years": round(cert_years, 1),
                "solo_years": round(solo_years, 1),
                # 分类：准师傅(取证久但单驾短)、资深师傅(两项都高)、新手
                "category": _categorize_experience(cert_years, solo_years)
            })

    # 4. 排班压力预警 - 籍贯分布（只统计司机）+ 政治面貌统计（统计所有人）
    hometown_stats = {}
    political_stats = {"中共党员": 0, "中共预备党员": 0, "共青团员": 0, "群众": 0, "其它": 0}

    # 籍贯统计只统计司机
    for row in driver_rows:
        hometown = row.get("hometown") or "未填写"
        # 河南省内细分到市/县，省外只显示省份
        location = _extract_location(hometown)
        hometown_stats[location] = hometown_stats.get(location, 0) + 1

    # 政治面貌统计所有人员
    for row in rows:
        political = row.get("political_status") or "未填写"
        if political in political_stats:
            political_stats[political] += 1
        else:
            political_stats["其它"] += 1

    # 5. 职业稳定性分析 - 司龄 vs 工龄散点图（只统计司机）
    stability_scatter = []
    for row in driver_rows:
        tenure = row.get("tenure_years")
        working = row.get("working_years")
        if tenure is not None and working is not None:
            stability_scatter.append({
                "name": row.get("name"),
                "emp_no": row.get("emp_no"),
                "tenure": round(tenure, 1),
                "working": round(working, 1),
                # 分类：应届入职、社招新员工、社招老员工
                "category": _categorize_stability(tenure, working)
            })

    return jsonify({
        "risk_distribution": risk_levels,
        "team_power": team_power,
        "experience_scatter": experience_scatter,
        "hometown_stats": hometown_stats,
        "political_stats": political_stats,
        "stability_scatter": stability_scatter,
        "total_count": len(rows),
        "driver_count": len(driver_rows)
    })


def _categorize_experience(cert_years: float, solo_years: float) -> str:
    """分类经验等级"""
    if cert_years >= 5 and solo_years < 3:
        return "准师傅"  # 取证很久但单驾时间较短
    elif cert_years >= 5 and solo_years >= 5:
        return "资深师傅"  # 两项指标都高
    elif cert_years < 2:
        return "新手"
    else:
        return "普通"


def _categorize_stability(tenure: float, working: float) -> str:
    """分类职业稳定性

    Args:
        tenure: 司龄（在本单位工作年限）
        working: 工龄（总工作年限）

    Returns:
        分类标签：应届入职、社招(新)、社招(老)
    """
    work_exp_diff = working - tenure  # 入职前的工作经验

    if work_exp_diff < 1:
        # 工龄和司龄相近，基本是应届生或毕业后很快入职
        return "应届入职"
    elif tenure < 3:
        # 有工作经验，但在本单位时间不长
        return "社招(新)"
    else:
        # 有工作经验，且在本单位时间较长
        return "社招(老)"


def _extract_location(hometown: str) -> str:
    """提取地域信息

    河南省内细分到市/县，其他省份只显示省外或省份名称

    Args:
        hometown: 籍贯字符串，如"河南郑州"、"河南省洛阳市"、"江苏南京"等

    Returns:
        地域标签：河南省内返回市/县名，省外返回省份名或"省外"
    """
    if not hometown or hometown == "未填写":
        return "未填写"

    hometown = hometown.strip()

    # 河南省内的地级市和县
    henan_cities = [
        "郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁",
        "新乡", "焦作", "濮阳", "许昌", "漯河", "三门峡",
        "南阳", "商丘", "信阳", "周口", "驻马店", "济源"
    ]

    # 常见县级市/县（可根据实际情况扩展）
    henan_counties = [
        "巩义", "荥阳", "新密", "新郑", "登封", "中牟",
        "兰考", "杞县", "通许", "尉氏", "偃师", "孟津",
        "新安", "栾川", "嵩县", "汝阳", "宜阳", "洛宁",
        "伊川", "汝州", "舞钢", "林州", "卫辉", "辉县",
        "沁阳", "孟州", "禹州", "长葛", "义马", "灵宝",
        "永城", "项城", "邓州", "固始", "鹿邑", "新蔡"
    ]

    # 检查是否为河南省内
    is_henan = False
    if "河南" in hometown:
        is_henan = True
    else:
        # 如果没有明确写"河南"，但包含河南的市/县名，也认为是河南
        for city in henan_cities + henan_counties:
            if city in hometown:
                is_henan = True
                break

    if is_henan:
        # 河南省内，提取市/县名
        # 优先匹配县级市/县（更具体）
        for county in henan_counties:
            if county in hometown:
                return f"河南·{county}"

        # 再匹配地级市
        for city in henan_cities:
            if city in hometown:
                return f"河南·{city}"

        # 如果只写了"河南"，返回"河南·未详"
        return "河南·未详"

    else:
        # 非河南省，提取省份
        provinces = [
            "北京", "天津", "上海", "重庆",
            "河北", "山西", "辽宁", "吉林", "黑龙江",
            "江苏", "浙江", "安徽", "福建", "江西", "山东",
            "湖北", "湖南", "广东", "海南",
            "四川", "贵州", "云南", "陕西", "甘肃",
            "青海", "台湾", "内蒙古", "广西", "西藏",
            "宁夏", "新疆", "香港", "澳门"
        ]

        for province in provinces:
            if hometown.startswith(province) or province in hometown:
                return f"省外·{province}"

        # 如果无法识别，返回"省外·其他"
        return "省外·其他"


# ==================== 个人综合能力画像 API ====================

@personnel_bp.route('/capability-profile')
@login_required
def capability_profile():
    """个人综合能力画像页面"""
    return render_template('personnel_capability_profile.html', title='个人综合能力画像')


@personnel_bp.route('/api/key-personnel-config')
@login_required
def api_key_personnel_config():
    """API: 获取关键人员配置参数（供前端动态显示使用）"""
    from services.algorithm_config_service import AlgorithmConfigService

    try:
        algo_config = AlgorithmConfigService.get_active_config()
        key_personnel_config = algo_config.get('key_personnel', {})

        return jsonify({
            'success': True,
            'config': {
                'comprehensive_threshold': key_personnel_config.get('comprehensive_threshold', 75),
                'monthly_violation_threshold': key_personnel_config.get('monthly_violation_threshold', 3)
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'config': {
                'comprehensive_threshold': 75,
                'monthly_violation_threshold': 3
            }
        })


@personnel_bp.route('/nine-grid')
@login_required
def page_nine_grid():
    """人才九宫格页面"""
    return render_template('personnel_nine_grid.html')


@personnel_bp.route('/api/departments')
@login_required
def api_departments():
    """API: 获取可访问部门列表"""
    return jsonify(get_accessible_departments())


@personnel_bp.route('/api/nine-grid-data')
@login_required
def api_nine_grid_data():
    """API: 获取九宫格数据"""
    from datetime import datetime
    
    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    tr = parse_time_range(request.args, ['month'], default_grain='month', default_range='current_month')
    start_month = tr['start_month']
    end_month = tr['end_month']
    department_filter = request.args.get('department_id')

    # 读取配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    
    rows = list_personnel()
    
    # 筛选
    if department_filter:
        try:
            dept_id_filter = int(department_filter)
            rows = [r for r in rows if r.get('department_id') == dept_id_filter]
        except ValueError:
            pass

    data = []
    
    # 获取权重配置
    score_weights = algo_config['comprehensive']['score_weights']
    nine_grid_weights = algo_config.get('nine_grid', {}).get('y_axis_weights', {
        'stability': 0.4,
        'learning': 0.6,
    })
    
    # 三维分权重归一化（去除稳定性和学习能力后的相对权重）
    w_perf = score_weights.get('performance', 35)
    w_safe = score_weights.get('safety', 30)
    w_train = score_weights.get('training', 20)
    w_x_total = w_perf + w_safe + w_train
    if w_x_total <= 0: w_x_total = 1

    for row in rows:
        try:
            scores = _calculate_single_employee_score(row, start_month, end_month, algo_config, cur)
            
            # 计算X轴（三维综合分）
            x_raw = (scores['performance'] * w_perf + 
                     scores['safety'] * w_safe + 
                     scores['training'] * w_train)
            x_score = round(x_raw / w_x_total * 100, 1) if w_x_total > 1 else round(x_raw, 1) # 假设配置是百分比整数(35)或小数(0.35)
            # 如果配置是小数(0.35)，w_x_total=0.85。 x_raw = P*0.35 ... -> x_raw / 0.85 * 100? No.
            # 如果配置是0.35，x_raw 是加权后的分。
            # 如果 P=100, x_raw = 35. 35/0.85 approx 41. ???
            # 综合分通常是加权和。
            # x_score 应该是满分100。
            # 如 P=100, S=100, T=100. x_raw = 35+30+20 = 85.
            # 那么 x_score = 85 / 0.85 = 100. Correct.
            if w_x_total < 5: # 检测是否为小数配置 (e.g. 0.35)
                # 系数本来就是小数，不需要 * 100 ?
                # 0.35+0.30+0.20 = 0.85
                # raw = 100*0.35 + ... = 85
                # 85 / 0.85 = 100.
                x_score = round(x_raw / w_x_total, 1)
            else:
                # 配置是整数 (35, 30, 20) -> Sum 85
                # raw = 100*35 + ... = 8500
                # 8500 / 85 = 100 ?
                # 通常 weighted average = sum(val * weight) / sum(weights)
                x_score = round(x_raw / w_x_total, 1)

            
            # 计算Y轴（稳定 + 学习）
            y_w_stab = nine_grid_weights.get('stability', 0.4) 
            y_w_learn = nine_grid_weights.get('learning', 0.6)
            y_total = y_w_stab + y_w_learn
            if y_total <= 0: y_total = 1
            
            y_raw = (scores['stability'] * y_w_stab + scores['learning'] * y_w_learn)
            y_score = round(y_raw / y_total, 1)

            # 判定九宫格位置 (3x3)
            # 简单的三分法：<75(Low), 75-90(Mid), >=90(High)
            # 可以根据实际需求调整阈值
            
            x_level = 1
            if x_score >= 90: x_level = 3
            elif x_score >= 75: x_level = 2
            
            y_level = 1
            if y_score >= 90: y_level = 3
            elif y_score >= 75: y_level = 2
            
            # 映射到行和列
            # 界面布局：
            # Row 1 (High Y): Cell 1-1, 1-2, 1-3
            # Row 2 (Mid Y)
            # Row 3 (Low Y)
            # Col 1 (Low X), Col 2 (Mid X), Col 3 (High X)
            
            grid_row = 4 - y_level # 3->1, 2->2, 1->3
            grid_col = x_level     # 1->1, 2->2, 3->3
            
            data.append({
                'emp_no': row.get('emp_no'),
                'name': row.get('name'),
                'department_name': row.get('department_name'),
                'x_score': x_score,
                'y_score': y_score,
                'grid_row': grid_row,
                'grid_col': grid_col,
                'details': scores
            })
        except Exception as e:
            current_app.logger.error(f"Error calculating score for {row.get('name')}: {e}")
            continue

    close_db()
    return jsonify(data)


@personnel_bp.route('/nine-grid/export')
@login_required
def export_nine_grid():
    """导出人才九宫格数据"""
    from datetime import datetime
    
    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    tr = parse_time_range(request.args, ['month'], default_grain='month', default_range='current_month')
    start_month = tr['start_month']
    end_month = tr['end_month']
    department_filter = request.args.get('department_id')

    # 读取配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    
    rows = list_personnel()
    
    # 筛选
    if department_filter:
        try:
            dept_id_filter = int(department_filter)
            rows = [r for r in rows if r.get('department_id') == dept_id_filter]
        except ValueError:
            pass

    # 获取权重配置
    score_weights = algo_config['comprehensive']['score_weights']
    nine_grid_weights = algo_config.get('nine_grid', {}).get('y_axis_weights', {
        'stability': 0.4,
        'learning': 0.6,
    })
    
    # 三维分权重归一化
    w_perf = score_weights.get('performance', 35)
    w_safe = score_weights.get('safety', 30)
    w_train = score_weights.get('training', 20)
    w_x_total = w_perf + w_safe + w_train
    if w_x_total <= 0: w_x_total = 1

    # 九宫格标签映射
    grid_labels = {
        (1, 1): '培养对象',
        (1, 2): '潜力新星',
        (1, 3): '明星员工',
        (2, 1): '改善对象',
        (2, 2): '中坚力量',
        (2, 3): '骨干员工',
        (3, 1): '问题员工',
        (3, 2): '需关注',
        (3, 3): '待观察稳定'
    }

    # 汇总数据：按九宫格位置统计
    summary_data = {}
    detail_data = []
    
    for row in rows:
        try:
            scores = _calculate_single_employee_score(row, start_month, end_month, algo_config, cur)
            
            # 计算X轴（三维综合分）
            x_raw = (scores['performance'] * w_perf + 
                     scores['safety'] * w_safe + 
                     scores['training'] * w_train)
            if w_x_total < 5:
                x_score = round(x_raw / w_x_total, 1)
            else:
                x_score = round(x_raw / w_x_total, 1)
            
            # 计算Y轴（稳定 + 学习）
            y_w_stab = nine_grid_weights.get('stability', 0.4) 
            y_w_learn = nine_grid_weights.get('learning', 0.6)
            y_total = y_w_stab + y_w_learn
            if y_total <= 0: y_total = 1
            
            y_raw = (scores['stability'] * y_w_stab + scores['learning'] * y_w_learn)
            y_score = round(y_raw / y_total, 1)

            # 判定九宫格位置
            x_level = 1
            if x_score >= 90: x_level = 3
            elif x_score >= 75: x_level = 2
            
            y_level = 1
            if y_score >= 90: y_level = 3
            elif y_score >= 75: y_level = 2
            
            grid_row = 4 - y_level
            grid_col = x_level
            
            # 汇总统计
            grid_key = (grid_row, grid_col)
            if grid_key not in summary_data:
                summary_data[grid_key] = []
            summary_data[grid_key].append(row.get('name'))
            
            # 明细数据
            detail_data.append({
                'emp_no': row.get('emp_no'),
                'name': row.get('name'),
                'department_name': row.get('department_name'),
                'x_score': x_score,
                'y_score': y_score,
                'grid_label': grid_labels.get(grid_key, ''),
                'performance': scores['performance'],
                'safety': scores['safety'],
                'training': scores['training'],
                'stability': scores['stability'],
                'learning': scores['learning']
            })
        except Exception as e:
            current_app.logger.error(f"Error calculating score for {row.get('name')}: {e}")
            continue

    close_db()

    # 创建Excel工作簿
    wb = Workbook()
    
    # 第一个工作表：汇总（9个单元格）
    ws_summary = wb.active
    ws_summary.title = '汇总'
    
    # 设置汇总表的标题和数据
    ws_summary['A1'] = '九宫格汇总'
    ws_summary['A1'].font = ws_summary['A1'].font.copy()
    
    # 按照九宫格布局填充汇总数据
    # 行标题（Y轴）
    ws_summary['A2'] = '高'
    ws_summary['A3'] = '中'
    ws_summary['A4'] = '低'
    
    # 列标题（X轴）
    ws_summary['B1'] = '低'
    ws_summary['C1'] = '中'
    ws_summary['D1'] = '高'
    
    # 填充9个单元格
    for row_idx in range(1, 4):
        for col_idx in range(1, 4):
            grid_key = (row_idx, col_idx)
            cell_row = row_idx + 1
            cell_col = col_idx + 1
            cell = ws_summary.cell(row=cell_row, column=cell_col)
            
            label = grid_labels.get(grid_key, '')
            count = len(summary_data.get(grid_key, []))
            names = ', '.join(summary_data.get(grid_key, []))
            
            cell.value = f"{label}\n({count}人)\n{names}"
            cell.alignment = cell.alignment.copy()
    
    # 第二个工作表：明细
    ws_detail = wb.create_sheet('明细')
    ws_detail.append(['工号', '姓名', '部门', 'X分数', 'Y分数', '九宫格位置', '绩效', '安全', '培训', '稳定性', '学习能力'])
    
    for item in detail_data:
        ws_detail.append([
            item['emp_no'],
            item['name'],
            item['department_name'],
            item['x_score'],
            item['y_score'],
            item['grid_label'],
            item['performance'],
            item['safety'],
            item['training'],
            item['stability'],
            item['learning']
        ])
    
    # 保存文件
    export_filename = f"人才九宫格_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    xlsx_path = os.path.join(EXPORT_DIR, export_filename)
    wb.save(xlsx_path)
    
    return send_file(xlsx_path, as_attachment=True, download_name=export_filename)


def _calculate_single_employee_score(row, start_month, end_month, algo_config, cur):
    """
    辅助函数：计算单个员工的各项评分
    """
    from services.domain.safety_utils import extract_score_from_assessment
    from datetime import datetime, timedelta
    import calendar

    emp_no = row.get('emp_no')
    emp_name = row.get('name')
    dept_id = row.get('department_id')
    entry_date = row.get('entry_date')

    # 计算取证年限
    cert_date = row.get('certification_date')
    cert_years = calculate_years_from_date(cert_date) if cert_date else None

    # 1. 培训能力
    training_query = "SELECT score, is_qualified, is_disqualified, training_date FROM training_records WHERE emp_no = %s"
    training_params = [emp_no]

    if start_month:
        training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') >= %s"
        training_params.append(start_month)
    if end_month:
        training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') <= %s"
        training_params.append(end_month)

    training_query += " ORDER BY training_date ASC"
    cur.execute(training_query, training_params)
    training_records_list = cur.fetchall()

    if start_month and end_month:
        try:
            start_dt = datetime.strptime(start_month + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_month + '-01', '%Y-%m-%d')
            ey, em = int(end_month.split('-')[0]), int(end_month.split('-')[1])
            last_day = calendar.monthrange(ey, em)[1]
            end_dt = end_dt.replace(day=last_day)
            duration_days = max(1, (end_dt - start_dt).days + 1)
        except Exception:
            duration_days = 30
    else:
        duration_days = 30

    training_result = calculate_training_score_with_penalty(training_records_list, duration_days, cert_years, algo_config)
    training_score = training_result['radar_score']

    # 2. 安全意识
    safety_query = "SELECT assessment, inspection_date FROM safety_inspection_records WHERE inspected_person = %s"
    safety_params = [emp_name]

    if start_month:
        safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') >= %s"
        safety_params.append(start_month)
    if end_month:
        safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') <= %s"
        safety_params.append(end_month)

    safety_query += " ORDER BY inspection_date ASC"
    cur.execute(safety_query, safety_params)
    safety_rows = cur.fetchall()

    violations_list = []
    for s_row in safety_rows:
        score = extract_score_from_assessment(s_row['assessment'])
        if score > 0:
            violations_list.append(float(score))

    months_active = 1
    if start_month:
        try:
            start = datetime.strptime(start_month + '-01', '%Y-%m-%d')
            end = datetime.strptime(end_month + '-01', '%Y-%m-%d') if end_month else datetime.now()
            if not end_month:
                months_active = max(1, int((end - start).days / 30) + 1)
            else:
                months_active = max(1, int((end - start).days / 30) + 1)
        except Exception:
            months_active = 1
    elif entry_date:
        try:
             # entry_date可能是date对象或str
            entry = entry_date if isinstance(entry_date, datetime) or hasattr(entry_date, 'year') else datetime.strptime(str(entry_date), '%Y-%m-%d')
            months_active = max(1, int((datetime.now() - entry).days / 30))
        except Exception:
            months_active = 1

    safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
    safety_score = safety_result['final_score']

    # 3. 工作绩效
    is_monthly = (start_month == end_month) if start_month and end_month else True
    perf_query = "SELECT score, grade, year, month FROM performance_records WHERE emp_no = %s"
    perf_params = [emp_no]

    if start_month:
        perf_query += " AND CAST(CONCAT(year, '-', LPAD(month, 2, '0')) AS CHAR) >= %s"
        perf_params.append(start_month)
    if end_month:
        perf_query += " AND CAST(CONCAT(year, '-', LPAD(month, 2, '0')) AS CHAR) <= %s"
        perf_params.append(end_month)

    perf_query += " ORDER BY year, month"
    cur.execute(perf_query, perf_params)
    perf_rows = cur.fetchall()

    if perf_rows:
        if is_monthly and len(perf_rows) == 1:
            perf_row = perf_rows[0]
            raw_score = float(perf_row['score']) if perf_row['score'] else 95
            grade = perf_row['grade'] if perf_row['grade'] else 'B+'
            perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
            performance_score = perf_result['radar_value']
        else:
            grade_list = [row['grade'] if row['grade'] else 'B+' for row in perf_rows]
            grade_dates = [f"{row['year']}-{row['month']:02d}" for row in perf_rows]
            perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
            performance_score = perf_result['radar_value']
    else:
        performance_score = 0

    # 4. 学习能力
    is_long_term = False
    if start_month and end_month and start_month != end_month:
        is_long_term = True
    
    learning_score = 0
    calculated_learning = False
    
    if is_long_term:
        try:
            monthly_counts = {}
            start_dt = datetime.strptime(start_month + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_month + '-01', '%Y-%m-%d')
            
            # 预查上月数据
            pre_period_month = (start_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
            pre_period_count = 0
            try:
                cur.execute("SELECT assessment FROM safety_inspection_records WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s", [emp_name, pre_period_month])
                pre_rows = cur.fetchall()
                pre_period_count = sum(1 for r in pre_rows if extract_score_from_assessment(r['assessment']) > 0)
            except Exception:
                pre_period_count = None

            curr = start_dt
            months_seq = []
            while curr <= end_dt:
                m_str = curr.strftime('%Y-%m')
                monthly_counts[m_str] = 0
                months_seq.append(m_str)
                curr = (curr.replace(day=1) + timedelta(days=32)).replace(day=1)
            
            # 填充违规计数
            for row in safety_rows:
                if extract_score_from_assessment(row['assessment']) > 0:
                    insp_date = row['inspection_date']
                    m_str = insp_date.strftime('%Y-%m') if hasattr(insp_date, 'strftime') else str(insp_date)[:7]
                    
                    if m_str in monthly_counts:
                        monthly_counts[m_str] += 1
            
            # 班组平均
            period_group_avg = 1.0 
            if dept_id:
                try:
                    cur.execute("SELECT COUNT(*) / COUNT(DISTINCT e.name) / %s as avg_viol FROM safety_inspection_records s JOIN employees e ON s.inspected_person = e.name WHERE e.department_id = %s AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') >= %s AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') <= %s", [max(1, len(months_seq)), dept_id, start_month, end_month])
                    avg_res = cur.fetchone()
                    if avg_res and avg_res['avg_viol']:
                         period_group_avg = float(avg_res['avg_viol'])
                except Exception:
                    pass

            monthly_scores = []
            last_violations = pre_period_count
            
            for m_str in months_seq:
                curr_viol = monthly_counts[m_str]
                res = calculate_learning_ability_new(curr_viol, last_violations, period_group_avg, algo_config)
                monthly_scores.append(res['learning_score'])
                last_violations = curr_viol
            
            # 加权平均
            total_weight = 0
            weighted_sum = 0
            time_decay_rate = algo_config.get('learning_new', {}).get('time_decay_rate', 0.2)
            for i, score in enumerate(monthly_scores):
                weight = 1.0 + (i * time_decay_rate) 
                weighted_sum += score * weight
                total_weight += weight
            
            learning_score = weighted_sum / total_weight if total_weight > 0 else 0
            calculated_learning = True
            
        except Exception as e:
            # Fallback
            pass
            
    if not calculated_learning:
        # 单月/Short term logic
        current_violations = 0
        if end_month:
            target_month = end_month
        else:
            target_month = datetime.now().strftime('%Y-%m')
            
        current_violations = sum(1 for r in list(filter(lambda x: x['inspection_date'].strftime('%Y-%m') == target_month, safety_rows)) if extract_score_from_assessment(x['assessment']) > 0)

        previous_violations = None
        if start_month:
            try:
                curr_dt = datetime.strptime(start_month + '-01', '%Y-%m-%d')
                prev_month = (curr_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
                cur.execute("SELECT assessment FROM safety_inspection_records WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s", [emp_name, prev_month])
                prev_rows = cur.fetchall()
                previous_violations = sum(1 for r in prev_rows if extract_score_from_assessment(r['assessment']) > 0)
            except Exception:
                pass
        
        group_avg = 1.0
        if dept_id and start_month:
            try:
                cur.execute("SELECT COUNT(*) / COUNT(DISTINCT e.name) as avg_viol FROM safety_inspection_records s JOIN employees e ON s.inspected_person = e.name WHERE e.department_id = %s AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') = %s", [dept_id, start_month])
                res = cur.fetchone()
                if res and res['avg_viol']: group_avg = float(res['avg_viol'])
            except Exception:
                pass
                
        l_res = calculate_learning_ability_new(current_violations, previous_violations, group_avg, algo_config)
        learning_score = l_res['learning_score']

    # 5. 稳定性（波动型）
    stability_result = calculate_stability_for_employee(
        emp_name,
        start_month,
        end_month,
        algo_config,
        cur,
        safety_score_for_tip=safety_score
    )
    stability_score = stability_result.get('stability_score', 50)
    if not isinstance(stability_score, (int, float)):
        stability_score = 50

    return {
        'performance': performance_score,
        'safety': safety_score,
        'training': training_score,
        'learning': round(float(learning_score), 1),
        'stability': purity_score(stability_score) # Just ensure float/int
    }

def purity_score(s):
    if isinstance(s, (int, float)): return round(float(s), 1)
    return 50.0



@personnel_bp.route('/api/students-list')
@login_required
def api_students_list():
    """API: 获取人员列表及综合评分（带权限过滤和关键人员标记）"""
    from datetime import datetime
    from services.domain.safety_utils import extract_score_from_assessment

    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    tr = parse_time_range(request.args, ['month'], default_grain='month', default_range='current_month')
    start_month = tr['start_month']
    end_month = tr['end_month']
    department_filter = request.args.get('department')
    position_filter = request.args.get('position')

    # 获取当前月份（用于关键人员标记）
    current_month = datetime.now().strftime('%Y-%m')

    # 读取算法配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    score_weights = algo_config['comprehensive']['score_weights']
    key_personnel_config = algo_config['key_personnel']

    # 安全获取字典值的辅助函数
    def safe_get(obj, key, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        else:
            try:
                return obj[key] if obj[key] is not None else default
            except (KeyError, IndexError):
                return default

    # 使用现有的 list_personnel() 函数获取权限过滤后的人员列表
    rows = list_personnel()

    # 应用部门和岗位筛选
    if department_filter:
        rows = [r for r in rows if safe_get(r, 'department_name') == department_filter]

    if position_filter:
        position_filter_lower = position_filter.lower()
        rows = [r for r in rows if position_filter_lower in (safe_get(r, 'position') or '').lower()]

    students = []
    for row in rows:
        emp_no = safe_get(row, 'emp_no')
        emp_name = safe_get(row, 'name')
        dept_id = safe_get(row, 'department_id')

        # 计算取证年限（用于培训和稳定性算法）
        cert_date = safe_get(row, 'certification_date')
        cert_years = calculate_years_from_date(cert_date) if cert_date else None

        # 获取部门名称
        if dept_id:
            cur.execute("SELECT name FROM departments WHERE id = %s", (dept_id,))
            dept_row = cur.fetchone()
            dept_name = dept_row['name'] if dept_row else None
        else:
            dept_name = None

        # 1. 培训能力（使用高级评分算法，应用日期筛选）
        training_query = """
            SELECT score, is_qualified, is_disqualified, training_date
            FROM training_records
            WHERE emp_no = %s
        """
        training_params = [emp_no]

        if start_month:
            training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') >= %s"
            training_params.append(start_month)

        if end_month:
            training_query += " AND DATE_FORMAT(training_date, '%%Y-%%m') <= %s"
            training_params.append(end_month)

        training_query += " ORDER BY training_date ASC"
        cur.execute(training_query, training_params)
        training_records_list = cur.fetchall()

        # 计算统计周期天数
        if start_month and end_month and start_month == end_month:
            duration_days = 30
        elif start_month and end_month:
            try:
                start_dt = datetime.strptime(start_month + '-01', '%Y-%m-%d')
                end_dt = datetime.strptime(end_month + '-01', '%Y-%m-%d')
                import calendar
                ey, em_i = int(end_month.split('-')[0]), int(end_month.split('-')[1])
                last_day = calendar.monthrange(ey, em_i)[1]
                end_dt = end_dt.replace(day=last_day)
                duration_days = max(1, (end_dt - start_dt).days + 1)
            except Exception:
                duration_days = 30
        else:
            duration_days = 30

        # 使用新的评分算法
        training_result = calculate_training_score_with_penalty(training_records_list, duration_days, cert_years, algo_config)
        training_score = training_result['radar_score']

        # 2. 安全意识（使用双轨评分模型）
        # 构建日期筛选条件
        safety_query = """
            SELECT assessment, inspection_date
            FROM safety_inspection_records
            WHERE inspected_person = %s
        """
        safety_params = [emp_name]

        if start_month:
            safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') >= %s"
            safety_params.append(start_month)

        if end_month:
            safety_query += " AND DATE_FORMAT(inspection_date, '%%Y-%%m') <= %s"
            safety_params.append(end_month)

        safety_query += " ORDER BY inspection_date ASC"
        cur.execute(safety_query, safety_params)
        safety_rows = cur.fetchall()

        # 收集所有违规扣分
        violations_list = []
        for s_row in safety_rows:
            assessment = s_row['assessment']
            score = extract_score_from_assessment(assessment)
            if score > 0:
                violations_list.append(float(score))

        # 计算统计周期月数
        months_active = 1
        if start_month and end_month:
            try:
                start = datetime.strptime(start_month + '-01', '%Y-%m-%d')
                end = datetime.strptime(end_month + '-01', '%Y-%m-%d')
                months_active = max(1, int((end - start).days / 30) + 1)
            except Exception:
                months_active = 1
        elif start_month:
            try:
                start = datetime.strptime(start_month + '-01', '%Y-%m-%d')
                months_active = max(1, int((datetime.now() - start).days / 30) + 1)
            except Exception:
                months_active = 1
        elif entry_date:
            # 没有日期筛选，使用入职以来的月数
            try:
                entry = datetime.strptime(entry_date, '%Y-%m-%d')
                months_active = max(1, int((datetime.now() - entry).days / 30))
            except Exception:
                months_active = 1

        # 使用双轨评分模型
        safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
        safety_score = safety_result['final_score']
        safety_status_color = safety_result['status_color']
        safety_alert_tag = safety_result['alert_tag']

        # 3. 工作绩效（使用双算法系统，应用日期筛选）
        is_monthly = (start_month == end_month) if start_month and end_month else True

        perf_query = """
            SELECT score, grade, year, month
            FROM performance_records
            WHERE emp_no = %s
        """
        perf_params = [emp_no]

        if start_month:
            perf_query += f" AND ({get_year_month_concat()}) >= %s"
            perf_params.append(start_month)

        if end_month:
            perf_query += f" AND ({get_year_month_concat()}) <= %s"
            perf_params.append(end_month)

        perf_query += " ORDER BY year, month"
        cur.execute(perf_query, perf_params)
        perf_rows = cur.fetchall()

        if perf_rows:
            if is_monthly and len(perf_rows) == 1:
                # 月度快照算法
                perf_row = perf_rows[0]
                raw_score = float(perf_row['score']) if perf_row['score'] else 95
                grade = perf_row['grade'] if perf_row['grade'] else 'B+'
                perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
                performance_score = perf_result['radar_value']
            else:
                # 周期加权算法（带时间衰减）
                grade_list = [row['grade'] if row['grade'] else 'B+' for row in perf_rows]
                grade_dates = [f"{row['year']}-{row['month']:02d}" for row in perf_rows]  # 构建日期列表
                perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
                performance_score = perf_result['radar_value']
        else:
            # 没有绩效数据
            performance_score = 0

        # 4. 学习能力评估（新版：基于违规数量变化趋势）
        current_violations = 0
        previous_violations = None
        learning_result = None  # 初始化，避免UnboundLocalError
        
        # 判断是否为长周期（跨月）
        is_long_term = False
        if start_month and end_month and start_month != end_month:
            is_long_term = True
            
        # 初始化班组平均违规数（默认值）
        group_avg_violations = 1.0
            
        if is_long_term:
            # 长周期模式：使用 calculate_learning_ability_longterm 保持与详情页一致
            try:
                # 1. 初始化每月计数
                monthly_counts = {}
                start_dt = datetime.strptime(start_month + '-01', '%Y-%m-%d')
                end_dt = datetime.strptime(end_month + '-01', '%Y-%m-%d')

                # 预先查询周期前一个月的数据（作为第一个月的previous基础）
                pre_period_month = (start_dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
                pre_period_count = 0
                try:
                    cur.execute("""
                        SELECT assessment FROM safety_inspection_records
                        WHERE inspected_person = %s AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s
                    """, [emp_name, pre_period_month])
                    pre_rows = cur.fetchall()
                    pre_period_count = sum(1 for r in pre_rows if extract_score_from_assessment(r['assessment']) > 0)
                except Exception:
                    pre_period_count = None # 无数据

                # 构建月份序列
                curr = start_dt
                months_seq = []
                while curr <= end_dt:
                    m_str = curr.strftime('%Y-%m')
                    monthly_counts[m_str] = 0
                    months_seq.append(m_str)
                    curr = (curr.replace(day=1) + timedelta(days=32)).replace(day=1)

                # 2. 填充数据
                for row in safety_rows:
                    if extract_score_from_assessment(row['assessment']) > 0:
                        insp_date = row['inspection_date']
                        m_str = insp_date.strftime('%Y-%m') if hasattr(insp_date, 'strftime') else str(insp_date)[:7]
                        if m_str in monthly_counts:
                            monthly_counts[m_str] += 1

                # 获取班组平均作为参考（使用周期内的整体平均）
                period_group_avg = 1.0
                if dept_id:
                     try:
                        cur.execute("""
                            SELECT COUNT(*) / COUNT(DISTINCT e.name) / %s as avg_viol
                            FROM safety_inspection_records s
                            JOIN employees e ON s.inspected_person = e.name
                            WHERE e.department_id = %s
                            AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') >= %s
                            AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') <= %s
                        """, [max(1, len(months_seq)), dept_id, start_month, end_month])
                        avg_res = cur.fetchone()
                        if avg_res and avg_res['avg_viol']:
                            period_group_avg = float(avg_res['avg_viol'])
                     except Exception:
                        pass

                # 3. 构建违规数量列表（按月份顺序）
                score_list = [monthly_counts[m] for m in months_seq]

                # 4. 调用 calculate_learning_ability_longterm（与详情页一致，包含风险惯性惩罚）
                learning_result = calculate_learning_ability_longterm(
                    score_list=score_list,
                    config=algo_config,
                    group_avg=period_group_avg,
                    initial_prev_viol=pre_period_count
                )

                # 补全部分前端需要的字段
                if learning_result['risk_level'] == 'SAFE':
                    learning_result['trend_type'] = 'safe'
                elif learning_result['risk_level'] in ['HIGH_RISK', 'PRE_ACCIDENT']:
                    learning_result['trend_type'] = 'deterioration'
                elif learning_result['risk_level'] == 'WATCH_LIST':
                    learning_result['trend_type'] = 'yellow_alert'
                else:
                    learning_result['trend_type'] = 'fluctuation'

                # 设置current/previous用于后续兼容（使用最后两个月数据）
                current_violations = monthly_counts[months_seq[-1]]
                if len(months_seq) >= 2:
                    previous_violations = monthly_counts[months_seq[-2]]
                else:
                    previous_violations = pre_period_count

            except Exception as e:
                current_app.logger.error(f": 长周期学习能力计算异常: {e}")
                is_long_term = False # 回退到单月模式


        if not is_long_term:
            # 单月模式：本月 vs 上月
            if end_month:
                end_target = end_month
                current_violations = sum(1 for r in safety_rows 
                                       if (r['inspection_date'].strftime('%Y-%m') if hasattr(r['inspection_date'], 'strftime') 
                                           else str(r['inspection_date'])[:7]) == end_target 
                                       and extract_score_from_assessment(r['assessment']) > 0)
            else:
                current_violations = len(violations_list)

            # 获取上月违规数
            if start_month:
                try:
                    current_dt = datetime.strptime(start_month + '-01', '%Y-%m-%d')
                    prev_dt = current_dt.replace(day=1) - timedelta(days=1)
                    prev_month = prev_dt.strftime('%Y-%m')

                    cur.execute("""
                        SELECT assessment FROM safety_inspection_records
                        WHERE inspected_person = %s
                        AND DATE_FORMAT(inspection_date, '%%Y-%%m') = %s
                    """, [emp_name, prev_month])
                    prev_rows = cur.fetchall()
                    previous_violations = sum(1 for r in prev_rows if extract_score_from_assessment(r['assessment']) > 0)
                except Exception:
                    previous_violations = None
        
        if not learning_result:
            # 获取班组平均违规数
            group_avg_violations = 1.0
            if dept_id and start_month:
                try:
                    cur.execute("""
                        SELECT COUNT(*) / COUNT(DISTINCT e.name) as avg_viol
                        FROM safety_inspection_records s
                        JOIN employees e ON s.inspected_person = e.name
                        WHERE e.department_id = %s
                        AND DATE_FORMAT(s.inspection_date, '%%Y-%%m') = %s
                    """, [dept_id, start_month])
                    avg_result = cur.fetchone()
                    if avg_result and avg_result['avg_viol']:
                        group_avg_violations = float(avg_result['avg_viol'])
                except Exception:
                    pass
            
            # 调用新算法
            learning_result = calculate_learning_ability_new(
                current_violations=current_violations,
                previous_violations=previous_violations,
                group_avg_violations=group_avg_violations,
                config=algo_config
            )
        learning_score = learning_result['learning_score']

        # 5. 稳定性（波动型）
        stability_result = calculate_stability_for_employee(
            emp_name,
            start_month,
            end_month,
            algo_config,
            cur,
            safety_score_for_tip=safety_score
        )
        stability_score = stability_result.get('stability_score', 50)

        # 异常情况：使用简单计算作为降级方案
        # Note: The new stability algorithm is designed to be robust.
        # If `calculate_stability_score_new` itself fails, it should handle its own defaults.
        # The previous fallback logic for `entry_date` is now less relevant
        # as the new algorithm doesn't primarily rely on `entry_date` for its core calculation.
        # However, if `stability_score` is still not set or is invalid, a final fallback can be applied.
        if stability_score is None:
            current_app.logger.debug(f" [api_students_list-员工{emp_no}]: 稳定性算法返回None，使用默认值50")
            stability_score = 50
        elif not isinstance(stability_score, (int, float)):
            current_app.logger.debug(f" [api_students_list-员工{emp_no}]: 稳定性算法返回非数值，使用默认值50")
            stability_score = 50

        # 综合评分（加权平均 - 使用配置权重）
        comprehensive_score = round(
            performance_score * score_weights['performance'] +
            safety_score * score_weights['safety'] +
            training_score * score_weights['training'] +
            stability_score * score_weights['stability'] +
            learning_score * score_weights['learning'],
            1
        )

        # 判断是否为关键人员（基于筛选日期范围）（使用配置阈值）
        # 复用已计算的违规数据和月数，避免重复查询
        import math
        violation_count = len(violations_list)
        avg_freq = math.ceil(violation_count / months_active) if months_active > 0 else 0

        is_key_personnel = (comprehensive_score < key_personnel_config['comprehensive_threshold']) or (avg_freq >= key_personnel_config['monthly_violation_threshold'])

        students.append({
            'emp_no': emp_no,
            'name': emp_name,
            'department_name': dept_name,
            'position': safe_get(row, 'position'),
            'comprehensive_score': comprehensive_score,
            'is_key_personnel': bool(is_key_personnel),  # 显式转换为JSON兼容的布尔值
            'safety_status_color': safety_status_color,
            'safety_alert_tag': safety_alert_tag
        })

    # 按综合分升序排序
    students.sort(key=lambda x: x['comprehensive_score'])

    return jsonify(students)
@personnel_bp.route('/api/student-detail/<emp_no>')
@login_required
def api_student_detail(emp_no):
    """API: 获取学员详细数据（培训雷达图数据：按项目分类的平均分）"""
    conn = get_db()
    cur = conn.cursor()

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # 获取时间筛选参数
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    # 构建时间筛选条件
    time_filter = ""
    time_params = [emp_no]
    if year and month:
        time_filter = " AND YEAR(training_date) = %s AND MONTH(training_date) = %s"
        time_params.extend([str(year), str(month).zfill(2)])
    elif year:
        time_filter = " AND YEAR(training_date) = %s"
        time_params.append(str(year))

    # 查询该学员各项目分类的平均分
    query = f"""
        SELECT
            c.name as category_name,
            ROUND(AVG(tr.score), 1) as avg_score,
            COUNT(*) as count
        FROM training_records tr
        LEFT JOIN training_projects p ON tr.project_id = p.id
        LEFT JOIN training_project_categories c ON p.category_id = c.id
        WHERE tr.emp_no = %s AND c.name IS NOT NULL{time_filter}
        GROUP BY c.id, c.name
        ORDER BY c.display_order ASC
    """
    cur.execute(query, time_params)
    student_data = {}
    for row in cur.fetchall():
        student_data[row['category_name']] = {
            'avg_score': row['avg_score'],
            'count': row['count']
        }

    # 查询团队平均（基于权限过滤的可见员工）
    accessible_dept_ids = get_accessible_department_ids()
    if not accessible_dept_ids:
        return jsonify({
            'student_data': student_data,
            'team_data': {},
            'categories': sorted(list(student_data.keys()))
        })

    placeholders = ','.join(['%s'] * len(accessible_dept_ids))

    # 构建团队查询的时间筛选
    team_time_filter = ""
    team_time_params = accessible_dept_ids.copy()
    if year and month:
        team_time_filter = " AND YEAR(tr.training_date) = %s AND MONTH(tr.training_date) = %s"
        team_time_params.extend([str(year), str(month).zfill(2)])
    elif year:
        team_time_filter = " AND YEAR(tr.training_date) = %s"
        team_time_params.append(str(year))

    query = f"""
        SELECT
            c.name as category_name,
            ROUND(AVG(tr.score), 1) as avg_score
        FROM training_records tr
        LEFT JOIN employees e ON tr.emp_no = e.emp_no
        LEFT JOIN training_projects p ON tr.project_id = p.id
        LEFT JOIN training_project_categories c ON p.category_id = c.id
        WHERE (e.department_id IN ({placeholders}) OR e.emp_no IS NULL)
            AND c.name IS NOT NULL{team_time_filter}
        GROUP BY c.id, c.name
        ORDER BY c.display_order ASC
    """
    cur.execute(query, team_time_params)
    team_data = {}
    for row in cur.fetchall():
        team_data[row['category_name']] = row['avg_score']

    # 合并所有分类
    all_categories = set(student_data.keys()) | set(team_data.keys())

    return jsonify({
        'student_data': student_data,
        'team_data': team_data,
        'categories': sorted(list(all_categories))
    })


@personnel_bp.route('/api/student-growth/<emp_no>')
@login_required
def api_student_growth(emp_no):
    """API: 获取学员成长趋势数据（按时间的平均分变化）"""
    conn = get_db()
    cur = conn.cursor()

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # 查询该学员按月份的平均分趋势
    query = """
        SELECT
            DATE_FORMAT(training_date, '%%Y-%%m') as month,
            ROUND(AVG(score), 1) as avg_score,
            COUNT(*) as count
        FROM training_records
        WHERE emp_no = %s
        GROUP BY month
        ORDER BY month ASC
    """
    cur.execute(query, (emp_no,))

    growth_data = []
    for row in cur.fetchall():
        growth_data.append({
            'month': row['month'],
            'avg_score': row['avg_score'],
            'count': row['count']
        })

    return jsonify(growth_data)


# ============================================================
# Risk Mining API - 风险挖掘接口
