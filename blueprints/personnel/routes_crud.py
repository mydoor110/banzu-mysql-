#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""人员管理 - CRUD 路由（列表/模板/导入/详情/更新/删除）"""
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session, current_app

from config.settings import APP_TITLE, EXPORT_DIR
from models.database import get_db, close_db, get_year_month_concat
from ..decorators import login_required, manager_required
from ..helpers import (
    current_user_id, require_user_id, get_accessible_department_ids,
    get_accessible_departments, calculate_years_from_date, get_user_department,
    validate_employee_access, log_import_operation
)
from . import personnel_bp
from . import (
    PERSONNEL_FIELD_SCHEME, PERSONNEL_DB_COLUMNS, PERSONNEL_DATE_FIELDS,
    PERSONNEL_SELECT_OPTIONS, PERSONNEL_IMPORT_HEADER_MAP,
    list_personnel, get_personnel, upsert_personnel, bulk_import_personnel,
    update_personnel_field, delete_employee, _serialize_person,
    _build_personnel_charts, _sanitize_person_payload
)
import json
import os
import pymysql
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook

@personnel_bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    """人员管理首页"""
    if request.method == 'POST':
        # 🔒 权限检查: 创建/更新员工需要管理员权限
        from flask import session
        user_role = session.get('role', 'user')
        if user_role not in ['admin', 'manager']:
            flash("您没有权限执行此操作，需要部门管理员或系统管理员权限", "danger")
            return redirect(url_for("personnel.index"))

        form_payload = {}
        for field in PERSONNEL_FIELD_SCHEME:
            key = field["name"]
            if field["input_type"] == "textarea":
                form_payload[key] = request.form.get(key, "")
            else:
                form_payload[key] = request.form.get(key, "").strip()
        saved = upsert_personnel(form_payload)
        if saved:
            flash("人员信息已保存。", "success")
        else:
            flash("请填写有效的工号和姓名。", "warning")
        return redirect(url_for("personnel.index"))

    rows = list_personnel()
    accessible_departments = get_accessible_departments()

    return render_template(
        "personnel.html",
        title=f"人员管理 | {APP_TITLE}",
        rows=rows,
        field_scheme=PERSONNEL_FIELD_SCHEME,
        select_options=PERSONNEL_SELECT_OPTIONS,
        accessible_departments=accessible_departments,
    )


@personnel_bp.route('/template')
@login_required
def template():
    """下载人员导入模板"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人员导入模板"

    headers = [field["label"] for field in PERSONNEL_FIELD_SCHEME]
    sheet.append(headers)

    examples = {
        "emp_no": "1001",
        "name": "张三",
        "class_name": "一班",
        "position": "班长",
        "birth_date": "1990-01-01",
        "marital_status": "已婚",
        "hometown": "江苏南京",
        "political_status": "群众",
        "education": "本科",
        "graduation_school": "某某大学",
        "work_start_date": "2012-07-01",
        "entry_date": "2018-03-15",
        "specialty": "摄影、篮球",
    }
    sheet.append([examples.get(field["name"], "") for field in PERSONNEL_FIELD_SCHEME])

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"personnel_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@personnel_bp.route('/import', methods=['POST'])
@manager_required
def import_data():
    """批量导入人员数据"""
    file_obj = request.files.get("file")
    if not file_obj or file_obj.filename == "":
        flash("请选择包含花名册数据的 Excel 文件。", "warning")
        return redirect(url_for("personnel.index"))
    ext = file_obj.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx"}:
        flash("目前仅支持上传 .xlsx 文件。", "warning")
        return redirect(url_for("personnel.index"))
    try:
        workbook = load_workbook(file_obj, data_only=True)
        sheet = workbook.active
    except Exception as exc:  # noqa: BLE001
        flash(f"无法读取 Excel 文件：{exc}", "danger")
        return redirect(url_for("personnel.index"))

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        flash("Excel 文件为空。", "warning")
        return redirect(url_for("personnel.index"))

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    field_map = [PERSONNEL_IMPORT_HEADER_MAP.get(header) for header in headers]

    if "emp_no" not in field_map or "name" not in field_map:
        flash('Excel 首行必须包含"工号"与"姓名"列。', "warning")
        return redirect(url_for("personnel.index"))

    # 获取部门映射，用于处理Excel中的部门信息
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM departments")
    dept_name_map = {row['name']: row['id'] for row in cur.fetchall()}

    # 获取当前用户可访问的部门ID列表（用于权限验证）
    accessible_dept_ids = get_accessible_department_ids()

    records: List[Dict[str, Optional[str]]] = []
    skipped_no_dept = 0
    skipped_no_permission = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: Dict[str, Optional[str]] = {}
        for idx, cell in enumerate(row):
            field = field_map[idx] if idx < len(field_map) else None
            if not field:
                continue
            record[field] = cell
            
        # 处理部门ID：支持名称匹配
        raw_dept = record.get('department_id')
        final_dept_id = None

        if raw_dept:
            raw_dept_str = str(raw_dept).strip()
            if raw_dept_str.isdigit():
                final_dept_id = int(raw_dept_str)
            elif raw_dept_str in dept_name_map:
                final_dept_id = dept_name_map[raw_dept_str]

        if not final_dept_id:
            # 未填写部门或部门无效
            skipped_no_dept += 1
        elif final_dept_id not in accessible_dept_ids:
            # 部门存在但无权限导入到该部门
            skipped_no_permission += 1
        else:
            # 部门有效且有权限
            record['department_id'] = str(final_dept_id)
            records.append(record)

    if not records:
        msg_parts = ["未导入任何数据。"]
        if skipped_no_dept > 0:
            msg_parts.append(f"{skipped_no_dept} 条记录因未填写部门或部门无效被跳过。")
        if skipped_no_permission > 0:
            msg_parts.append(f"{skipped_no_permission} 条记录因无权限导入到该部门被跳过。")
        if not skipped_no_dept and not skipped_no_permission:
            msg_parts.append("未识别到任何有效行。")
        flash(" ".join(msg_parts), "warning")

        # 记录失败的导入操作
        log_import_operation(
            module='personnel',
            operation='import',
            file_name=file_obj.filename,
            total_rows=skipped_no_dept + skipped_no_permission,
            success_rows=0,
            failed_rows=0,
            skipped_rows=skipped_no_dept + skipped_no_permission,
            error_message=" ".join(msg_parts),
            import_details={
                'skipped_no_dept': skipped_no_dept,
                'skipped_no_permission': skipped_no_permission
            }
        )
        return redirect(url_for("personnel.index"))

    imported = bulk_import_personnel(records)

    # 计算总行数
    total_rows = len(records) + skipped_no_dept + skipped_no_permission

    # 构建提示消息
    msg = f"已导入/更新 {imported} 名员工信息。"
    msg_parts = []
    if skipped_no_dept > 0:
        msg_parts.append(f"{skipped_no_dept} 条记录因未填写部门或部门无效被跳过")
    if skipped_no_permission > 0:
        msg_parts.append(f"{skipped_no_permission} 条记录因无权限导入到该部门被跳过")

    if msg_parts:
        msg += " 另有 " + "、".join(msg_parts) + "。"
        flash_type = "warning"
    else:
        flash_type = "success"

    flash(msg, flash_type)

    # 记录导入操作日志
    log_import_operation(
        module='personnel',
        operation='import',
        file_name=file_obj.filename,
        total_rows=total_rows,
        success_rows=imported,
        failed_rows=0,
        skipped_rows=skipped_no_dept + skipped_no_permission,
        import_details={
            'imported': imported,
            'skipped_no_dept': skipped_no_dept,
            'skipped_no_permission': skipped_no_permission,
            'accessible_departments': len(accessible_dept_ids)
        }
    )

    return redirect(url_for("personnel.index"))


@personnel_bp.route('/<emp_no>')
@login_required
def preview(emp_no):
    """查看人员详情"""
    person = get_personnel(emp_no)
    if not person:
        flash("未找到该员工。", "warning")
        return redirect(url_for("personnel.index"))
    return render_template(
        "personnel_preview.html",
        title=f"{person.get('name', '')} | 人员档案 · {APP_TITLE}",
        person=person,
        field_scheme=PERSONNEL_FIELD_SCHEME,
        select_options=PERSONNEL_SELECT_OPTIONS,
    )


@personnel_bp.route('/<emp_no>/update', methods=['POST'])
@login_required
@manager_required
def update(emp_no):
    """更新人员信息字段（仅限部门管理员及以上权限）"""
    payload = request.get_json(silent=True) or request.form
    field = (payload.get("field") or "").strip()
    value = payload.get("value")
    if field in PERSONNEL_DATE_FIELDS and isinstance(value, str):
        value = value.strip()
    if not field:
        return jsonify({"ok": False, "message": "未指定字段"}), 400
    updated = update_personnel_field(emp_no, field, value)
    if not updated:
        return jsonify({"ok": False, "message": "更新失败或字段不受支持"}), 400
    person = get_personnel(emp_no)
    return jsonify({"ok": True, "person": person})


@personnel_bp.route('/batch-delete', methods=['POST'])
@login_required
@manager_required
def batch_delete():
    """批量删除员工（仅限部门管理员及以上权限）"""
    emp_nos = request.form.getlist('emp_nos')

    if not emp_nos:
        flash("未选择要删除的员工", "warning")
        return redirect(url_for("personnel.index"))

    uid = require_user_id()
    from flask import session
    user_role = session.get('role', 'user')

    conn = get_db()
    cur = conn.cursor()

    deleted_count = 0
    skipped_count = 0
    for emp_no in emp_nos:
        emp_no = emp_no.strip()
        if emp_no:
            # 🔒 权限检查: 非管理员需要验证是否有权删除每个员工
            if user_role != 'admin':
                if not validate_employee_access(emp_no):
                    skipped_count += 1
                    continue

            cur.execute("DELETE FROM employees WHERE emp_no=%s", (emp_no,))
            if cur.rowcount > 0:
                deleted_count += 1

    conn.commit()

    if deleted_count > 0:
        message = f"成功删除 {deleted_count} 名员工"
        if skipped_count > 0:
            message += f"，跳过 {skipped_count} 名无权删除的员工"
        flash(message, "success")
    elif skipped_count > 0:
        flash(f"跳过 {skipped_count} 名无权删除的员工", "warning")
    else:
        flash("未删除任何员工", "info")

    return redirect(url_for("personnel.index"))


