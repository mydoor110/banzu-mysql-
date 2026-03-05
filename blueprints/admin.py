#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
系统管理模块
负责用户管理、备份管理等系统级功能
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from werkzeug.security import generate_password_hash
import pymysql
import json
from datetime import datetime, timedelta
from models.database import get_db
from utils.backup import BackupManager, get_backup_statistics, BackupTaskManager
from .decorators import admin_required
from openpyxl import Workbook, load_workbook
import csv
import io
import os
from config.settings import EXPORT_DIR
from blueprints.helpers import parse_time_range, build_date_filter_sql

# 创建 Blueprint
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')
APP_TITLE = "乘务数字化管理平台"

_USER_IMPORT_ALIASES = {
    'username': ('username', '用户名', '账号', '登录名', 'user_name'),
    'password': ('password', '密码', '初始密码', '登录密码'),
    'display_name': ('display_name', '姓名', '真实姓名', 'name'),
    'role': ('role', '角色', '用户角色', '权限'),
    'department_id': ('department_id', '部门id', '所属部门id'),
    'department_name': ('department_name', '部门名称', '所属部门', '部门'),
    'dingtalk_userid': ('dingtalk_userid', '钉钉userid', '钉钉用户id', '钉钉userid'),
    'dingtalk_unionid': ('dingtalk_unionid', '钉钉unionid', 'unionid'),
}

_ROLE_ALIAS = {
    'admin': 'admin',
    '系统管理员': 'admin',
    '管理员': 'admin',
    'manager': 'manager',
    '部门管理员': 'manager',
    'user': 'user',
    '普通用户': 'user',
}


def _normalize_import_text(value):
    text = '' if value is None else str(value).strip()
    if not text:
        return ''
    text = text.lower()
    for ch in (' ', '\t', '\n', '_', '-'):
        text = text.replace(ch, '')
    return text


def _normalize_cell_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_user_import_header_map():
    mapping = {}
    for canonical, aliases in _USER_IMPORT_ALIASES.items():
        for alias in aliases:
            mapping[_normalize_import_text(alias)] = canonical
    return mapping


_USER_IMPORT_HEADER_MAP = _build_user_import_header_map()


def _parse_user_import_rows(file_storage):
    filename = (file_storage.filename or '').lower()
    if filename.endswith('.csv'):
        return _parse_user_import_rows_from_csv(file_storage)
    return _parse_user_import_rows_from_excel(file_storage)


def _parse_user_import_rows_from_excel(file_storage):
    try:
        wb = load_workbook(file_storage, data_only=True)
    except Exception as e:
        raise ValueError(f'Excel 解析失败: {e}') from e

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError('导入文件缺少表头')

    canonical_headers = []
    for header in headers:
        canonical_headers.append(
            _USER_IMPORT_HEADER_MAP.get(_normalize_import_text(header), '')
        )
    if not any(canonical_headers):
        raise ValueError('未识别到可导入字段，请使用模板文件')

    parsed_rows = []
    for row_no, row in enumerate(rows, start=2):
        parsed = {}
        for idx, canonical in enumerate(canonical_headers):
            if not canonical:
                continue
            cell = row[idx] if idx < len(row) else None
            value = _normalize_cell_value(cell)
            if value:
                parsed[canonical] = value
        parsed_rows.append((row_no, parsed))
    return parsed_rows


def _parse_user_import_rows_from_csv(file_storage):
    raw = file_storage.read()
    text = None
    for encoding in ('utf-8-sig', 'gb18030', 'utf-8'):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError('CSV 编码无法识别，请使用 UTF-8 或 GBK')

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError('CSV 缺少表头')

    if not any(_USER_IMPORT_HEADER_MAP.get(_normalize_import_text(h), '') for h in reader.fieldnames):
        raise ValueError('未识别到可导入字段，请使用模板文件')

    parsed_rows = []
    for row_no, row in enumerate(reader, start=2):
        parsed = {}
        for header, value in row.items():
            canonical = _USER_IMPORT_HEADER_MAP.get(_normalize_import_text(header), '')
            if not canonical:
                continue
            normalized = _normalize_cell_value(value)
            if normalized:
                parsed[canonical] = normalized
        parsed_rows.append((row_no, parsed))
    return parsed_rows


def _normalize_role(role_raw):
    if role_raw is None:
        return ''
    role_text = str(role_raw).strip()
    if not role_text:
        return ''
    role = _ROLE_ALIAS.get(role_text)
    if role:
        return role

    role = _ROLE_ALIAS.get(_normalize_import_text(role_text))
    if role:
        return role

    raise ValueError(f'角色无效: {role_text}')


def _resolve_department_id(department_id_raw, department_name_raw, department_by_id, department_name_map):
    if department_id_raw:
        dept_id_text = str(department_id_raw).strip()
        try:
            department_id = int(float(dept_id_text))
        except (TypeError, ValueError):
            raise ValueError(f'部门ID格式不正确: {dept_id_text}')
        if department_id not in department_by_id:
            raise ValueError(f'部门ID不存在: {department_id}')
        return department_id

    if department_name_raw:
        normalized_name = _normalize_import_text(department_name_raw)
        dept_ids = department_name_map.get(normalized_name, [])
        if not dept_ids:
            raise ValueError(f'部门不存在: {department_name_raw}')
        if len(dept_ids) > 1:
            raise ValueError(f'部门名称重复，请改用部门ID: {department_name_raw}')
        return dept_ids[0]

    return None

# ========== 用户管理 ==========

@admin_bp.route('/users', methods=['GET', 'POST'])
@admin_required
def users():
    """用户管理页面"""
    conn = get_db()
    cur = conn.cursor()
    # action可能来自URL参数(GET)或表单数据(POST)
    action = request.args.get('action') or request.form.get('action')

    if request.method == 'POST' and action == 'batch_import':
        import_file = request.files.get('import_file')
        if not import_file or not import_file.filename:
            flash('请选择导入文件（.xlsx 或 .csv）', 'warning')
            return redirect(url_for('admin.users'))

        try:
            parsed_rows = _parse_user_import_rows(import_file)
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('admin.users'))

        if not parsed_rows:
            flash('导入文件没有可处理的数据行', 'warning')
            return redirect(url_for('admin.users'))

        cur.execute("SELECT id, name FROM departments")
        departments = cur.fetchall()
        department_by_id = {row['id']: row['name'] for row in departments}
        department_name_map = {}
        for row in departments:
            key = _normalize_import_text(row['name'])
            department_name_map.setdefault(key, []).append(row['id'])

        cur.execute("""
            SELECT id, username, password_hash, department_id, role, display_name,
                   dingtalk_userid, dingtalk_unionid
            FROM users
        """)
        existing_users = cur.fetchall()
        users_by_id = {row['id']: dict(row) for row in existing_users}
        users_by_username = {}
        users_by_dingtalk = {}
        for row in existing_users:
            users_by_username[row['username']] = row['id']
            if row.get('dingtalk_userid'):
                users_by_dingtalk[row['dingtalk_userid']] = row['id']

        total_rows = len(parsed_rows)
        success_rows = 0
        failed_rows = 0
        skipped_rows = 0
        detail_items = []

        for row_no, row in parsed_rows:
            if not row:
                skipped_rows += 1
                continue

            try:
                username = row.get('username', '').strip()
                dingtalk_userid = row.get('dingtalk_userid', '').strip()
                if not username and dingtalk_userid:
                    username = dingtalk_userid
                if not username:
                    raise ValueError('缺少用户名（可留空并填写钉钉UserId自动生成）')

                role_specified = bool(row.get('role', '').strip())
                role = _normalize_role(row.get('role', '')) if role_specified else ''

                department_id_raw = row.get('department_id', '').strip()
                department_name_raw = row.get('department_name', '').strip()
                has_department = bool(department_id_raw or department_name_raw)
                department_id = _resolve_department_id(
                    department_id_raw,
                    department_name_raw,
                    department_by_id,
                    department_name_map
                ) if has_department else None

                password_raw = row.get('password', '').strip()
                display_name = row.get('display_name', '').strip()
                dingtalk_unionid = row.get('dingtalk_unionid', '').strip()

                user_id_by_username = users_by_username.get(username)
                user_id_by_dingtalk = users_by_dingtalk.get(dingtalk_userid) if dingtalk_userid else None
                if user_id_by_username and user_id_by_dingtalk and user_id_by_username != user_id_by_dingtalk:
                    raise ValueError('用户名与钉钉UserId分别匹配到不同用户，无法自动合并')

                target_user_id = user_id_by_username or user_id_by_dingtalk
                if target_user_id:
                    old_user = users_by_id[target_user_id]
                    old_dept_id = old_user.get('department_id')
                    old_role = old_user.get('role')

                    new_username = username
                    new_password_hash = generate_password_hash(password_raw) if password_raw else old_user['password_hash']
                    new_department_id = department_id if has_department else old_dept_id
                    new_role = role if role_specified else (old_role or 'user')
                    new_display_name = display_name if display_name else old_user.get('display_name')
                    new_dingtalk_userid = dingtalk_userid if dingtalk_userid else old_user.get('dingtalk_userid')
                    new_dingtalk_unionid = dingtalk_unionid if dingtalk_unionid else old_user.get('dingtalk_unionid')

                    cur.execute(
                        """
                        UPDATE users
                        SET username=%s, password_hash=%s, department_id=%s, role=%s,
                            display_name=%s, dingtalk_userid=%s, dingtalk_unionid=%s
                        WHERE id=%s
                        """,
                        (
                            new_username,
                            new_password_hash,
                            new_department_id,
                            new_role,
                            new_display_name,
                            new_dingtalk_userid,
                            new_dingtalk_unionid,
                            target_user_id
                        )
                    )

                    if old_role == 'manager' and old_dept_id and (new_role != 'manager' or old_dept_id != new_department_id):
                        cur.execute(
                            "UPDATE departments SET manager_user_id = NULL WHERE id = %s AND manager_user_id = %s",
                            (old_dept_id, target_user_id)
                        )
                    if new_role == 'manager' and new_department_id:
                        cur.execute(
                            "UPDATE departments SET manager_user_id = %s WHERE id = %s",
                            (target_user_id, new_department_id)
                        )

                    old_username = old_user['username']
                    old_dingtalk_userid = old_user.get('dingtalk_userid')
                    if old_username != new_username:
                        users_by_username.pop(old_username, None)
                    if old_dingtalk_userid and old_dingtalk_userid != new_dingtalk_userid:
                        users_by_dingtalk.pop(old_dingtalk_userid, None)

                    users_by_id[target_user_id] = {
                        'id': target_user_id,
                        'username': new_username,
                        'password_hash': new_password_hash,
                        'department_id': new_department_id,
                        'role': new_role,
                        'display_name': new_display_name,
                        'dingtalk_userid': new_dingtalk_userid,
                        'dingtalk_unionid': new_dingtalk_unionid,
                    }
                    users_by_username[new_username] = target_user_id
                    if new_dingtalk_userid:
                        users_by_dingtalk[new_dingtalk_userid] = target_user_id

                    success_rows += 1
                    detail_items.append({'row': row_no, 'username': new_username, 'action': 'updated'})
                else:
                    new_role = role if role_specified else 'user'
                    new_password = password_raw or '123456'
                    new_password_hash = generate_password_hash(new_password)
                    cur.execute(
                        """
                        INSERT INTO users(
                            username, password_hash, department_id, role,
                            display_name, dingtalk_userid, dingtalk_unionid
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            username,
                            new_password_hash,
                            department_id,
                            new_role,
                            display_name or None,
                            dingtalk_userid or None,
                            dingtalk_unionid or None
                        )
                    )
                    new_user_id = cur.lastrowid

                    if new_role == 'manager' and department_id:
                        cur.execute(
                            "UPDATE departments SET manager_user_id = %s WHERE id = %s",
                            (new_user_id, department_id)
                        )

                    users_by_id[new_user_id] = {
                        'id': new_user_id,
                        'username': username,
                        'password_hash': new_password_hash,
                        'department_id': department_id,
                        'role': new_role,
                        'display_name': display_name or None,
                        'dingtalk_userid': dingtalk_userid or None,
                        'dingtalk_unionid': dingtalk_unionid or None,
                    }
                    users_by_username[username] = new_user_id
                    if dingtalk_userid:
                        users_by_dingtalk[dingtalk_userid] = new_user_id

                    success_rows += 1
                    detail_items.append({'row': row_no, 'username': username, 'action': 'created'})

            except (ValueError, pymysql.err.IntegrityError) as e:
                failed_rows += 1
                detail_items.append({'row': row_no, 'error': str(e), 'action': 'failed'})
            except Exception as e:
                failed_rows += 1
                detail_items.append({'row': row_no, 'error': f'系统错误: {e}', 'action': 'failed'})

        conn.commit()

        try:
            cur.execute(
                """
                INSERT INTO import_logs (
                    module, operation, user_id, username, user_role,
                    department_id, department_name, file_name,
                    total_rows, success_rows, failed_rows, skipped_rows,
                    error_message, import_details, ip_address, created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """,
                (
                    'admin',
                    'user_batch_import',
                    session.get('user_id'),
                    session.get('username') or 'unknown',
                    session.get('role') or 'admin',
                    None,
                    None,
                    import_file.filename,
                    total_rows,
                    success_rows,
                    failed_rows,
                    skipped_rows,
                    '' if failed_rows == 0 else f'存在 {failed_rows} 行导入失败',
                    json.dumps(detail_items[:300], ensure_ascii=False),
                    request.remote_addr
                )
            )
            conn.commit()
        except Exception:
            conn.rollback()

        flash(
            f'批量导入完成：总计 {total_rows} 行，成功 {success_rows} 行，失败 {failed_rows} 行，跳过 {skipped_rows} 行',
            'success' if failed_rows == 0 else 'warning'
        )
        return redirect(url_for('admin.users'))

    if request.method == 'POST' and not action:
        # 创建用户
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        department_id = request.form.get('department_id')
        role = request.form.get('role', 'user')
        display_name = request.form.get('display_name', '').strip()
        dingtalk_userid = request.form.get('dingtalk_userid', '').strip()
        dingtalk_unionid = request.form.get('dingtalk_unionid', '').strip()

        if not username or not password:
            flash('请输入用户名和密码', 'warning')
        else:
            try:
                department_id = int(department_id) if department_id else None
                cur.execute(
                    """
                    INSERT INTO users(
                        username, password_hash, department_id, role,
                        display_name, dingtalk_userid, dingtalk_unionid
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        username,
                        generate_password_hash(password),
                        department_id,
                        role,
                        display_name or None,
                        dingtalk_userid or None,
                        dingtalk_unionid or None
                    )
                )
                new_user_id = cur.lastrowid

                # 如果角色是部门管理员且指定了部门,同步更新部门的负责人
                if role == 'manager' and department_id:
                    cur.execute(
                        "UPDATE departments SET manager_user_id = %s WHERE id = %s",
                        (new_user_id, department_id)
                    )

                conn.commit()
                flash('用户已创建', 'success')
            except pymysql.err.IntegrityError:
                flash('用户名或钉钉UserId已存在', 'danger')
            except Exception as e:
                flash(f'创建失败: {e}', 'danger')

    # 处理其他操作
    if action == 'edit_user':
        user_id = request.form.get('user_id', type=int)
        username = request.form.get('username', '').strip()
        department_id = request.form.get('department_id')
        role = request.form.get('role', 'user')
        display_name = request.form.get('display_name', '').strip()
        dingtalk_userid = request.form.get('dingtalk_userid', '').strip()
        dingtalk_unionid = request.form.get('dingtalk_unionid', '').strip()
        if username:
            try:
                # 获取用户的旧信息
                cur.execute("SELECT department_id, role FROM users WHERE id=%s", (user_id,))
                old_info = cur.fetchone()
                old_dept_id = old_info['department_id'] if old_info else None
                old_role = old_info['role'] if old_info else None

                department_id = int(department_id) if department_id else None
                cur.execute(
                    """
                    UPDATE users
                    SET username=%s, department_id=%s, role=%s,
                        display_name=%s, dingtalk_userid=%s, dingtalk_unionid=%s
                    WHERE id=%s
                    """,
                    (
                        username,
                        department_id,
                        role,
                        display_name or None,
                        dingtalk_userid or None,
                        dingtalk_unionid or None,
                        user_id
                    )
                )

                # 双向同步部门负责人
                # 1. 如果从manager变为非manager,清除原部门的负责人设置
                if old_role == 'manager' and role != 'manager' and old_dept_id:
                    cur.execute(
                        "UPDATE departments SET manager_user_id = NULL WHERE id = %s AND manager_user_id = %s",
                        (old_dept_id, user_id)
                    )

                # 2. 如果用户换了部门且之前是manager,清除旧部门的负责人
                if old_role == 'manager' and old_dept_id and old_dept_id != department_id:
                    cur.execute(
                        "UPDATE departments SET manager_user_id = NULL WHERE id = %s AND manager_user_id = %s",
                        (old_dept_id, user_id)
                    )

                # 3. 如果新角色是manager且有部门,设置为该部门的负责人
                if role == 'manager' and department_id:
                    cur.execute(
                        "UPDATE departments SET manager_user_id = %s WHERE id = %s",
                        (user_id, department_id)
                    )

                conn.commit()
                flash('用户信息更新成功', 'success')
            except pymysql.err.IntegrityError:
                flash('用户名或钉钉UserId已存在', 'danger')
        return redirect(url_for('admin.users'))

    elif action == 'reset':
        user_id = request.args.get('id', type=int)
        new_pw = request.args.get('newpw', default='123456')
        if user_id:
            cur.execute("UPDATE users SET password_hash=%s WHERE id=%s",
                       (generate_password_hash(new_pw), user_id))
            conn.commit()
            flash(f'已重置用户 {user_id} 密码', 'success')
        return redirect(url_for('admin.users'))

    elif action == 'delete':
        if request.method != 'POST':
            flash('删除操作必须使用 POST 请求', 'danger')
            return redirect(url_for('admin.users'))
            
        user_id = request.form.get('id', type=int)
        if user_id == 1:
            flash('默认超级管理员(ID=1)不可删除', 'warning')
        elif user_id:
            # 防止删除当前登录用户
            if session.get('user_id') == user_id:
                flash('不可删除当前正在使用的账号', 'warning')
                return redirect(url_for('admin.users'))
                
            # 防止删除最后一个系统管理员
            cur.execute("SELECT COUNT(*) as admin_count FROM users WHERE role = 'admin'")
            admin_count = cur.fetchone()['admin_count']
            cur.execute("SELECT role, username FROM users WHERE id = %s", (user_id,))
            target_user = cur.fetchone()
            
            if target_user and target_user['role'] == 'admin' and admin_count <= 1:
                flash('系统中必须至少保留一个系统管理员', 'warning')
                return redirect(url_for('admin.users'))
                
            cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
            conn.commit()
            flash(f'已删除用户 {target_user["username"] if target_user else user_id}', 'success')
            
            # 记录安全日志
            try:
                from utils.logger import SecurityLogger
                SecurityLogger.suspicious_activity('user_deleted', {'deleted_user_id': user_id, 'by_admin_id': session.get('user_id')})
            except Exception:
                pass
                
        return redirect(url_for('admin.users'))

    # 获取用户列表
    cur.execute("""
        SELECT u.id, u.username, u.created_at, u.role,
               u.department_id, u.display_name, u.dingtalk_userid, u.dingtalk_unionid,
               d.name as department_name
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        ORDER BY u.id
    """)
    users_list = cur.fetchall()

    # 获取部门列表
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments_list = cur.fetchall()

    return render_template('admin_users.html', title='用户管理 | ' + APP_TITLE,
                         users=users_list, departments=departments_list)


@admin_bp.route('/users/import-template')
@admin_required
def download_users_import_template():
    """下载批量导入用户模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'users_import_template'

    headers = [
        '用户名', '密码', '姓名', '角色',
        '所属部门ID', '所属部门',
        '钉钉UserId', '钉钉UnionId'
    ]
    ws.append(headers)
    ws.append([
        'zhangsan', '123456', '张三', 'user',
        '', '客舱一部',
        '0123456789', ''
    ])

    ws2 = wb.create_sheet('说明')
    ws2.append(['字段', '说明', '是否必填'])
    ws2.append(['username', '用户名。可留空，若留空则自动使用 dingtalk_userid', '否'])
    ws2.append(['password', '密码。留空时：新建用户默认 123456，更新用户保持原密码', '否'])
    ws2.append(['display_name', '姓名', '否'])
    ws2.append(['role', '角色：user/manager/admin 或 中文角色名', '否'])
    ws2.append(['department_id', '所属部门ID（优先级高于 department_name）', '否'])
    ws2.append(['department_name', '所属部门名称（department_id 为空时生效）', '否'])
    ws2.append(['dingtalk_userid', '钉钉UserId。可用于匹配并更新已有用户', '否'])
    ws2.append(['dingtalk_unionid', '钉钉UnionId', '否'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'用户批量导入模板_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========== 备份管理 ==========

@admin_bp.route('/backups')
@admin_required
def backups():
    """备份管理页面"""
    try:
        manager = BackupManager()
        backups_list = manager.list_backups()
        stats = get_backup_statistics()
        return render_template('backup_management.html', title='备份管理 | ' + APP_TITLE,
                             backups=backups_list, stats=stats)
    except Exception as e:
        flash(f'加载备份列表失败: {e}', 'danger')
        return redirect(url_for('personnel.dashboard'))

@admin_bp.route('/backups/create', methods=['POST'])
@admin_required
def create_backup():
    """创建备份"""
    try:
        description = request.form.get('description', '').strip()
        backup_type = request.form.get('backup_type', 'full')
        
        manager = BackupManager()
        
        # Start background task
        task = BackupTaskManager.create_task(
            'backup', 
            description,
            manager.create_backup,
            backup_type=backup_type,
            description=description
        )
        
        return jsonify({'success': True, 'task_id': task.id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@admin_bp.route('/backups/restore', methods=['POST'])
@admin_required
def restore_backup():
    """恢复备份"""
    try:
        backup_name = request.form.get('backup_name')
        restore_database = request.form.get('restore_database') == 'true'
        restore_config = request.form.get('restore_config') == 'true'
        restore_uploads = request.form.get('restore_uploads') == 'true'
        
        if not backup_name:
            return jsonify({'success': False, 'error': '未指定备份文件'}), 400
            
        manager = BackupManager()
        
        # Start background task
        task = BackupTaskManager.create_task(
            'restore', 
            f"Restoring {backup_name}",
            manager.restore_backup,
            backup_name=backup_name,
            restore_database=restore_database,
            restore_config=restore_config,
            restore_uploads=restore_uploads
        )
        
        return jsonify({'success': True, 'task_id': task.id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@admin_bp.route('/backups/delete', methods=['POST'])
@admin_required
def delete_backup():
    """删除备份"""
    try:
        backup_name = request.form.get('backup_name')
        if not backup_name:
            return jsonify({'success': False, 'error': '未指定备份文件'}), 400
        manager = BackupManager()
        success = manager.delete_backup(backup_name)
        if success:
            flash(f'备份已删除: {backup_name}', 'success')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '备份文件不存在'}), 404
    except Exception as e:
        flash(f'备份删除失败: {e}', 'danger')
        return jsonify({'success': False, 'error': str(e)}), 500

@admin_bp.route('/backups/download/<backup_name>')
@admin_required
def download_backup(backup_name):
    """下载备份"""
    try:
        from utils.backup import BackupConfig
        import os
        backup_path = os.path.join(BackupConfig.BACKUP_DIR, backup_name)
        if not os.path.exists(backup_path):
            flash('备份文件不存在', 'danger')
            return redirect(url_for('admin.backups'))
        return send_file(backup_path, as_attachment=True, download_name=backup_name, mimetype='application/zip')
    except Exception as e:
        flash(f'备份下载失败: {e}', 'danger')
        return redirect(url_for('admin.backups'))


@admin_bp.route('/backups/task/<task_id>')
@admin_required
def get_backup_task(task_id):
    """获取备份任务状态"""
    task = BackupTaskManager.get_task(task_id)
    if not task:
        return jsonify({'success': False, 'error': 'Task not found'}), 404
        
    return jsonify({'success': True, 'task': task.to_dict()})


# ========== 导入日志审查 ==========

@admin_bp.route('/import-logs')
@admin_required
def import_logs():
    """导入日志审查页面"""
    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    module_filter = request.args.get('module', '').strip()
    user_filter = request.args.get('user', '').strip()
    tr = parse_time_range(request.args, ['day'], default_grain='day', default_range=None)
    start_date, end_date = tr['start_date'], tr['end_date']
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # 构建查询条件
    conditions = []
    params = []

    if module_filter:
        conditions.append("module = %s")
        params.append(module_filter)

    if user_filter:
        conditions.append("username LIKE %s")
        params.append(f"%{user_filter}%")

    date_conds, date_params = build_date_filter_sql('DATE(created_at)', start_date, end_date)
    conditions.extend(date_conds)
    params.extend(date_params)

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    # 获取总记录数
    cur.execute(f"SELECT COUNT(*) AS cnt FROM import_logs WHERE {where_clause}", params)
    total_count = cur.fetchone()['cnt']

    # 获取分页数据
    offset = (page - 1) * per_page
    cur.execute(f"""
        SELECT
            il.*,
            d.name as dept_name,
            CASE
                WHEN il.user_role = 'admin' THEN '系统管理员'
                WHEN il.user_role = 'manager' THEN '部门管理员'
                ELSE '普通用户'
            END as role_display,
            CASE il.module
                WHEN 'personnel' THEN '人员管理'
                WHEN 'performance' THEN '绩效管理'
                WHEN 'training' THEN '培训管理'
                WHEN 'safety' THEN '安全管理'
                ELSE il.module
            END as module_display
        FROM import_logs il
        LEFT JOIN departments d ON il.department_id = d.id
        WHERE {where_clause}
        ORDER BY il.created_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    logs = cur.fetchall()

    # 解析 JSON 详情
    logs_with_details = []
    for log in logs:
        log_dict = dict(log)
        if log_dict.get('import_details'):
            try:
                log_dict['details_parsed'] = json.loads(log_dict['import_details'])
            except Exception:
                log_dict['details_parsed'] = {}
        else:
            log_dict['details_parsed'] = {}
        logs_with_details.append(log_dict)

    # 计算分页信息
    total_pages = (total_count + per_page - 1) // per_page

    # 获取统计信息
    cur.execute("""
        SELECT
            module,
            COUNT(*) as count,
            SUM(total_rows) as total_rows,
            SUM(success_rows) as success_rows,
            SUM(failed_rows) as failed_rows,
            SUM(skipped_rows) as skipped_rows
        FROM import_logs
        GROUP BY module
        ORDER BY module
    """)
    stats_by_module = cur.fetchall()

    # 最近7天的导入趋势
    cur.execute("""
        SELECT
            DATE(created_at) as date,
            COUNT(*) as import_count,
            SUM(success_rows) as total_success
        FROM import_logs
        WHERE DATE(created_at) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
        GROUP BY DATE(created_at)
        ORDER BY date DESC
    """)
    recent_trend = cur.fetchall()

    return render_template(
        'admin_import_logs.html',
        title='导入日志审查 | ' + APP_TITLE,
        logs=logs_with_details,
        stats_by_module=stats_by_module,
        recent_trend=recent_trend,
        total_count=total_count,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        # 筛选参数
        module_filter=module_filter,
        user_filter=user_filter,
        start_date=start_date,
        end_date=end_date
    )


@admin_bp.route('/import-logs/<int:log_id>')
@admin_required
def import_log_detail(log_id):
    """查看导入日志详情"""
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            il.*,
            d.name as dept_name,
            CASE
                WHEN il.user_role = 'admin' THEN '系统管理员'
                WHEN il.user_role = 'manager' THEN '部门管理员'
                ELSE '普通用户'
            END as role_display,
            CASE il.module
                WHEN 'personnel' THEN '人员管理'
                WHEN 'performance' THEN '绩效管理'
                WHEN 'training' THEN '培训管理'
                WHEN 'safety' THEN '安全管理'
                ELSE il.module
            END as module_display
        FROM import_logs il
        LEFT JOIN departments d ON il.department_id = d.id
        WHERE il.id = %s
    """, (log_id,))

    log = cur.fetchone()

    if not log:
        flash('日志记录不存在', 'warning')
        return redirect(url_for('admin.import_logs'))

    log_dict = dict(log)
    if log_dict.get('import_details'):
        try:
            log_dict['details_parsed'] = json.loads(log_dict['import_details'])
        except Exception:
            log_dict['details_parsed'] = {}
    else:
        log_dict['details_parsed'] = {}

    return render_template(
        'admin_import_log_detail.html',
        title=f'导入日志详情 #{log_id} | ' + APP_TITLE,
        log=log_dict
    )


@admin_bp.route('/import-logs/export')
@admin_required
def export_import_logs():
    """导出导入日志为Excel"""
    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数（与列表页相同）
    module_filter = request.args.get('module', '').strip()
    user_filter = request.args.get('user', '').strip()
    tr = parse_time_range(request.args, ['day'], default_grain='day', default_range=None)
    start_date, end_date = tr['start_date'], tr['end_date']

    # 构建查询条件
    conditions = []
    params = []

    if module_filter:
        conditions.append("module = %s")
        params.append(module_filter)

    if user_filter:
        conditions.append("username LIKE %s")
        params.append(f"%{user_filter}%")

    date_conds, date_params = build_date_filter_sql('DATE(created_at)', start_date, end_date)
    conditions.extend(date_conds)
    params.extend(date_params)

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    # 查询数据
    cur.execute(f"""
        SELECT
            il.id,
            CASE il.module
                WHEN 'personnel' THEN '人员管理'
                WHEN 'performance' THEN '绩效管理'
                WHEN 'training' THEN '培训管理'
                WHEN 'safety' THEN '安全管理'
                ELSE il.module
            END as module_display,
            il.operation,
            il.username,
            CASE
                WHEN il.user_role = 'admin' THEN '系统管理员'
                WHEN il.user_role = 'manager' THEN '部门管理员'
                ELSE '普通用户'
            END as role_display,
            d.name as dept_name,
            il.file_name,
            il.total_rows,
            il.success_rows,
            il.failed_rows,
            il.skipped_rows,
            il.error_message,
            il.ip_address,
            il.created_at
        FROM import_logs il
        LEFT JOIN departments d ON il.department_id = d.id
        WHERE {where_clause}
        ORDER BY il.created_at DESC
    """, params)

    logs = cur.fetchall()

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "导入日志"

    # 写入表头
    headers = ['ID', '模块', '操作', '用户', '角色', '部门', '文件名',
               '总行数', '成功', '失败', '跳过', '错误信息', 'IP地址', '操作时间']
    ws.append(headers)

    # 写入数据
    for log in logs:
        ws.append([
            log['id'],
            log['module_display'],
            log['operation'],
            log['username'],
            log['role_display'],
            log['dept_name'] or '',
            log['file_name'] or '',
            log['total_rows'],
            log['success_rows'],
            log['failed_rows'],
            log['skipped_rows'],
            log['error_message'] or '',
            log['ip_address'] or '',
            log['created_at'],
        ])

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['N'].width = 20

    # 保存文件
    filename = f"导入日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
